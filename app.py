# -*- coding: utf-8 -*-
"""
APP DE RADON - VERSION STREAMLIT
=================================
Migración de la app de escritorio (Tkinter + OpenCV) a Streamlit.

Conserva TODA la lógica original:
  - Gestión de Centros y Detectores en SQLite
  - Marcado del punto del detector sobre el plano
  - Generación de informe PDF (idéntica, con logo y cabecera)
  - Ajustes (técnico/empresa)

Cambios respecto a la versión Windows:
  - La cámara ya no usa cv2.VideoCapture (ventana propia de escritorio):
    usa st.camera_input, que abre la cámara nativa del navegador
    (funciona igual en PC como en el móvil Android).
  - El punto sobre el plano se marca con un clic usando el componente
    streamlit-image-coordinates.
  - Al generar el PDF aparece un botón "Enviar por WhatsApp" que usa la
    Web Share API de Android para abrir el diálogo nativo de compartir
    con el PDF ya adjunto.
"""

import os
import io
import re
import json
import html
import base64
import sqlite3
import zipfile
from datetime import datetime, date, timedelta

import streamlit as st
import streamlit.components.v1 as components
from PIL import Image, ImageDraw, ImageOps

try:
    from streamlit_image_coordinates import streamlit_image_coordinates
    IMG_COORD_DISPONIBLE = True
except ImportError:
    IMG_COORD_DISPONIBLE = False

# Favicon personalizado (favicon.png debe estar en la misma carpeta que
# este script). Si por lo que sea no se encuentra, se usa el emoji de
# radiactividad como respaldo para que la app no falle al arrancar.
_carpeta_script = os.path.dirname(os.path.abspath(__file__))
_ruta_favicon = os.path.join(_carpeta_script, "favicon.png")
_icono_pagina = _ruta_favicon if os.path.exists(_ruta_favicon) else "☢️"

st.set_page_config(page_title="Detectores Rn", page_icon=_icono_pagina, layout="wide")

# Imagen de fondo de la app (fondo_app.jpg debe estar en la misma
# carpeta que este script). Se codifica en base64 para poder ponerla
# como fondo por CSS sin depender de servir el archivo como URL. Si no
# se encuentra, la app sigue funcionando sin fondo (solo el color
# oscuro de siempre).
_ruta_fondo = os.path.join(_carpeta_script, "fondo_app.jpg")
_fondo_css = ""
if os.path.exists(_ruta_fondo):
    with open(_ruta_fondo, "rb") as _f_fondo:
        _fondo_b64 = base64.b64encode(_f_fondo.read()).decode("utf-8")
    _fondo_css = f"""
    [data-testid="stAppViewContainer"] {{
        background-image:
            linear-gradient(rgba(8, 12, 18, 0.48), rgba(8, 12, 18, 0.55)),
            url("data:image/jpeg;base64,{_fondo_b64}");
        background-size: cover;
        background-position: center;
        background-attachment: fixed;
        background-repeat: no-repeat;
    }}
    [data-testid="stHeader"] {{
        background-color: rgba(0, 0, 0, 0) !important;
    }}
    """

# Ajustes visuales generales: texto en negrita, títulos más pequeños,
# texto pequeño (captions/etiquetas) más grande, botones destacados en
# naranja, y fondo gris oscuro en todos los paneles/tarjetas.
st.markdown(
    """
    <style>
    """ + _fondo_css + """
    html, body, [class*="css"] {
        font-weight: 700 !important;
    }

    /* Títulos más pequeños (pantalla de inicio y el resto de pantallas) */
    h1 { font-size: 1.55rem !important; }
    h2 { font-size: 1.25rem !important; }
    h3 { font-size: 1.1rem !important; }

    /* Texto pequeño (captions, ayudas, pies de foto...) más grande y legible */
    [data-testid="stCaptionContainer"], .stCaption, small {
        font-size: 1rem !important;
    }
    label, .stMarkdown p {
        font-size: 1.02rem !important;
    }

    /* Botones "primary" (acción principal: Abrir, Guardar...) en verde */
    button[kind="primary"], button[kind="primaryFormSubmit"] {
        background-color: #F5A623 !important;
        border-color: #F5A623 !important;
        color: #000000 !important;
    }
    button[kind="primary"]:hover, button[kind="primaryFormSubmit"]:hover {
        background-color: #d68f10 !important;
        border-color: #d68f10 !important;
        color: #000000 !important;
    }

    /* Botones "secondary" y "tertiary" (acciones neutras: Nuevo centro,
       Importar centro, Nuevo detector, Añadir plano, Generar
       documentos...) con estilo neutro de contorno, no relleno de
       color, para que solo la acción principal (amarilla) destaque. */
    button[kind="secondary"], button[kind="secondaryFormSubmit"],
    button[kind="tertiary"], button[kind="tertiaryFormSubmit"],
    .stDownloadButton button {
        background-color: #262626 !important;
        border: 1px solid #666666 !important;
        color: #ffffff !important;
    }
    button[kind="secondary"]:hover, button[kind="secondaryFormSubmit"]:hover,
    button[kind="tertiary"]:hover, button[kind="tertiaryFormSubmit"]:hover,
    .stDownloadButton button:hover {
        background-color: #333333 !important;
        border-color: #999999 !important;
        color: #ffffff !important;
    }

    /* Botón "Eliminar" (marcado con un div justo antes, ver código
       Python): con borde rojo y texto rojo, sin relleno, para que se
       vea claramente como una acción de peligro/secundaria. Se ancla
       al nivel exacto "stElementContainer" (sin ">", el marcador queda
       varios niveles por debajo) para no afectar por error a otros
       botones vecinos en la misma fila de columnas (p.ej. "Abrir"). */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-eliminar) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
        width: auto !important;
        min-width: 0 !important;
        display: inline-flex !important;
        opacity: 0.9;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-eliminar) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
        opacity: 1;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-eliminar) + div[data-testid="stElementContainer"] {
        display: flex !important;
        justify-content: flex-start !important;
    }

    /* Botón "Generar documentos": lo más pequeño posible, alineado a
       la izquierda (ya va en una columna estrecha). Fondo negro con
       borde y letra amarilla, igual que el resto de acciones
       principales de la app. Se ancla al nivel exacto
       "stElementContainer" (sin restringir a hijo directo, el marcador
       queda varios niveles por debajo) para no afectar a otros
       botones de columnas vecinas. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-generar) + div[data-testid="stElementContainer"] button {
        font-size: 0.75rem !important;
        padding: 0.15rem 0.6rem !important;
        min-height: 1.7rem !important;
        height: 1.7rem !important;
        line-height: 1 !important;
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-generar) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-generar) + div[data-testid="stElementContainer"] {
        display: flex !important;
        justify-content: flex-start !important;
    }

    /* Cabecera del bloque desplegable actualmente abierto: texto en
       amarillo (igual que el logo) para distinguirlo del resto. Se
       usa ":has(div...)" sin ">" (descendiente, no solo hijo directo)
       porque el marcador queda varios niveles por debajo del div que
       realmente tiene como hermano el botón. */
    div:has(div.marcador-bloque-activo) + div button {
        color: #F5A623 !important;
    }

    /* Botón "Importar centro": fondo oscuro con las letras en blanco
       (igual que el "+" de "Nuevo centro"), con borde amarillo. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-importar) + div[data-testid="stElementContainer"] button {
        background-color: #262626 !important;
        border: 1px solid #F5A623 !important;
        color: #ffffff !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-importar) + div[data-testid="stElementContainer"] button:hover {
        background-color: #333333 !important;
        border-color: #F5A623 !important;
        color: #ffffff !important;
    }

    /* Botón "Nuevo centro": mismo fondo oscuro de siempre, con borde
       amarillo (igual que "Importar centro"). */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-nuevo-centro) + div[data-testid="stElementContainer"] button {
        border: 1px solid #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-nuevo-centro) + div[data-testid="stElementContainer"] button:hover {
        border-color: #F5A623 !important;
    }

    /* Botones "Nuevo centro" / "Importar centro": cuando su formulario
       está abierto, el texto se pone en naranja (y vuelve a su color
       normal en cuanto se abre el otro, ya que solo uno lleva este
       marcador a la vez). Se restringe al nivel "stElementContainer"
       (con ">") porque, al estar en columnas, un selector más genérico
       también coincide con la COLUMNA vecina y pintaría el botón
       equivocado. */
    div[data-testid="stElementContainer"]:has(div.marcador-activo-naranja) + div[data-testid="stElementContainer"] button,
    div[data-testid="stElementContainer"]:has(div.marcador-activo-naranja) + div[data-testid="stElementContainer"] button:hover {
        color: #F5A623 !important;
    }

    /* Botones "Volver a Centros" / "Volver al centro": letra en azul
       claro. */
    /* Botones "Volver..." (Volver a Centros, Volver al centro,
       Volver): fondo azul claro y letra negra. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-azul-claro) + div[data-testid="stElementContainer"] button {
        background-color: #93C5FD !important;
        border: 1px solid #93C5FD !important;
        color: #000000 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-azul-claro) + div[data-testid="stElementContainer"] button:hover {
        background-color: #7CAEEB !important;
        border-color: #7CAEEB !important;
        color: #000000 !important;
    }

    /* Casilla de "Nº personas" en categorías profesionales: ancho
       máximo para que solo quepan 2 dígitos (en vez de ocupar todo el
       ancho disponible, sobre todo notorio en móvil). */
    div[data-testid="stElementContainer"]:has(div.marcador-num-personas) + div[data-testid="stElementContainer"] div[data-testid="stNumberInput"] > div:not([data-testid="stWidgetLabel"]) {
        max-width: 8rem !important;
    }

    /* Todos los botones dentro de "Imagen exterior" (Subir archivo,
       Activar/Tomar cámara, y el de capturar dentro del propio visor):
       fondo negro, borde y letra amarilla. Se restringe a la columna
       exacta ("stColumn") que contiene el marcador; un selector sin
       esa restricción coincide con TODA la página (cualquier ancestro
       que también contenga el marcador en su árbol) y pinta botones
       completamente ajenos. */
    div[data-testid="stColumn"]:has(div.marcador-imagen-exterior) button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stColumn"]:has(div.marcador-imagen-exterior) button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* Dentro del formulario de importar: el botón de subir archivo y
       el de confirmar "Importar", en naranja sobre fondo negro. */
    div:has(div.marcador-uploader-importar) + div button,
    div:has(div.marcador-btn-confirmar-importar) + div button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-uploader-importar) + div button:hover,
    div:has(div.marcador-btn-confirmar-importar) + div button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-crear-centro) + div button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-crear-centro) + div button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-guardar-ajustes) + div button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-guardar-ajustes) + div button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-guardar-centro) + div button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-guardar-centro) + div button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* Botón "➕ Añadir" (categorías profesionales): fondo negro, borde
       y letra amarilla. Está dentro de columnas, así que se restringe
       al nivel exacto "stElementContainer" para no afectar a otros
       botones cercanos (p.ej. "Eliminar seleccionadas"). */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-anadir-categoria) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-anadir-categoria) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* "Planos del centro": botones "Añadir plano", "Subir archivo" (del
       selector de imagen) y "Guardar plano", todos en fondo negro con
       borde y letra amarilla. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-plano-amarillo) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-plano-amarillo) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    /* "Añadir plano" iba con el texto muy pegado al borde: un poco
       más de espacio a los lados (afecta a los 3 botones de este
       mismo marcador: Añadir plano, Subir archivo y Guardar plano). */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-plano-amarillo) + div[data-testid="stElementContainer"] button {
        padding-left: 1rem !important;
        padding-right: 1rem !important;
    }

    /* "Detectores colocados": Nuevo detector y Abrir detector, en
       fondo negro con borde y letra amarilla. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-nuevo-detector) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
        padding-left: 1.1rem !important;
        padding-right: 1.1rem !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-nuevo-detector) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-abrir-detector) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-abrir-detector) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* "Retirada de detectores": Capturar fecha y hora / Guardar, en
       fondo negro con borde y letra amarilla. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-retirada-amarillo) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-retirada-amarillo) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* "Guardar detector" (al final de la pantalla del detector): fondo
       negro con borde y letra amarilla. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-guardar-detector) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-guardar-detector) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* Descargar PDF / Excel / fotos, y Seleccionar todas / Quitar
       selección: fondo negro con borde y letra amarilla. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-descarga-amarillo) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-descarga-amarillo) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* Título grande de la pantalla de inicio ("Detectores de Radón"),
       al doble de tamaño que un título normal. Usa una clase (en vez
       de estilo en línea) porque Streamlit elimina por completo el
       atributo style="" de cualquier HTML si contiene "!important". */
    p.titulo-home {
        color: #F5A623 !important;
        font-size: 1.55rem !important;
        font-weight: 800 !important;
        line-height: 1.2 !important;
        margin: 0.5rem 0 !important;
    }

    /* Título del nombre del centro (pantalla del centro), al doble de
       tamaño que un título normal. */
    p.titulo-centro {
        color: #F5A623 !important;
        font-size: 1.55rem !important;
        font-weight: 800 !important;
        line-height: 1.2 !important;
        margin: 0.5rem 0 !important;
    }

    /* Subtítulos de sección en amarillo (p.ej. "Centros registrados"),
       con el mismo truco de clase por la limitación de style="" con
       !important explicada arriba. */
    p.subtitulo-amarillo {
        color: #F5A623 !important;
        font-size: 1.1rem !important;
        font-weight: 700 !important;
        margin: 0.8rem 0 0.3rem 0 !important;
    }

    /* Botón "Abrir centro": fondo negro con letras naranjas (en vez
       del amarillo/negro normal de los botones "primary"). */
    div:has(div.marcador-btn-abrir-centro) + div button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-abrir-centro) + div button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* "Ventanas" (tarjetas, expanders, formularios) con fondo gris oscuro */
    div[data-testid="stVerticalBlockBorderWrapper"],
    div[data-testid="stExpander"],
    div[data-testid="stForm"] {
        background-color: #262626 !important;
        border-radius: 10px;
    }

    /* Títulos de cada ventana/panel en naranja */
    h1, h2, h3 {
        color: #f5f5f5 !important;
    }

    /* Cabecera de los desplegables tipo acordeón (expanders) en blanco,
       para que combine con las casillas (p.ej. "➕ Nuevo centro") */
    div[data-testid="stExpander"] summary {
        background-color: #ffffff !important;
        border-radius: 8px !important;
    }

    /* Ocultar el texto "Press Enter to apply" que Streamlit muestra
       bajo las casillas de texto mientras se escribe */
    [data-testid="InputInstructions"] {
        display: none !important;
    }

    /* Casillas de texto y desplegables: SOLO el recuadro donde se
       escribe/selecciona queda en blanco con letra negra, un 25% más
       grande (1rem -> 1.25rem) y con esquinas rectas (sin redondear).
       La etiqueta de encima (el texto que describe el campo, con
       testid "stWidgetLabel") se excluye a propósito para que quede
       sobre el fondo gris de la ventana.
       Se usa "*" para pintar TODO lo de dentro del recuadro (incluida
       la parte del desplegable de Streamlit, que anida varios niveles
       de <div> internos) y luego se fuerza la etiqueta a transparente
       para que no quede afectada. */
    div[data-testid="stTextInput"] > div:not([data-testid="stWidgetLabel"]),
    div[data-testid="stTextInput"] > div:not([data-testid="stWidgetLabel"]) *,
    div[data-testid="stTextArea"] > div:not([data-testid="stWidgetLabel"]),
    div[data-testid="stTextArea"] > div:not([data-testid="stWidgetLabel"]) *,
    div[data-testid="stNumberInput"] > div:not([data-testid="stWidgetLabel"]),
    div[data-testid="stNumberInput"] > div:not([data-testid="stWidgetLabel"]) *,
    div[data-testid="stDateInput"] > div:not([data-testid="stWidgetLabel"]),
    div[data-testid="stDateInput"] > div:not([data-testid="stWidgetLabel"]) *,
    div[data-testid="stSelectbox"] > div:not([data-testid="stWidgetLabel"]),
    div[data-testid="stSelectbox"] > div:not([data-testid="stWidgetLabel"]) * {
        background-color: #ffffff !important;
        color: #000000 !important;
        font-size: 1.1rem !important;
        font-weight: 700 !important;
        border-radius: 8px !important;
    }
    div[data-testid="stTextInput"] input,
    div[data-testid="stTextArea"] textarea,
    div[data-testid="stNumberInput"] input,
    div[data-testid="stDateInput"] input,
    div[data-testid="stSelectbox"] div[data-baseweb="select"] * {
        background-color: #ffffff !important;
        color: #000000 !important;
        font-size: 1.1rem !important;
        font-weight: 700 !important;
        border-radius: 8px !important;
    }
    /* La etiqueta de encima NO se toca: se fuerza a volver al color de
       texto normal de la app, sobre el fondo gris de la ventana. */
    div[data-testid="stWidgetLabel"],
    div[data-testid="stWidgetLabel"] * {
        background-color: transparent !important;
    }
    /* Menú desplegable del selectbox al abrirlo: se renderiza en un
       "popover" flotante fuera del recuadro, así que se cubre con
       TODOS los posibles selectores que puede usar Streamlit/BaseWeb
       (testid, data-baseweb de popover/menu/menu-item, y roles ARIA
       listbox/option) y con "*" para llegar a todas las capas
       internas de texto/fondo, sea cual sea la estructura exacta. */
    div[data-baseweb="popover"],
    div[data-baseweb="popover"] *,
    div[data-baseweb="menu"],
    div[data-baseweb="menu"] *,
    li[data-baseweb="menu-item"],
    li[data-baseweb="menu-item"] *,
    ul[data-testid="stSelectboxVirtualDropdown"],
    ul[data-testid="stSelectboxVirtualDropdown"] *,
    [role="listbox"],
    [role="listbox"] *,
    [role="option"],
    [role="option"] * {
        background-color: #ffffff !important;
        color: #000000 !important;
        border-radius: 8px !important;
    }
    div[data-baseweb="popover"] li,
    li[data-baseweb="menu-item"],
    ul[data-testid="stSelectboxVirtualDropdown"] li,
    [role="listbox"] li,
    [role="option"] {
        font-size: 1.1rem !important;
    }
    /* Opción resaltada al pasar el ratón/dedo: un gris muy claro para
       distinguirla, siempre con texto negro */
    div[data-baseweb="popover"] li:hover,
    li[data-baseweb="menu-item"]:hover,
    ul[data-testid="stSelectboxVirtualDropdown"] li:hover,
    [role="listbox"] li:hover,
    [role="option"]:hover,
    div[data-baseweb="popover"] li[aria-selected="true"],
    li[data-baseweb="menu-item"][aria-selected="true"],
    [role="listbox"] li[aria-selected="true"],
    [role="option"][aria-selected="true"] {
        background-color: #eeeeee !important;
        color: #000000 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# BASE DE DATOS  (idéntica a la app de escritorio)
# ============================================================

DB_NAME = "radon_data.db"


def get_data_dir():
    """Carpeta donde se guardan la BD y las imágenes.

    Se usa una carpeta junto al propio script para que funcione igual
    en local, en un servidor o en Streamlit Community Cloud. Si prefieres
    guardar los datos en el perfil del usuario como hacía la app de
    Windows, cambia la línea siguiente por:
        data_dir = os.path.join(os.path.expanduser("~"), "RadonApp")
    """
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "RadonApp_data")
    os.makedirs(data_dir, exist_ok=True)
    return data_dir


def get_db_path():
    return os.path.join(get_data_dir(), DB_NAME)


def init_db():
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS centros
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  nombre TEXT, zona TEXT, fecha_medicion TEXT, imagen_exterior_path TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS detectores
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  centro_id INTEGER,
                  planta TEXT, sala TEXT, fecha TEXT, detector_codigo TEXT,
                  plano_path TEXT, punto_x REAL, punto_y REAL,
                  foto_situacion_path TEXT, foto_detector_path TEXT,
                  fecha_creacion TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS settings
                 (id INTEGER PRIMARY KEY CHECK (id=1), tecnico TEXT)''')
    c.execute("INSERT OR IGNORE INTO settings (id, tecnico) VALUES (1, '')")
    # Planos del centro (sin límite de cantidad), cada uno con su nombre
    # (p.ej. "Planta baja", "Planta 1"...). El punto rojo de cada
    # detector se guarda en el propio detector (punto_x/punto_y),
    # apuntando a uno de estos planos mediante plano_centro_id.
    c.execute('''CREATE TABLE IF NOT EXISTS planos_centro
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  centro_id INTEGER, nombre TEXT, ruta TEXT, orden INTEGER)''')

    # Categorías profesionales expuestas en el centro (p.ej. "Enfermería"
    # -> 8 personas), con número de personas expuestas de cada una.
    c.execute('''CREATE TABLE IF NOT EXISTS categorias_centro
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  centro_id INTEGER, categoria TEXT, num_personas INTEGER)''')

    # --- Migraciones: añadir columnas nuevas sin perder los datos ya
    # guardados, por si la base de datos viene de una versión anterior. ---
    c.execute("PRAGMA table_info(centros)")
    columnas_centros = {fila[1] for fila in c.fetchall()}
    if "tecnico" not in columnas_centros:
        c.execute("ALTER TABLE centros ADD COLUMN tecnico TEXT DEFAULT ''")
    if "direccion" not in columnas_centros:
        c.execute("ALTER TABLE centros ADD COLUMN direccion TEXT DEFAULT ''")

    c.execute("PRAGMA table_info(detectores)")
    columnas_detectores = {fila[1] for fila in c.fetchall()}
    if "codigo_sala" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN codigo_sala TEXT DEFAULT ''")
    if "profesionales_sala" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN profesionales_sala TEXT DEFAULT ''")
    if "hora_colocacion" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN hora_colocacion TEXT DEFAULT ''")
    if "turno_trabajo" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN turno_trabajo TEXT DEFAULT ''")
    if "nivel" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN nivel TEXT DEFAULT ''")
    if "plano_centro_id" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN plano_centro_id INTEGER")
    if "fecha_retirada_real" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN fecha_retirada_real TEXT DEFAULT ''")
    if "hora_retirada_real" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN hora_retirada_real TEXT DEFAULT ''")

    # Datos de la empresa (globales, no cambian de un centro a otro):
    # nombre de la empresa y CIF, con sus valores por defecto.
    c.execute("PRAGMA table_info(settings)")
    columnas_settings = {fila[1] for fila in c.fetchall()}
    if "empresa" not in columnas_settings:
        c.execute(
            "ALTER TABLE settings ADD COLUMN empresa TEXT DEFAULT 'Área Sanitaria da Coruña e Cee'"
        )
    if "cif" not in columnas_settings:
        c.execute("ALTER TABLE settings ADD COLUMN cif TEXT DEFAULT 'Q151569009B'")

    # Migración de planos antiguos: los detectores que ya tenían un
    # plano propio (plano_path, de antes de este cambio) y todavía no
    # están vinculados a un plano del centro (plano_centro_id vacío) se
    # migran automáticamente: se crea un plano del centro con esa
    # misma imagen (agrupando detectores que ya compartían el mismo
    # archivo) y se les asigna. Así no se pierde ningún plano ni punto
    # ya marcado.
    c.execute('''SELECT id, centro_id, plano_path FROM detectores
                 WHERE plano_centro_id IS NULL
                 AND plano_path IS NOT NULL AND plano_path != ''
                 ORDER BY centro_id, id''')
    pendientes = c.fetchall()
    if pendientes:
        cache_por_ruta = {}  # (centro_id, ruta) -> nuevo plano_centro_id
        contador_por_centro = {}
        for det_id, centro_id, ruta in pendientes:
            clave = (centro_id, ruta)
            if clave not in cache_por_ruta:
                contador_por_centro[centro_id] = contador_por_centro.get(centro_id, 0) + 1
                nombre_auto = f"Planta {contador_por_centro[centro_id]}"
                c.execute(
                    "INSERT INTO planos_centro (centro_id, nombre, ruta, orden) VALUES (?,?,?,?)",
                    (centro_id, nombre_auto, ruta, contador_por_centro[centro_id] - 1),
                )
                cache_por_ruta[clave] = c.lastrowid
            c.execute(
                "UPDATE detectores SET plano_centro_id=? WHERE id=?",
                (cache_por_ruta[clave], det_id),
            )

    conn.commit()
    conn.close()


def fetch_planos_centro(centro_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT id, centro_id, nombre, ruta, orden FROM planos_centro WHERE centro_id=? ORDER BY orden, id",
               (centro_id,))
    rows = c.fetchall()
    conn.close()
    return rows


def get_plano_centro(plano_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT id, centro_id, nombre, ruta, orden FROM planos_centro WHERE id=?", (plano_id,))
    row = c.fetchone()
    conn.close()
    return row


def insert_plano_centro(centro_id, nombre, ruta, orden):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("INSERT INTO planos_centro (centro_id, nombre, ruta, orden) VALUES (?,?,?,?)",
              (centro_id, nombre, ruta, orden))
    rowid = c.lastrowid
    conn.commit()
    conn.close()
    return rowid


def delete_plano_centro(plano_id):
    """Borra un plano del centro. Los detectores que lo tuvieran
    asignado se quedan sin plano ni punto marcado (ya no tendría
    sentido, al desaparecer la imagen sobre la que estaba)."""
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE detectores SET plano_centro_id=NULL, punto_x=-1, punto_y=-1 WHERE plano_centro_id=?",
              (plano_id,))
    c.execute("DELETE FROM planos_centro WHERE id=?", (plano_id,))
    conn.commit()
    conn.close()


def fetch_categorias_centro(centro_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT id, centro_id, categoria, num_personas FROM categorias_centro WHERE centro_id=? ORDER BY id",
              (centro_id,))
    rows = c.fetchall()
    conn.close()
    return rows


def insert_categoria_centro(centro_id, categoria, num_personas):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("INSERT INTO categorias_centro (centro_id, categoria, num_personas) VALUES (?,?,?)",
              (centro_id, categoria, num_personas))
    rowid = c.lastrowid
    conn.commit()
    conn.close()
    return rowid


def delete_categoria_centro(categoria_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("DELETE FROM categorias_centro WHERE id=?", (categoria_id,))
    conn.commit()
    conn.close()


def crear_centro(nombre):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("INSERT INTO centros (nombre, zona, fecha_medicion, imagen_exterior_path, direccion) VALUES (?, '', ?, NULL, '')",
              (nombre, datetime.now().strftime("%d/%m/%Y")))
    rowid = c.lastrowid
    conn.commit()
    conn.close()
    return rowid


def fetch_centros():
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT id, nombre, zona, fecha_medicion, imagen_exterior_path FROM centros ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows


def get_centro(centro_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT id, nombre, zona, fecha_medicion, imagen_exterior_path, tecnico, direccion FROM centros WHERE id=?", (centro_id,))
    row = c.fetchone()
    conn.close()
    return row


def update_centro(centro_id, nombre, zona, fecha, imagen_path, direccion=""):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE centros SET nombre=?, zona=?, fecha_medicion=?, imagen_exterior_path=?, direccion=? WHERE id=?",
              (nombre, zona, fecha, imagen_path, direccion, centro_id))
    conn.commit()
    conn.close()


def set_tecnico_centro(centro_id, valor):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE centros SET tecnico=? WHERE id=?", (valor, centro_id))
    conn.commit()
    conn.close()


def delete_centro(centro_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("DELETE FROM detectores WHERE centro_id=?", (centro_id,))
    c.execute("DELETE FROM centros WHERE id=?", (centro_id,))
    conn.commit()
    conn.close()


def insert_detector(data):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute('''INSERT INTO detectores
                 (centro_id, planta, sala, fecha, detector_codigo, plano_path,
                  punto_x, punto_y, foto_situacion_path, foto_detector_path, fecha_creacion,
                  codigo_sala, profesionales_sala, hora_colocacion, turno_trabajo, nivel,
                  plano_centro_id, fecha_retirada_real, hora_retirada_real)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', data)
    rowid = c.lastrowid
    conn.commit()
    conn.close()
    return rowid


def update_detector(detector_id, data):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute('''UPDATE detectores SET
                 centro_id=?, planta=?, sala=?, fecha=?, detector_codigo=?, plano_path=?,
                 punto_x=?, punto_y=?, foto_situacion_path=?, foto_detector_path=?, fecha_creacion=?,
                 codigo_sala=?, profesionales_sala=?, hora_colocacion=?, turno_trabajo=?, nivel=?,
                 plano_centro_id=?, fecha_retirada_real=?, hora_retirada_real=?
                 WHERE id=?''', data + (detector_id,))
    conn.commit()
    conn.close()


def actualizar_retirada_detector(detector_id, fecha_retirada_real, hora_retirada_real):
    """Guarda solo la fecha/hora real de retirada de un detector, sin
    necesidad de tocar el resto de sus datos (se usa desde el bloque
    "Retirada de detectores")."""
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE detectores SET fecha_retirada_real=?, hora_retirada_real=? WHERE id=?",
              (fecha_retirada_real, hora_retirada_real, detector_id))
    conn.commit()
    conn.close()


def fetch_detectores(centro_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT * FROM detectores WHERE centro_id=? ORDER BY id ASC", (centro_id,))
    rows = c.fetchall()
    conn.close()
    return rows


def get_detector(detector_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT * FROM detectores WHERE id=?", (detector_id,))
    row = c.fetchone()
    conn.close()
    return row


def delete_detector(detector_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("DELETE FROM detectores WHERE id=?", (detector_id,))
    conn.commit()
    conn.close()


def get_tecnico():
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT tecnico FROM settings WHERE id=1")
    row = c.fetchone()
    conn.close()
    return row[0] if row else ""


def set_tecnico(valor):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE settings SET tecnico=? WHERE id=1", (valor,))
    conn.commit()
    conn.close()


def get_empresa():
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT empresa FROM settings WHERE id=1")
    row = c.fetchone()
    conn.close()
    return row[0] if row else ""


def set_empresa(valor):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE settings SET empresa=? WHERE id=1", (valor,))
    conn.commit()
    conn.close()


def get_cif():
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT cif FROM settings WHERE id=1")
    row = c.fetchone()
    conn.close()
    return row[0] if row else ""


def set_cif(valor):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE settings SET cif=? WHERE id=1", (valor,))
    conn.commit()
    conn.close()


# ============================================================
# UTILIDADES
# ============================================================

def _slug(texto):
    texto = (texto or "centro").lower()
    texto = re.sub(r"[^a-z0-9]+", "_", texto).strip("_")
    return texto or "centro"


def generar_plano_con_punto(plano_path, px, py, destino_path):
    """Copia la imagen de un plano dibujando encima, en rojo, el punto
    de un detector concreto (posición relativa 0-1). Se usa para poder
    descargar/enviar el plano YA con el punto marcado, en vez de la
    imagen del plano "en blanco". Devuelve True si se generó bien."""
    try:
        with Image.open(plano_path) as im:
            im = ImageOps.exif_transpose(im)
            im = im.convert("RGB")
            if px is not None and py is not None and px >= 0 and py >= 0:
                draw = ImageDraw.Draw(im)
                w, h = im.size
                cx, cy = px * w, py * h
                r = max(6, int(min(w, h) * 0.015))
                draw.ellipse([cx - r, cy - r, cx + r, cy + r],
                             fill=(220, 20, 20), outline=(120, 0, 0), width=2)
            im.save(destino_path, quality=90)
        return True
    except Exception:
        return False


def guardar_bytes_imagen(file_bytes, prefijo, ext=".jpg"):
    """Guarda bytes de una imagen (subida o capturada con la cámara) en la
    carpeta de datos y devuelve la ruta del archivo guardado."""
    data_dir = get_data_dir()
    nombre = f"{prefijo}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}{ext}"
    destino = os.path.join(data_dir, nombre)
    with open(destino, "wb") as f:
        f.write(file_bytes)
    return destino


def extension_de(uploaded_file, por_defecto=".jpg"):
    if uploaded_file is None:
        return por_defecto
    nombre = getattr(uploaded_file, "name", "") or ""
    ext = os.path.splitext(nombre)[1]
    return ext if ext else por_defecto


# ============================================================
# LOGOTIPO (para cabecera del informe PDF) - idéntico al original
# ============================================================

LOGO_PNG_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAABkAAAADeCAMAAABSbjycAAAAwFBMVEX9/v4Ce8QChMrK1+q0yOMvhsnl6/SIqtVrmc/a5PBSlM6Wtdr+/v51pNRJi8mku9ytwt7AzeW80eckfsU7kM6lvuCpvd6avOCaweFeoNN9sduewN7///+lvuFgj8qux+QDe7+AntGww9/V3u3Z"
    "5PAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADdAvdLAAAAMHRSTlP+///+///+/v7///+9//////7/////vf//////i73/vf//vb29AAAAAAAAAAAAAABMCEqxAABi"
    "NklEQVR42u2diZriOrKgs9AutbxgmKzTfWef93/GUYQkW7JlY0ggyURxv9snC4wRYMev2D/+619VqlSpUqXK1fJfH//6qFKlSpUqVa6Wf1WAVKlSpUqVCpAqVapUqVIBUqVKlSpVKkCqVKlSpUoFSJUqVapUqVIBUqVKlSpVKkCqVKlSpUoFSJUqVapUqQCpUqVKlSoV"
    "IFWqVKlSpUoFSJUqVapUqQCpUqVKlSoVIFWqVKlSpQKkSpUqVapUgFSpUqVKlSqrAOFVqjxP6l1XpcqvAcj/JT2tUuWJQtRV1yhX6mrscPei7xR+zWeraujLwq/4zqvcFSD/5yyaKlWeJkybqwjCCaVHcp16UIR8KyL3rhc+GyFV9X1VFDlevS+pch+A/Dcj/1Sp8jw5"
    "CHrNra46Ywy9ljmt+VahZBcUFHXHdlXzfdn+8N9jBfE3AOS/f1aAVHmmsOsAQoTW0mnkaxQKNVZ/p7j1qgd9tiolA+Sv+9JtBUgFSJUKkLmStYcDs+11ABH68J2y9yPiZxNdBcgXAXI07PBH9xUg3+HCqgCp8voAob8YINUCuQdADhUg3wOQtgKkSgVIBUi1QKpUgFSp"
    "AKkAebsYyBcAUtOoK0CqVIBUgFSA3IICXnPgfhVAWNNI3QxVzVaAfBdAmNaasQqQdwAIV6Sv3/5vAkgjpRCSVYJUgHwbQCwkAbMKkDcACNSb1m//6QAZGGNRG7Dp73soFyla2gndVD1bAfItAGFMG2qsrAB5F4BUF9azATKwRjYeG6zRUt5P22ObC66oqHGZCpDvAog9"
    "OoVkdQXIGwBE1T4yTwcIQ2g4gqDHqZHW6sb9fPcJgFgKvydpBauKtgLk+QBx9oc1ClqM2K84sSpAfgpAKj+eDJABgxRChjCFs0UE/PM+ce9G+IJj5QgiHxNJH9DnVkMsFSBFcfg4gk7hhBpZAfJ+FsjK2IGNiQSbnaK5b+vZFzpmLl63et7w18ZohNK5ymtdHvZkgDAp"
    "TOvUu2ZBH8O/OxP//TXlrk1oZ8MdQTR7gBXizCfsCFsNnAqQkjh+hDtKHcULA+SH7Zv5Y5vW3wwQ0O/ZyjgZJdX4ycPzp6Dz86oZ487fGiutaDsy69TlTqnUfC0qZ1t8Qvk9zVLU2sdQ5Y8xM7g4+XIr6CsBAgZH2xphxyi3s0iAKGAxfFUps8bddvH2JdSd8s563q3e"
    "CmPwA3x9uVV+pQXS4l0JFoh9SYAopxuOR3qEe3+mDrYnk2wMLNmYpoEnTf7km2+2HIWCquzoW9yTRzmMbgUIdG3usl8KpwCA3u2zuTWKdv7hIHx2CrK2rq7r4KV+DM6lt+7a5C05id2FVe8sGD4ChOIbBllZg4ofg8w+Rt4lerGGxwNEio7QDBYQUXe3gLtlv6qRG5l2"
    "zHPfirivkgfWtXDfucvj7JZb/VgVIMsaEGvPypkf/7H69IIAgcbl1koQa/Mm99DS/DNtKf/p/tWNGsOpkk8UaHze0aNKb7XO/D2qsgp0Z/WfBf4ctSA/fhbb2H/+zc6ML4L1QodiXPBLAQS0llNe6SNCCAOqtxPCPcOnh01LSwABLZUcmC+rF9BqGfxEcMIuM2mobbRN"
    "9V1vJHRmjgepVrb+myQwHmAECO2MtHEt4YDWrUGmp4ePAcf01Ahp04+R/gLKrUEK8jyADEwLSvtutntnoPlp+1WCsFkPbLis72koAD6ot9g4Xte6OrIqQApVhNZtMf6KL9USPgogaBjpBuN4zOm+83HUZZz8tUsxNLrkjiZ5WOCEFTV+Xe45UyIIvp8wPjAJZ4gKTlFh"
    "y/KZfHK8z+JyIT9B0IdkPd0IEN63wpG4TV5GrbSGgJXn9LRTa+GpXkr8whZGFnx1s1MkX547fTAj3DfR0QwgrWwa3aVncvtyHaHhPpORwaEPNACA4FsT+EbdWcm0EHxl0xiVfQy3Vwbjj6YfgziYTPBza7BNIyl/GkCAH4QsSHGAzHkCSP2KQkYH1nL7c7cc4QHspOnL"
    "AwOnVptUgJTSeOnRqYVXrAPxyWFsyhjTScyGGH0adXWQ06gwOLVan05Bk8P0DDtOveKQtezOVHxD6Z4hcbuu7TkCxEqdv9Xp5M560gmI+NFAsDRLcROPGBx4I0BU60BqZ5rX+p+Ow5ciokukB88LL57C8VVkp0i/vMDeZZxbgYWTqPaIAS3iIwuAjJ+1lTaPbLg1uNea"
    "5CrqhTXjx5DjvIAcIMQtz72uU08DSGPdByYFnc4Qd19T9k38vnIb04H6Lr4m4FN+frghDlWdV4BE7QY6ELWg+eu2lRr/dbrNDnkMQJxCtkH9o5zACom6Sx1tgEoqNq6DU6fJp2e9QRDf2G1F4V9LDak6h53wWZyFA0WWatx5z4Zo+W9Km8QmMgiZ9BDGQpDpFQBChGzd"
    "lSISvYAA4REvo/7v5cqPRMBJZRpRogucoF35rKC83UVmepVcs+6im8LAawDhpJu5neDINv8YvXvnuDtopRBlgLiPera6/dIcrqsAAj4mTgqRCea+XwXlf7erelZ2JGI65T3cWI1sF2kQXa02qQAZXVfSejd+dP6giNvKQR4DEFDIkGgcwqMt6vAUIAxyIsujdREgenwW"
    "7Qcd91Tu1p3YkL+jZvoc9FgKEAj5zsIf+E0xOX5wcHlprKvx6+2BOSf4990nB94GEE6spKSz6UCvBCCg0cUFgDhtCKcoP7sFEAqmmDOApqiveztHECna/lqAuI/R+TXwEkA6dzHTAkA4laLP1/BggAwSAkIlT9UBPGng1mNf0CrluWwQNbyDEVLik9sh1a6NFSCwLYbg"
    "uekoPR6PqJetMP4f3meE8u0AcYhwq7CfwUcEDl5h7WcEiDsXmBHHTBJ1AQCJz1LjbZkQdeSggJldLuNoTywaMTlAlCLp+1D3TYGJoyfdBvBhJzBzgtESiMIkvTdBbgOIotKCR0VMkes5QKLmXQOI+9Jl7ygi2pL5BgAhKxeog4Byl9qklQAgKOEluwHCiZTkI19DChCw"
    "bEoAUdMangMQ5gw15QBS6nTYyA6jCrcCZFgaCEkgDoPpwxeV1pJPXLU1l/etAeKUmYUcdGdpfNIQl/SVX8fpH7i9htRvuR8ijwEItQetP0cPkLszjlNeJlog0AmonFOLALHjsxCydMpcBh+NwmcXa4aHx1ShHCAfPE/kJcCLgx7zYCDNCGIwaaYYhtTBDrp3+6mbAAIx"
    "ZKeJnXZJ7IQcICKzQHjhFB0cAsGOgqkBcY7iyzCEbhRmKHQZQFpK43tuAyTNI+7gY6g2XUMGkLIFMq2hnALwCIC4D+QMgpKZ0ehWuevZNHc0ENLr48tlhU3JbuadrGH0NwYIbsGPCjfT6eWRxTs5PAnHYD7RNwLE3evyIGeRCj4iAgCi11NqPEBUGgI02n2vIXKO9ski"
    "TqzOGphUBMhSJ54Oh38mB1YHt2yCuxDF8V4t8gIAUZ4c6thKUQRIl8VASkH0ECYHWBY+Eu+MbCCYvQztui2xexeIh5sUIO5YIkI+7l6ABEeZ+xh2Mkx2xEDAs2MgObhd/U3vDJAG4nUpQIbJFAGAwFputRMacSabF8gXs26HpmTgOGuq+rDeFiCQBmuOmMvKSxVuESPc"
    "V9Ipd4tmOUXPB0irD9aQdQcXAORjCyAiK/74674jTcfwCtNmtmhFRTOF4bcB4vQRm2DjT8j0IjdYUeTKnU2QmwASfFccPVnzLCyOWVimI2ka71jbzfNTeC8SL8ZvrS8rybOX4+vcW4/fGAKk544bXlXtBUj0XWUfY5aF1RYAQoL/THVafMGleBVAwNThI0CGpKmUAwg8"
    "191aCpKmAnCllsX66mtZtw4gpUuW1nr09wVIbHuV6CE1XXAfanHb8/1Neh8DEGcy2O5eAAGPGERFVNT/J2ZndDgiVWKe8CZAIAk1802Bh2xBJAxp4lnuGwW5CSBOYaNVkWl/rAPp4ac/djatA2mkY0GQSQnboJAhHF/Yg/hGJk6EafvUnRgjHVSKKU0ODR5OjY9q7waI"
    "tN20hggQqAPpfaqF1MU6kLgGd2GIXj0DIBr9dQAQFjvyOtFNw4bBl6Pw7sYo+iCz+hZCJyFjrjoWwN+o75kuA8RWgLwrQOZ1q6CC6WfXxRD1sWtbeswVHT9+2l2BkMcB5G4WCCQhOYCcVby99Ny15G71U/LFbwJEQcQ8S/HRK9FyjK3fOQpyC0B8HPtjpv0pJkJB/xEs"
    "4R4VMtT4yViFeR7DCxCFjwnBxUxeiLD4vgFJFb57XfgXyYLd6DEjnQBFy3cCBNfgOwUka+ix/5ZTnxC6m7iSAIS3MYe4l6IjTwQIcUqXwbdMaU9pBytsQgjjCwDpUoDQVggbed/SsZaT3txwq1ogFSBzB5bx9x1Xowo1vnbbb0vxH2PZWyz6xUza7wLIGZJ416IcXwMI"
    "xstneViqA1U/3phbAME8YJbYFarF05V0qgEsGfLdAFHdmEFghInVdAAQp9sgf9smKVS9jDv6xC5N8qGUsab4i6M7pXcKLRYohq8gqHrwI7U5QOASaxXfCRCupjW008dwAMG2f/gxEr/iBBBYg8p8WU8CCMC2aTC3m3NsVtwK3fjfgtPbAHIYnJWlRp8zga4wjc+uhEZb"
    "3WSE3FoTUo6BVIC8L0C0Hjs8eZsDfSvwxL/hCfTvwPv+9Q7pY2yCA66abwuiY56VuhNAPkjiwor/4jOr4jS93QZAoMLRvXnHU40O8Cnnkmm2zsFnAQTsOTO678b4MxTYS695RYKEYhYWhqHVGMneYKJ3n4hRj4noEIFTCJUBBJpgwVQ9sg8gpI2Z0yrNGkMOwsdwps9U"
    "aj4BRBHRhJ+LTB/jOQCB/jGQnIfN5iE7pZXB2XkrQHAQSDe6qyBxd7Q1mG9hFfZ/5Nauv40p1L9CFlatRX9PgPg9MB97Baqj0SE08hcBIsO//LPUfHpjhNA9TU4elsaLfqJiOykASHNzDATUlLMZ0m0WfCOpot8AiPrroGCTmofVykR0A0rwbn0zQKAPlA0VmRCnmADi"
    "w+XOajCTzijWgYAysp0/hdOJW6llcJ010aMHTZ9EfGvRhNeNAMEICVX7AIJ9FKePMdatWOmD9+AimgzDESCQ+xvX0Lk13N5RcT9AovUMuWcSDQafaI6eJW8kcHq7iwlcjJM/UKG3apgQMj1hbsqcKt6ltQ7kbQHCAjagQxTePdC0owCQkFhEsFMqmCqoq74HIFDu51S+"
    "LtcWBIDw/VlY8BE1newNx4fUtQSAkWMIfRMgBEo+0sA4MdgbpfgdYIGI/maAQGuwRgcvuZRNXM+UxuvsBNvG6HIZIND9MHjaoZvh5kdyujGUKnBQYXJ66xCiGAECUHC/w06ApGuQ4+/fi1AdyRWGZ9QcIKBfpzUweXNSw16AMNbETpyqN9gqmECj+x77C4dPyvvb49yM"
    "eVBEFx1tE4RI247WCb1khPiZUbNqR5Y0Sk5uyNmJwksrVd4AIHA9QQNCZ4z4pJciQEJ9NkZ+bXBRfxdAsIOhrx8nywFNsZCQq+I0unkdCMeiviRsrqAuJNlG8yP0UUyWtgoQCMDPgjOYv7XW5Bytk7P6XoC0mFUV+tUIHWtgZr2w2o1WJoCE9BRNu/mRiAwZVwoDIlMD"
    "mHH3HQECJX7u7cQOgPg1mLgGCdGTAJC451atNYs0XtUGF5d/obbdrQTZCZDBt7QZL7VALliFgHD/2CDH2JuTbYeAkDGm2Y45UuEJPgY72SY/JI5sZ/Mk5EUvrPnk9aHxL63Vhb8eICfq/VGC7QeIRgcBhIe/sRuv9n3cPxdNbX0vLHeXpimMag0givhQRNItNmBgPOI/"
    "zjhLI8OrAIHk5llYPCTx8rIhJTbqSZ4EEKyOGCs7KChrNQMIBzXXbwDEKfIuO8Xmbz4CxJEheV1nbRv8/2MYxV18sqVjm8MNgKi2EbTPPsYcINCzPsJ9BEj+8aFi8taevLsAAjm7YyDCcwI7R8KGHRPUVJpA5RTwcL35ESMeY5Np7gejT0+0/WidmNVZhXh3IVVlbqcs"
    "uvFiYGvIjghxJ5jwfmu6sJ+Yu8+MGaaDh/UTLeXCUPetlzbjN3370n4TQFo7AmTFheUjwb5Ph89Vp98HEBzDgS2nJJTQ8zlAcB5WMvUjHXAXW5n4wOXxrz2xvFADsrIS1xIxp7xeYw0gWF2el3yAg6hk878MQDju71U01qaawQQgH+ib56sAcdeEgBYFIRR8vtCVkEgZ"
    "ig6hviSEkKG0yL1uBhBkDDZovwAQ8KGkH6ONu/wEINh1vJ0DRCRrgJ20uDUtbg9AnGbNjWaOA2eFN0DabJ/jA+BXah3vOAqWziJmPoxgGJ9wvyzbWCoOhHEGTHoQDG9Pw48eTzlhgMl+CIu9zQgZgEEoOyZh4YDdePCweE6sCM532+DTtIbCa9G+uvzrbC3tl7iwuhQg"
    "PHhwCkF0n6ACEeaQ4wTJrd83kRAjjtpPAsl3vKGde9ZhPZ3NMevGa303XjXLutKfKsNJ2vukDBCn0LAERH0sAHLkq6GcbwYIxFhpNhYqJDClAPGJUGu1EsrpuTZtSbXIhuXJjHPuLpuYj5EVHYK+tGQOkBZuvMsAQcSpbA2G+iD6BBBsehUqXkLRIYc1JOkSoM0fBxCn"
    "kUy/nHQMWQrGtN08J+SGZNsGCxKHlAF89DLJ1I81tZ9uVs4U++mjAZMdxPQYSXHP9rNgPI5VVFN+3k07b+bnw7e+OubSp9aOWF3bdTBQni34AafJpfXinZ2rDPFrKL4SX2uRIWzH0uB1Rkj2KwHSTilV1m8MqV3NwkJtOrqwmm8ESBgqFed5qOsAckgnQGk5K2t3h2gW"
    "M4+gZRGz2bTXMkDmPUwCQE6vDBDcc6crdvaDNzZSgGAbKRorxmdRAujimw/DXXQlhAnkYduNESdBPsbeh+nabUNnAOHoY7kIEPwYdJZY3PqPMQEEfW3+XwEgcIpsmi5AjD4KILAzL6XAqryqJn3m2jmCTl0liiok7qrJJ5alY6kA1SLrtEnaFbuDhim+ghEW08VR8zgt"
    "N/EFpfOsOAwzbm5SjWO1kVNhF7HZjs5Lm2t0HD2/Ib0vMS1imiUO1pJA8SckXmxxQWdLG34jQGJKlTtOnn1A3efnZgAJ74sZr/7v7wyij0ZIZ7wjy3RJq9sxBnL24n7o46wb7yGW0aMFs6izhz5V4YtWOAkkw0IZIPxY6KJ1wYX17QBR3cxnA1UhuKXIAYKQQIUs05gB"
    "bEOh1WKbn6KdlVNwevbFCT7RyH8d+ZhCzwccCpgCxBclpQDpigBZfoywhtSFlRAouLCgjaRRM4/ejWlxlwGy3mh9O9Z3zb6962mbBMaZ2wHHqw9CFaPThUU17+7voXgmmjU6DnqZhbCA9ikL4x7eBwUGPFMeAlOX9f92tQn8zBcOlomtPBvSxeQFlyTguysTpFjxkr7U"
    "OyA3wzwbS/stAIlkAMeOxZQgZU5sBpCgQrGJSJi1p74dIH6LhpM1Tnr6sUvt3OdZWIcw1oSd/vl3YbRsMpgQ85vt38sA8SXn88TiyzGQz28FiJlHvHnr1TVo8aQeMtbY9d5dHwSCHdAKJHfboafoYx5l9b5gK6N/gyymTCmDv6K7Zk1SMy6cupqC6KYIECg5pLPmAf45"
    "kgOkCx8DZhASf8L8t5lq4+8OkOk6mOUF+gdm/xwNtivUDs4ZwXECadLX6McC629UlUPjr1Xe28IWms1GS6LyQ38+TnRA6UKCSvTpxHBF/gWumTiXAJI4HullgJBsoekHudzqgXvjbAnq5vKVwFX6nZYBMoUKfyVADt7Bwo9CRgdMeFEKkKBNwZyMDiPS2W8HyNSemk35"
    "l1gHoldTSWMMBE0XJg0tdicXzjTxzj20KzLGFAGijqLUGxG/sBWAqO8PonPazj8/MeHn7ZJfA+4xfwWgFzh6hJ1VAebFzPnidMbsR4eXjzJmANHZW/PeL4ZQmg64bU32kszhRNXKx8A1+HMRtfwYyr+3gg85Mz5vHTR8CSBDE1puYdeSzGPFvS03c2mp0ONlvwuogaYs"
    "0FQ51VQ46idN3G1iSYjvtgwX4VIzzVyxCuIcMazOi4LRdnfUMKt6cjuw4WsAuWSBHL4GkDHMM1wPkDFS9cYAYboLibwwFSPcXiIHSPgZ0KOvx7+/qxfW4gcUOo2D7CkktAT7Ca91w0XnEgt7Vc1m7UZKAPEOLLvIAcVCwtU6EDhP+71ZWHzpeOelZ8aHc/2xeYqPxa46"
    "3WoXXhf33VndTjJiYLlxXz3XB9/8GKrwTqtrvwdA3Pa9g1PHoDm2G5kmktCxF7F7Hg5wG3zU+mD77dUnwcQhM4UGUe3RCOnHssJGdKsAaeZOyDa07eLbahiC/rafoVw8GCB/vgwQTOGwcxuk2WmLqo1cBxnz338vQA4+XMzJWcYtNIfMVQ8QX6Z9Mn6z0jJwYB39zlyz"
    "lwAIRmN1OmN2Ry8sq6BrFaRfla8RzDCDpDTHycXYpxJA1LIEJASlL1Wid9/cyqTKfeQSQJwqgX06Fn5IDCFMWl3IkE/AcbJuPAIMOd/2f1/5YLh/IHCySEWiWVmh220PMZl6J0CkvGybOSUiGpFffar/CQDxhgS7CSA4aqZ5Y4D4PonqKGKROZgd+nCaAAJjNnGrjxF0"
    "//f3tXNf7nyOsLB4pewDiI+OM7mY9JQc5G46QrWex+EKAMEvTxeCbpu9sMwr9MKq8hSAYJtd6EyFWbbMh6HPJOQzu3/2fk9h3OPeRBhA7WNtzj7FM+YawD07zP1ndkqtJS2kRsMjqzGQBUBgINgOdQo2iJi7sH4EQNAHk782B8gicJWpC7OSo/sOAAFTg+CFqk/sBKn5"
    "sOsBgEA7EPIJnh7/oPuWYPar+5t04vsGShWu20+3MHEFQKAXFtb9nexn8T1DBTn1fRTVBYDARX5gsvBR1dEwdirWFvDX6MZb5TkAgcqLsbCDwayVVvjWJZCEjLWXGOUWput855FQx6F2RkGmuB8YLQsDyBo6RUKoMSYMEStmYS2D6PMshVV3jsjr1O8QRL8XQPL5Wv1i"
    "TB6Em4YyQNTipflrV7Pl3gIg2J2H0pARa92fDg+Nf7jDxiYnAQ/iBBBUq5D7zV4JIDQZ2LG7nbvyTbCKjl1HSIbVhjYZRbgKEOzMXt7rqBZCSxvzQI4fFSBvABDd9gTHGHvdBHFtSE0TQYv3PntAwD2nRo8IjpeC4/aF0KeM0YJVkVQPhji970ZsdNEdlu2awMm2s0kY"
    "MX6wcEIU/SIAwarUKNC8rc37ec/tiCwTTMjppfDaWfaF6sphkLcASKy6O7Hxbzb+6R884YPJsfvw8SyA4Gb+aoCADQL5WyUbBAqmmbVQqbggzAIgeB69MkiJQpfK0lMkzC6sAHkHgGCr+TZqGazCwHkgPpcWM8051HSjoWJCH8UBS+R2htETgBxLNWt+opTiqT+Gz4vM"
    "k5zjJBPOmSlyp6ry3WCSBDNzUyHhQwAiQsUKSuwew3Mn1lBeA8SNRvFVMLTPXlsuNH8TgEBZxCmWRgQZ/5weOxzSB18KIJJd7cIK26MVGwSMGo9SMY9szAESpuDScv7Oykz0j9eZiV7lOQChk7LGiAhYwE1yzQNfmhZHaEVVM8DztLNfdmGFvn6zPgJum7Q2y3yqJ8cW"
    "YdLuvFChsc2slQl7GYA0LJdGy7wJXps6opI14MfIX5rMVgkui+aNARLAcMhhUfrrcA0+nubCAj+RuR4gXouXv1OIkGDF4SICPgeId4WtBeugkITpRRTE9xO+7zSQCpDXBkjSZATS/5RTjY3MdsgwfR4ml4yqZpBttxMgEERXcVDVsJoLlrTcwulexTp00K2QQwLip4SJ"
    "3ZMW3D5fxJf2rdA729Y+ASBLWGJtS9oKLnVEZQDRpThRNn/evDVA8u61F0Sz7wWImu31UW2ezBV1IONAKfK5YiD4tpHgvFucaQYQvjU0aiRFu+ggZUtcqQBZ/F7B2fLzAdJO2gmy7xSmf0qbRC5gU4saM9FiVnT7AAJup45st8AdYOiWb4QYCrCLTaCCh8dAxSj2iroG"
    "IMbnKMcOgug2Gl4UIFNLlxjr0fsBAjRWqfXyzi4s3ZK9onaWgDwMIIp+0qNaqOibABIUeSGZl/i5KHIZ5c4BAg5inbcOmkVTwL8G0fJ8yVIf9HW//xsChEPzLOhTS3rCfzpAMgUCvirspj5FLnqJDpW8dBDGd+zNY8LeiZfmDPpBIR40xQPBseNbkJtx0NblLvdjvhI0"
    "NJ5eOo6pHF4TIGDjJXYEScPoFwDic9WSW7pk9r1LEJ1dlU1K9ek7AUKsBnUca4l9AcusEh2CC3xeM10ACA48h45Ux0JzXV3+vnOAqFAdQ9TKG/oh61glMi7Z+1Rvbbv0RgDx6qg70o7+bIDoNi8zgM68nCeTVsCdildRHnaGLfJegOAgp0sDNLCXLo57Kh8IvdqXTWjb"
    "GF9ZKYYYC/g5KXSwpSuh+lcAyJ90G5flFFwGSJqrBnHYoQJkH0DsNwPkhDNxnBwhpRgMotOUfV6aSEi7sSXSfCa6I4jVJ7uwQXwJJSvZJn+xW6Ma42tYS2Nmb9gl3R2Nt0HOsGJcstAw9PDz7rvqX2iBYAb+0Wmhn+3C0qYT+Vw/d8th848k9B0DIyzPndrfzAS7rV+e"
    "chSlqLptIbHD/Q6jwb3S1y024CeljBTSXjkM44kAScNQ2XiUiwBxv+o5sV5Kp68AKQJEfjdAMMrnwzZhJMh01TuLBD6RzsX+JyaVAECyIg5vgyzLyMlKqyqwQJoIENVCdczy/XTWH9j4RYZAEy758AB+/EKAEO6LFhT/2QBxFkeeWwvDpdznMqkFInoYq6XZLPlWyD9P"
    "kyHN302u+QkgxftZ9ePkqmLN07VjQZ4IkDQSyfuEFDsA0iTR02K79veJgZjj7iDI98dAwiQpX5HCThqaQmT5EM1CnY8mBtSMzPqQQOKIXr65MlqXYilQ76VtuO+VkbosybXlCILjEwNl/JIpUR8/HyDQ6xDH2ZYf/urZlW/X+BB8jEvnTwHIPDTB0IqeXFi8t7KbjY8N"
    "ANHP4wcrZ4OMAFE3AWR98uELAKSZUpSzUPgOgMyzhGsW1s/IwiJdtDz8vt6kNRiY27IQa8bW39bOOv3AgHVrF03XlbGyNJ8O0lxsNNmIsbIo1iQFijxAz+dDY5H/8RFK8XqAcB4HpxQVKb8AAUV6aPBOZzW9MOsLHs61iZreKlvBvF8unxR88tyCIhuA4iqbBrPidemx"
    "yfy0yPxVY1ditTJZ5iqAOFzMQw6Nu0aMSCcOSfdvncebIWQhm6cBZN7D5F4AgSLJVwVI0o8ua7yyByAbb/1mdSDsdDrpXXI6fXsdiMJuPmGokclzsmC/D3rhnIhXcTzeC7Q9zj28tKOLPCp36ZqOlN99PFrRc1kozeesU/oJfVYtNkIw9DE+mevbuRNC+yTDbjmwe8tQ"
    "Un6uJ3y77iwqVc04PhrHEE6mYU/p8p3AQ5X1ToqzMTiOzVCpjsr7D/V0LfUNftAscbD8wWGRMNSExrYe+cvUR+HBmwHSyPn8bCxkyyulpWzmIQzmzGn2NIA0KxP8vgyQy2Ohvgsgw1cAotO31u8MkAPbL4fvLyTkWOGEt3nBf6KWkuZhLfaRuMssjJRwj675PqZRFGpF"
    "OF9ZsR/K8xCH0tUAIb2Bvj44WM7A9L7sq0Fby6yFxyCHCPoJESw2M9EMjA/Dp21N0teCCGtwjh08Ok7BVb1JdRZ3C/GXDBht6cBAOquipkLatnzBwKTeNsmfWOg0bFvYdmHprT8zZqC2Ux5E6I0OD3bjg+pGgPyBoa9DYcOfVKIbWch2HW4qxLsdIK16BEA+Xhcgf4ap"
    "Tcv1ALEVIA+V51Si/6Sw8EPPfvVEQojnCK8enaKUIpsliM2Z1lSKx0tADkeAhMGznQlt+3AA3jQqkujGq2JsajdN6rbpvBXVxhHrMPIyrdecjc4lbbPmb/lQprFt0jx12UemG2nJybH1TSlIK7RMABIn7zpO3QEgpfSoPJNbrTQFfyI//rjvTZUu0wqQ6sKqAPn9cr0F"
    "4vbxLZbM+L24ScNJ7upxmn6tSYvTd2ZqRsejW8q9SLTRGFG+U3kIF2koV4XBAeFdo29jFSApImYAwVLMZqX+3wFEkKQwpxDSFenSvXMMACLIrJAHAGI2TrUbIGsekDR414nmibAo5qWG4uq5jZ4ChP8ugAzDV2IgNYheAfLuAKFJqF+kWlq1UJe2MrURY775djT0B2+S"
    "h7ERXzBhiJYxQxTO7H9s0A8LgPi9v1OnTTMaMDlA3NtAJZw9l3N+zOakYFATWYtMDwZkxSKX25lL7Z6v83qAYIN1lSQOAEEkG74XIMF/qO4GEA/d1wWIMwP7W7OwGvGWabyQ2OF2lheSqZjGbKKLR+FhumEVID8TIEmiGWrWLKpr6GLE3PhKKztVJFL6MJTKhTAIAGS8"
    "I2WwMzYBorFttioBBOIsOFtcXQ8Q0NMFswoBwpcPinsCZAposEaH6lcsl+TRozh24t0X+xh8XHIYFhvr8MQtAFlUe3wJINjh8XUB8oU03rTF3fsUEkIXMBiVdIEgUFwN27zNo6CiYeuoCpCfBBDvtErzSil4e0qzhDhtre1LF5wQ2cNgbBQAYjOA8DJA4DqNQJsBpMVW"
    "A+5gcgNATBEKz7BAIKdKh3kUwkeNsF16679jBcF9Kf3UCd1cJgjsBSGxz8o8AD9AghcInKwCZBMgycuzcfKXAZK2wHujVibQN+p4PFJjtwFisFWI2Zw9iJ1z8Fy6AuTnAwT9Ru1YGi2cbQBRjIKlAU8WlLdT7LO+e7yDFuIFgJgdAHGXoAjP5gBRBkYtgbnTXw8QB7lS"
    "atlTAIJ9CrHVLe3H/AOppQw0gRRoxxCLh13M38Wpt17yQncmYz6Zsdd0EXlDgNzeysTZLn2SAlH6nn8lQLSGJhpoLm8BxGLoECdmbJopPgMSO0JVgPx0gGAII1ZctqDked+WWrEqE22I2cNy/jCVIWC+4cJaAYh1kIjhdgeirG+EM5Q42kj8WoA4FJWW/iSAOHhh+VJ3"
    "9KF5gGADxQjgz+JYFkP9EXsAAl4w4hPE0966OKsCyzYhEa4CZAsgeTNFux8gQ4qed2rnHvpz+HazWwAJP8k2QMLd3toKkF8AEKdd42xgMD3gx5vl2U4avPTTOoDMwxIOIL7a2gEkpniBX6vbARDKHah8NWdmgaijj/2jIVJo3oeB/LXEKd42xasSAJJkYSVUIRdzsPa7"
    "sHAcNxTKYHkiNNZExY9NE+NjLVTJWnnZhZXODUznVUHuQdgGwNCo+wLk4xaAvGoQnWUjq69p5+6Yn9ixir7PQKkHAIRXgPw+gAhfHlHe5q8AhIgFQHoHkM4DRHfeqs3SeLcB4u5dbwClAAnGERxtTRFj0AS9xabIbTuvV+emXD8Cvbw1tkx2r0rqQDScqz37mhL1NYD4"
    "cRvCQEWln9U0Bs0PMCUVBzH5SUx6h77RYvpl/HwP390nmZK0sjW+kIX1LgAZmmyq9NpAqWUjlmHKoVvaLr8YIHh5+T5O6vipV+rK8SoMvwoxuikfhueKFxy1TemgCpAfBpCPLg6BJzJUWYxepPwHk2YfQNyR3gkEdSC4xe5bY5NCwi2AhLQungNExbgMJmOVHGk4Ui+I"
    "6PcDZHpZ8NtlD9qNuUq703iBIGFiE8TLJytjiOaJf2KPukkBwuF7hWi6EIamvUVvAEjBAonFPHSlWWJwban55C/+wgBhTZM1WlBrI21JPhPd50CkbRtS++8XA8SxA/KvPoPzgH46U1ku1D6DDj1jgiTUi8F1uYADw3OF6wl8rZDYxSpAfgFAgtM6qFnwIi1+xQIpLgOk"
    "CaNOoafJOD/iAkDg5oU7OQUIgeiIz3t1h9MSQJLq8UWHrw2AJJXoZKxEnx7sb+6FlUraNmj1icNOfZ/3EYOONF32ie8DkD6MUnfqfAUgod6f0P5lATJ7bgCUtyT/qlh5Dc6+aHz2HKTHwf6izRu4trI5vAFAIPfxGJvLKUUgf2qp9KHVODlGm1S5v+nS3QWTxI9jozoO"
    "5zp2i4B7BcgPdWHBLRETaLvCNn8tBlJ2YYUYCISQ/f46uTcvAARj3nDKFCA2lqBDqtbSv4YxkNXvYA0gCmMghSC62XMJ31KJfo+6v7yV8bIl5nUACY1r+Gx+F6cYs1HQ3WwFIF3XE2xmQEsAeYlmivMu3ZDKkH4akvecn9YA/b9NItAmNIe0e+Xw59cDhJ3cfaemDoHY"
    "B5AuSgUhsyppXO27ALYLgEiatbeGw2C4bAXIrwCIwgpCn8hj7DJjFwDS7gIIlSHY6C0QaN+LjURWAGIWAOmxG0pigXBM7MLkI9pa3RYAsp2F1RSvtGdXoj8AIKVPeyVAfMBpARAjwhzc8jti5pjB4P9ZzQACSiLrEfJNAFGQnJBAAJE4a8/JygBZdPrnM0bblYS5XwYQ"
    "vdw+wHwltigAmS+Lq07Oq9lLqTmmAuSHA0QFgGBEWeCtBuO2ZGFYijXXpvFOQXQTr4vLAMG2vG0CEFCJjQxrs7pgTlwACNW2rQBZCdAInMtO5wAR0oeCGrkCkDFUNPvmwYOhsNjluwEC5lnfrw4vQDfUnxWAbH/DdLWH2a8DSMFrXQLIMvWxs4sKwpL6qAD52QAZ60B6"
    "IWNEF6qZ9byWENpmFQsJRaGQkM4KCfnU3wRiLbLPAXLOAcLJGVpXjQnCqnPaLEabrSzk5F4qJLTFfXQFCMQFtGjbrmvzSwg358xX0JcBAn0g/fNzgDh2QMuA7x9p691pfKUv5rKV5T6A+OYBax1j3hUgy/7X9JUBwjl/jZHZ+9bB1cqgkW8BiKQzJQr60ql9iwFvimml"
    "dpGHBWZ7sZWJyYMSvnHhopVJZ0MdCAKEpomUMWI/AgTr1unU5xE6kQhcGyTcmkJZ+eVWJnEi8c8HiLorQHylipx/pQCQWDy3BhCWBlGyLYmDvWz2ZQU8FiCbX1Mn9PDnaoD4QlC2vrTf7sL62OnConMXFrPH13Bh+RjNo1TyVep710J8vgF5yHq/ChCY92t9hR74MqL0"
    "y/JAbKZYSsOitkkVrm+m+DEDCHQTCZtnTqxMElqmNK4JIFCSYoyM1yX8qxsXRwtR7ksAoQVS/EiAxPEdW5fEVYWEPre1kWIFIA4Q5daabQTIvPs/2LRaXqs7nwwQd092CxtpF0DAv7rx4X5jFtY4JxRyJtaysKaZrRxrY7tyFtYxy8KCfjzPBQgOHKVdC5Mp6Dwotji0"
    "XAMFS994HY5txdy1INBGrDD+FlLVIFVjcyEKjvKhvM/t5X6LCwt2ixjCUD53Ngqmcvcfi51lVm/AQ0vy7GFohBIvgRQgMHUqdnmHMpMxr6OXchxMNS5NUSysCymmFKP7cW3gXb8SIE4nWUGzCCnnPxMgWCV/aWd99Uh1JkXusuRuex6f6ooAiQ1T2BIg4oaJWM8FCM4x"
    "WKj3PQCBVk9bDWd+WyEhg6HmNhYS/sfCiHNWarHrJNxLTiHgJPRlISEcNfobHIh0ob3vYwECCUIwpP3EGIOR7jYdh7Q0sm3x2lf039KuTdfGHUZhMPzMlON+IT6DHxdyLC1EHeFbYvEwa6h6JYDAveobVPsq8OxbkHM1yjtrJz0zzjIHI2Z8GIkUb+EUIOBYmcpMxFQV"
    "0ko5ZejGpcHDUNqKz6i2SdU8LHoejLkAEEjsct/9pHX8huAHAsT3vNqUTkh2C5h630sL0/BgP9HEdzQ0TJJOZHoXFixXFSc49628YcLJEwHiu48VbKQMIGvBE2wpyNi7AATLx8dKdGdVrLQpgUKmsZXJWr16OFdsZVKuVn8gQHwSZ/KuDFTymp3hdgralJ5D3f+5DhCr"
    "C2Phs8vS2RVt5gjEhRznb+bW+xlMNPxfv1z13QCB/b+/3XswjrAIhJo8FI5VIQvd2hnhR6LjNMOYGQ9BRRNb/MHo8fgRU4AEOyIUAxrTheO7doSJU0pjjaDT+E0ACBYWZkmUVNrZBw4TCYs5l+FXbU2y9HYEiCDTi/jo1yJrg+6/GyAQsbDbcl079+m0Ii18SPbn0E/Y"
    "LGR6l6GxIjlAXBk9LwCEPhAgoXelLDEgAQhXKZNVlvyrqNBvBBDfzf2qXlh6sxfWBJAn98KChuLj0tg0nkSszRhi5S/PAeS0DZDC50q1KTQ8jQthyUJs7kdW42EMbKbYGeCuRsj1AOms222GjFgbe2DAtI1uBpDCYEJFzgb7jvvusUndLjyMyiOZHJsDBKLlwVCBTRyc"
    "BhsITqVdiQXio2v+WoIIeP7FEiln5dHQC2vSYm23/D7AsyNkLG3wliB2LZl0n09EStKZ8cFVr8b3AATnSTUbwm6bcMhYk5bcpX25WNPIhTTT22CrluSZm9TmYwCSZ2H5Kc4d0nHYXAPuVUeBhgRZ/WG3TpDaTNG963YzxU/y8R3NFL3zEb1tca/1D/rjYM506Q2Ee0qa"
    "Iy+6b/Vq4bI6W7DTpnfBd7JZA52wEDtfSOZG4UeINvkT/ds4DWvRmVW2ip4GEKCaV7XIgLFSz3SzabVUFL4i6IoY0mlFQhzwaHiq2DRKQkTKINWZOGeEOxBFDE0oUH2b5PC7dVq0ZcBeIbMf0Mw8kIradBvcFhKHONhPAikXsalgRlayefYOM2qyB18NIHEi4arcOCE3"
    "x1K6Py8RK3ub/AB2k9Z8BEBUT7vAAP9f3/MYEDdcdKOJFIpWpPfH+9SBjADB/fOeeSBOTx93zgN5MkDQ9QhIoNEhS46YkO2+IlI0JLy9dDVAwDkFnzIXNVtII9EZHTy/biEAh9RMUUeDawvOHe/egQaUdw2EXA2QzDafHDTLqQ+KlL4hPr0+M+vHh1XeqonnZxxv0OR4"
    "ni6NZ0eHoX1kkVG8WO3qr7W19NxPEV5VfPCVAPI75REAmbUy0ZfMs/V27gNrrOnTVrxrfrpfOpEQnbrHT3vaBAiqOkI35xaeIOkSx9Y8d6CUL6CHTLBJm7iP5H7oUwkgbvsvVwthtgHSAUAgLT660nOXerA/8oQkdXQ7Fn3K/Fww/3G2XrcozRZpK08FSGLW548Wjts4"
    "w76HF7VbfMdpFs+XOLb+qbbHePDFEgqv2neqCpC7AmS6IS8C5LC7maKcmU1sL8QW80CgY36Sa3g0kr0RQBw+wYCzF2aiO8P9bIzYnonuhyUYK5860lZhxYkWeadVcGpSWup9B4EubHVql6DYARC5quM5Ouexe9h8IW360RF47hvNkrMUFr9CgzL+fQCpch+pAHkQQNRl"
    "gJibuvFeYQUtB0rhCEmV6Jd3aKaYpE/5bNLDpmDOr75wkDtXo2FiyOGZACGfYBcVNa+itGhluM97Ykwu32IHQNZTOWDir/tR/hZenk1OUGdn7c1jvb5/dLHyvwKkAqQC5M4AsXcESDoNMvTRGt4HIM+TxwAE9vOHdOxa7v8uKHOo5bBCsoKy/gpAIBFhLRCezlRwOxQY"
    "vnVcVPhD/+KTvZsTqwKkAuT3AcQOF9To2BqTPxEgzgZJ8nzJWwyU+i0AwZ5bcv/G3QEHiy7cT2g/rwQI3QJIcGBd6mjnLkDJTnLFt7aSN1YBUgHyvgCZ9AYndhguqNEEIOxZAIFFTk2dVwr+3xkg3oOlLzm6vgMgRPgg/95vgZw1zK4kXUlZfwkg6MC6XOwKQ1eKnMEY"
    "ysnSCpAKkCorALnQC4Xlx/55GkBYUhvFSVvyYb0xQJzCgxR/azV7NYD4nFxD9qpJglXoVPEjGCLzUpAvAAQsm8OOSg4F0+dtsVgRPgvTlFeAVIBUKVkVF4PfbLoBIST5PIAMTNI0VFNICn5XgDCci44ds41DyIm9FEAIAISZ3d/dERKp4KeG7b5etJW9HSCKum9XX/6E"
    "HiDl48CI0feKoleAVID8DoAkI78uAGRIK8bprPHWQwHi3jozfgrFIG8KkNhUKrYKgvq81wEIB4Do/bt2sBN8/q4KpsidAMJJhwC5/OnMyQGv/BbQKkbvt6YqQCpAviaHHwGQJLNKmf+x1U6L6U6tFh0+FiDQtnha5llUgEz8IPnI+Fv9WI8BSKsPujCOZNXjpUNVOCZN"
    "2UUnjIsA+WcNINAMxpLLC3Zqfc1igoSAxhxVBUgFyMNlaDDn/icAZJpIotqtfowsGXEFjTafCpAhe/N2mYj1lgDBsmqVb/mN1a8DEGgjvD/1FdT82PORSghmXwOQM7YyabN+FuFY3894D0DEFkAsa+7VUrECpAJkkx9SQn/eWxtkPXWlk2Yep1mVATL1YgZ191SAuHdP"
    "Unn7pa/tLQFSqJVQxLwOQLALpG/iuO9woWMXRU6WIes9zRR1HMGNY7jjNYkGzWkPQODTrUyVQBPnXtPgK0DeHiAMjAytSx0CoUl76+eTv7wRkrQixVbf+5xd7dyL9GiAZDp8mpryzgDBaC8vqrlXAog4Z+3GeT7gZgYIPX5rkBJ4MupKgBxYk06TsilA8lZWs3XwYKlc"
    "AoitAKkAuZPilX6bI5ctxrH7Buekb23z6jYIS0sJt3Rvs6nBHw+QlF9LU+kNAaLLjZ/IbU6sxwDEfajsWH6czbchWQhdn8YYg4Jui7mF9TWAnLLKRJz9mojvV+4Bcq4AqQAp+2u+1nl9ZmP4ocqdWdoZjfSpTeoBuuwRibxJq6k1m4ll5eBLH9LDAfJHb0ZB3hEg5bpq"
    "vtZv9xUAos65eZQmRilzSviATQ1z99dOF9YouQtrBhBzyjoc+6KTaoFUgGzhI05gunF2RqrzcMal4r5Btm3mz4bcWNLeNCbwyWH0tFGIWbGZ8oZUVLKnA6TRadeVeU/F9wMIWwsu+Oq914mB5FZGPrP9RLMANiT6jW8EXU1Eysgb5oHwNReW+sztNB+7rwCpAFnVATDg"
    "TxgD041uHd+X+VOmbTs5z0rbGtHGgWGvHwUZkjwsuNNKK2Z5S9ze6OHPswGSd1JpZ47DGUDWowC/ByCrayLmVbKwFgD5gPoOlJOfNj4BBIDBZHIpHsWshmQPQMopX8sgOgzOCwvB4ksZAAJZWO1qFpbWNQvrPQECLifRdtTvS4p+p6vOll1Haj4o70cBZHS4+dBivwj9"
    "Dz4nIJl42S4LMZ4AkD+pETSfj54BBGZYbcq1TsyXBMi6KjXshdJ4ZdbTitAx+IAhC5rCBrNwwaqHMcgE++LORz1dAMhKznAhjTddCIsAuZDGC4CqdSDvCBDcQU9TGDkOZrtdtTez8OXcU/WzAJLjkPuvZsj5kc5HgGxf9h0AaWyXhvHXAKI6Ky/ItT/KzwIINzd0NHlg"
    "IWFazcH9KFknxzYDSCjjMPToBVqzwCSR5EPeDpBSIeG0EHzOf/qtQkLi3WAVIG8IkAb5kV3HOOf4RpWrbcfnd/+dAcLYyhxYp8/vXV6SRkGQIM5As+lc8tT8WBnJ8RSA6CwRq1kBCHyAOJO9LEY0PwIg0IAcyhm0LgNk9VyFRF52wnOJtSTfB7UykVDNUXgG9nBZdxHs"
    "t+7WOCVG4RciU8v3CwABWuljYR2OaB4uPFoZbAUS2Aurrb2w3hAggyxVXLU3WwfS0BlA8sLsrwPE7fplsZId7IG7Fyg2WUMMd1cRZ+BTULWd+y/Jh4DidM8/3wEQ1tikI1aeiJUAZD4xeSGqVMr+kgABj4nTtJ+laefXAgSr1glZs3Qe1guLlYPSHPobJnAB9eyHJo7C"
    "MNUstTpvBchmM0XVTdaJb6bYkbVPXrvxviNAwEdDSmbtrSlS8wlroM3uCBAGkBAYphmWiWRuj2blfdukQJeS2SeK9j1Rs1o1qMEo4OsZAMkTxmj2LhlALiq23v4IgDQSryNocn4VQM4FgGj/A/Gjsc9v516GC7i3yHikBH6cmPt/J8z/lx3SsPiXACKd+XAuvlZBA8UQ"
    "qIEsMTbfH04rZLrOA3lDgGhxLjf4L6rCi9p9eDBAQsya+Er2iSED7M+su3dJT42Q9yWIaOdfUdiuL257UcTuUwDyR68WM14JEPlTAILVCdTo62Ig7LAKEPJMgEA9+RhfWLDlNLXIhCg3OLCSPiQG/KhQbr8XIOdLA6XKKVQqDdZj9Ulq9SSHYZD/WAHydgBhKyOZsZzh"
    "elWr2bAEyKoL61pFj9nGBu0B1bs9FeQThfLHpoHRD3DrghNG6K9Xs2TQavc0qob3Lb7rcwCSNeXNAk+/EiBadhsAEWvxXGXKZYehS6Flzxxpy9jBmmNx+DmbLhpgA4bQM/mPU+1sKpf80kRCd92yYun+0WDbRhX/ZVkxGRiyNsonqAD5JQCB8HJTKPBYvdxLTfkYnqSk"
    "mr0Wl6aVjRTzGIgoAwTrKvQVysoXuPugA+z23b+gcAU3ZBD9Dd1XIQTZGnHHKsXGrbm/gBBOSLf2nk8CyGoU5NcBBC62UGuO2/NFacvqGyicI1s4l88zgp1/oU7mMQCBHT3EMUghAsKSZpBHrOGgJG9QRdAeGL2rXwGI8qYFXbwaOr8kTiv3XZ/Sqyx//b2qQCpAXhEg"
    "bhcNuUOLDKFm9ZqDzI8ZhHyxul5qyQG50tje6S0p82QMtZbGG2q7r4r2TyFrZ2e0BGMRvfs/4v6k07M+FfmOfixM1lWb13y3Xj3zHID8GXRaNp/w/7cBhMGg839GgNiTXsQ1ZGljD1/LX1sYmm7NCBAs43sOQD7Ipz0UZvy5vdEJo9o8hh5YydMFyDiNlPkKQDAh7HBa"
    "fEiFBlkCFgips0WbY2T44W7jpCpAXhAgsHnvuq6ddzdMB+kt4m654o8driDPc2GaaCxPEwRCJ8mM7o9COD4DCDVyv//K5gmzXUcwbcjXVnFO6Wx80B17NQ7Mh154USuhzSPk6vu5F/cxyWnZ6OQCQNBlBy8l9FKCLeiu8X0ygLhbnl/KvwpC6JXBr6cDxG12jTM3wynU"
    "kZpPM0/A1eXRq2oecmeYGxsvHUjq+vw0Cxw9CCBR89L8hgHXbjLpXrXOxBKkFJ/QuZ+r0ME+DaLrMlRxId6JNV8IGHdpdQc/et/aMU0eUd6zZe9mgFSAvBxAmAQXjK9gzdQc0+e1n513uXUQuuiibprbJsybN5ZwKMROIwboVi5XooOdcEWkBTrj5vdPbnJz2ubPt/dM"
    "xoJOphhkmTME03qNkFtRF/f1xFE+YBgN16nnlo4vvQQQR7nwPn2ah4WBo53S02UrltcCCOhWlWYw4O5hRhDGVrrxzg+zNDsXnIwsCPIggEDyEmzpLT1OtjNFGE6xauVr9FTBbnBkOf1zTAAi3IlKndhjLyyqVvrFQyeSsJAEH34h6UeHCvjDSYu/I1SUx0xzvwhIBcjL"
    "AWTyW/JZgUfaX2d518pDWq5uxmkEC00GvVAQIEgojRvhoF0XfpckBkKI6gXb7UXKU+Z5nwNE0Tw9hFN55zp3MEMM7cOUbcVD4fCm8TGljo1lelc6iMDu8y+1FwNG7quP9YCpQw3Sm9udYq7xKn4TQAralCxqAAvzQLiaB8mZLgxSUk8DCKj9E3wdY4QD7hf3yMmO7Q1A"
    "Z5d7C2Me8Ck8A2fCQPtxFPg7fAfEA+QzefYIz06mBZa262whlvmxwLmzC6JE2JYxzpqH32Ylu7cC5HcAJNkeoInAEvXUbQGEZeXq46ZjCQUHECuZ7GGf3GDGSB/6atm5Gp8AQqB4S+zdkM8B8kGyIj6IhfDHAuQPWiEQsjdRp0MMHwtPhosvvLFRyJ/0pZde627tUkcS"
    "rJ3ZLa/eymQnQPKJ6MHrOc/m+m6AcAyFQ5/1qQkWKugujXwxWy7SgPTbmLAMAGFa25kEigJAsN5+/mxizsSFfH6e4bqGEv9lZB0nyx/igj/dev3L7tXFpALkFQGSNn7ObZC9ABm06JKLrZu5yT1ABgmOEtOEBoNOSm0ZpzRe0tKvAAR2/9OaaEceDpD/5QtOWNZ68PUn"
    "8z5YvgEgZAdAoLBt3qHHNnMyNN8LkJDohIL15ezAWLbtd5exPolyl0IYEhLD6Dgqiy0kKHbk1KHwbBqrzBfCZgvJfNKwzumwk72n/VEB8nIAyTJrs6g2W+9fg3dtWq7Okztxppt9c5FBt87uaBs2+dwLbRkTC+RrAOHqmHRR77r5fvP+AElTltnd5m9VgFwJENCnaeAC"
    "43IlgGDMyrev9U1Plscw6QN7aTwFyiKeBhCnkD/BiI3Xk7bC0GNys7ldmKDlTBd8rvVwUe5vu5Dp2VbY5fOztBOoThwpNF9Iepyx8hQPA1uE3tP+qAB5bYBg+QVLsrBWg+iyOUzKm2zqZq9KIWPIAQS7M3IfJV8WCzb2TgDJwubEzD/IwwBS5ZsB4l59nPyXHHO5TakN"
    "IiQMffqtDPSvLbXhhdSkxBeKJ1vC6IEAiZr7lEBv1l6to6up9pR2QXm7s2CDtnMirXvpcf3Ztp2locBC/Ewr7NxIj4qX9Tsex3xCtbg3PipAXtqF5cNyYy1HYy6n8TIYIpjn9y0skEa7Uzaig7yqJB2QL9OORoDw3nwRIGSqkXAw4hUgbwEQTOOFTpYkqrMzbVf66EKJ"
    "hy8yFeVh6KCxz+eY0Ad69oxZRU8ECO6z6BkjIC2251z31C5fOcJvbNCWicowu3x2uZBPXMjnvE/o4rjOx2woXT2sAuT3BNFntX1ujzXEu3+9kDA6n5pZd5JlEJ1BbFli0yh37rSNBF8OlBoBQq8DiF40LiVdDJy7K5pUgLwJQLz73Qrs3cePf23sTluyQU5j+9qVI3RW"
    "SGhk6VyPBUh0nRXaqz1f9i7kkQuuAHkxgLC8tg9TLoJWZ6IrE2SarL1oTrIcuoe5pgLb93BiclotCglDShgYJ18FCO2j/e7kowLkTQASemHR9V5YV4q2vhyKnJ/ZC6tKBcjPAMgw7wY4zT0aZPneGJspDWw+L6RQSNhY01GD/dTAC5D98KrLB0qFgkQonxBfBAgnvW8k"
    "pI5LO7oC5JcDxN4TIGMrk0MFSAVIBchC+4q8ZduUo8uaohMLbJRoveQVWWrZJAsAQiH9SrpzqZ7M736Zm0NYlk2NlfZrMZAp9EEKYcYKkDe0QNhuOaUuLW0/SbDMK0AqQCpASvXMMwz0IcF2gDYnahEli1YGk2KeqmGWA4cygMycSfN27kPjx4dKLW13FUBK9WMIEF4B"
    "8n4AwSZNSh3TiYTgSt0rbgc1m0iIOYMVIHNVQAi5f5y8AuRnAQQ6SVE1vylZMUYec6dCAhZdOr9YQbl3dBUgixX5AgrWyK8DhII7TJFCMjLvKkB+N0BwQkbaRhGs5f1i0hf6aRvG6gqQ+R2GyWGqAuS9AfJn0a3GN/30VYDYajHZc8S+fe6OnAdPaHHoN06Z9S1E1Szp"
    "EF8Sy7Zn1YfXAaQtjUKjvgNgESBN1e6/GCDYSBc6biQcaMlsasa6kM/khVg2h205KkBmN5j7VgQl1QJ5d4AUAumRBdh5pIcOp9CZ3E+2wBD7vAAEI+LFfT0UgjR/igCBPQyWvc7bLMGw8asAcl5eUNAKt+toadoTTMytZeK/FyBhLEg6/2mjNWhBaEqLsZNIBchMlbuv"
    "RTyXHxUgrwiQWTsS71vCueI4KHZsGk76zvcmhw58C370K0P3BuYAcigDBPKyKD1jK9qsxd9VAGHzZOJ4rbWmbfuSj5aY6sP61QBZ6vnbAXLhxG8KEGx85Qw99VEB8vYA8f6oRahDQuWU+18RuoaDVyA+ZmY1ItgEhRXNG4cgiTMJlwD58EOflJo1NrkKIIi/4iVu3MqL"
    "OyTVG10J8k4AOegKkPtq8g4GGT7ZgVUB8qIAQSWcbdW5gvmCUcb0lCjtopXz6tBvMGEEdFGkBYBkcZfmFoAMTMuV0YnEWFGeZYOj3ZqKkLcCiLripPvrR94UIO6GZ+yOo2orQH42QAbMyZ39VFuiZv4rs6aPYeRe314GCCdtk0ZlKNkFEN/et3w5qfNqjgj3s56qin8f"
    "gHzuV3aqZRUgF76io9B2dQ5uBci7AQSL+G5OyOP9ekwByUT1ZYDEkeCHg1+MA8gOG2FRxZIpl3bdxMYxOVXFvw1AmNm9FFRSFSAXAGJg/vlHBUgFyEgQcyNBsADksAWQTkPVyCZAHIXENPpOQC8sI2VzyUho7PrAZU7PG2VOELWpKv59AGJ3T+bG6d4VINvfuzrevVV7"
    "BciPBshaMtMeftj1iarurB1pdTkLa+5VakOsBccR4BycS9lSjejUOiLOG24LwF7N5X0bgBzKw8NL4jbXhwqQy1/9t1g+FSAvCxCsO1c38UNvGTbSdCZWom+4sHxKbxRCcLJAoTdKLnrj/lW0vQCQGkZ/H4BsTNicXzZWV4C8quusAuR1AYIBabX2w+HsGr6In3PS2c1Y"
    "BdSMSLZSB8LT+t95nB4qT0xz0QIh64Roycd9LRCogfk9+n14J4Ac2M7tUXmE4W8BSBgfpRSvAKlyX4Bs2CCcnM++p1yXB6b5cgLIQuc2uhlWWpmM0o9/jIIMaS8CxFkZCwlnN6bwXBRyfQwE6ip/S/IvFok++tO8FkD2VS0oaq9p9/4ogOTNVfjqAMCibOHjCIMFKT1u"
    "hAf52vvxIndwG8hXl77mcbhq4RUgPwIgazaIb3vjLjsjLHTXTe0PsV2SB4WEaxYIRDnWxZxxjPol5IkWaJN4vxA9HNsGi/NGx7zr03gbOYb1h3H3PoC8rhGwZRgKK1d+vNVPlDwx/DCAHHblGV47RORBAMFB5SEi+HmGybolja+OZ4OHtaMYuG1W3QhQ3IXNg2Ccrzvl"
    "ysC4tnwKdaSFJyB4mc5mdzszvxC/Lu+PLryoi/HOUbrLX08FyEsDpNR/9yOMCIH+JdjeJGm9iPbHcElRCWFZMQZCjG+MUhaojr8IEDj9oh13C+UfAJCtFt5CXl2M3gjxv6Hexe/efbtJ5ltBQsnKkDu4/KcIZTZ4AD7kDxzC0+HV0IHYHwuPTUfHF0/1OqyJTy89amx6"
    "y+yvJvw9jCccwvcmQlOw9DxDsE3Cmxxii+TpCRa/APajAOLOfbFwgRNxjQPrYQCh2LwxvRvs3+OyExDCbn7fmG6t+Mlgg7Bx7knZJFPEllPW1NHognogZ6Hd6tSIn9maNDbE58tAE1us/HKeQwXIiwME+u8uG474yYHMezz0mKwF9selQDSDXlqm3M7dJ+qmGnGWodtf"
    "BgicfiH9DgvkfLUF4tDpW4RpKfzYEuCi/9MpYtzTj19HI/EprK+Ebi4IWu2PBBK7hyx2iYH6fugQI1GZ4wh5eB22IXMrRADL+P00/v2C5eAOsNOHwPcPy/IrhL9YfAWofL9M+A+2q/UrZ/geY3tJfK3BpcfThpPJ8cNCnjUuu/lRADnoS6ULkDvOXgAgnEqdTsGKbeXz"
    "W5OHgpU5QMofEko3Tsy3htQ4NKucuA+zF50qL57gpJfFg5iEOTXFUvQ0W/lJC0OPs9LjkKlQAfLLAII2yPznUa3XFIeDV+zhsltp4L7AABQSNiWAgIoHbaSn/W72ytYBRF88fT93F/soCFe9aTc6eJPZOMTL3ian23DGROPuB0JIZ2zTCAoOMyhrx54tYz+WAW6anlIc"
    "kQKOwdZBgInz2bivTLqdv3vobIxvU+n+gu4w7f8YhkY4O55QpxZRb7uzwiCL+DWD9deTHpaBsBHmbHRii3W0791a3GtM52dYuBV26OATjT/hgP+x7kTu8zviuPUMcX3xG+1gTdTIw4hodwpcYg8dNaHFGCycdkb/LIA4q3bbBrnW/ngcQEC/QvgQxVt9Jz2r3XMAOcVu"
    "wZOsOeqIBWy4k9h//9tiU1SHA1KKAWmn9AuqHCyQMkD0KQUIWiB+5dFehVVlpot/l9nSZQXILwDIn2bRWyobPTuM5SLgIxouu30cBhxAnFFcGCgVWvEaAIm7pA/zGkHS6Yu4oz5hK4x7GButcNIbcxwfn9ED9N+VAMGmjdAzUmN/Ytq1RjYGtbNjiX8YbC0PG2n8aCKw"
    "WKAUBiDgFEgPj9i2te6GoqDmET9eJQNsGpx7dISpvk5xO60OhOxNtAYs6nL3reF5LRhfUfEDToAabQt0QrOsBYIY0vuFoAGBnAMD40w8QKBPGTYLiJByP1iHLxYIEPhYXUc7hw0D8SX3sX0vNPgCxKtaIF4xLW0JnFW4XluaDgKZnevEng4QDZGK6HB1Kv/EZgQBbQqu"
    "3tw3u/IeTrH7OSlweVGIpEMkpAAQo8GEKD2zHyAsXbl7Ggd6pVoFXVg4aSV1PdcYyG8AyHLUB9bmDuM+PDwLs22bHXGDFirRIVsK1PYiFwMEFDo18/aGuwACDhcMwJ1DIO58hmBcr9A+smmILhMwBa50YUF4yBHPgqVxpqBPWynbHvSwdDutAWZvAQSGABtIPPBRI2Ao"
    "dgsTPXBLC1DC7lsBv1XXt+Adgm8Jvs8Gn2t9J8u2swz6HsMt6kMVosUnjINXMH/GyS3a+Agp+KvcK/EwMICMWyGc0DfFRIAYfHvS4iEOeWgkRkg1PjYC217f4cbga0HbOFCCIoMvwAElHvKCAEFF+W9bpIFdizFDXp4tvgI+vn7qQCkEiG2ParxBAOgnsBmSt0OA6EWq"
    "4ZrmBY8rxdk+cN9BmdXn38IkT/8dFOyYKwCiaXAF+Fvbon8qDZ8AQDTGRi4vvQLkZwHkzx9tW7UIv8UJUzpk+u4sxAOPSN+CsnMAWb9/uHcJ3Q6QTCi9BBA/geS6CIhoj6D/HUgcfwS4pK0z1loZAsvQvgtuJjYCBL1zDC0XAKSjKADEOOOkQ4BAdnOLpokG/7bTyQgQ"
    "6ywL3Od3nWjA5kAPWQQI6HanwL35Q6LidybPGVjWNBjCAD+W9G4qQ87SBiKNADFhSXAk/DYUFdQIEDiPD5GD9QRv0fjTOcMILCrTwtKaVw2iF0ba5tm8qpyitEIcPNdzR9p6gCRn9rNAGYxonzS406aNQ8qu93fbPUyFzD/xMloC+h8st/mU0k2ANHOA2OR9OARftL9B"
    "eAaQ9vpmvhUgPwEgzaK+m4PXnaV5vo4pds9YP9DwbgcCBX+EbN0/7hbJR5Xvc2HB7nnpogoxELMZA7muDgTUZut9QeaMriLaOQsEHFB+X27RteSx5Ky4DhMxwagAOji7yKlhcBU5td0GgIBPATMMwOHV+aNhm2+oM0ucHQFaWsBpHDYSgPghX2cB5g++YzAbe/CBMccD"
    "g2+GYBCwQnBr5BaIgQw3gEoLADH4aUJDTHBhubWECDmeTPghYsK7xYTEcC7g50WD6OioUhg0PqwQYV4Oq44darnF4Se44P3V8myA5GcOejgNcYBVcdKf+96fWnaazX/iqpBee3bvIXUp5/kKgPxDZmwG8ye1nipAfjNA/gxy6cRCr/vEj9g893LoudG6GRp092/fP5xk"
    "OcH7ALLaCwuA1PGtd+vsNXm8zNkG4EA2Tn+eKfZdaSVM2qZgLkCEAbS5CXEBJlv0NbtvTaOfy3SdRRXslDuNABkGtEAcEcDHZM7e49She8mOADFjcD4AZEAdCBBw7+hYMXhLoW/R0eX9XBp+xQ4B0rfxhClAMMRv3PvCjdxCgMMGgGAoHvbvYIwCjGygFyUYwJ9i6C8a"
    "AwnZPKslHd6PlW/v16LnUZG6q+V7AeLbf8JJxlvzKoAYuyvLCVQ9JBOypWVzowUyhjwgE7MC5C0AsmwxBerYQFw3bN3wQtsXUYFkDFCU5BJAsC7kaoDQDYDQTYDQawDiNuCgVsGHJbwF0nmAoAUC/GgjVQYfO3DqGkMRFowveKWAhCoIFUK4w+mStokAwSj22al0zI6C"
    "VpIRIBh3gWw0NA8iQDB3ucVItoMRfAymHUD8T+LY1aJZwmSwQFo4AZywBYDYAJAgLb55O4bRwQI5QzQFzQsNDsIRIGjwWEdQiGuZl03jvQSQaIXwMZUDHINrAPE7FEK/HSCowVnyBT0AIO6wE1jFWjNB7geQDyR08v4IkHMFyG8FCASMFwoZXSF8vND2RRCcPSwBID41"
    "dPv+4Z28I0DIfQECxgakdMG2vQVN2nqAdBaiyT4bDJO7bKgCbEJ4xkBrehBQ1RDQxgywAJAGACJDhoFjjPQuLOr1vdvtd5gy5uw99CB6gPjMWvDD+Gwy5n1mJAIEvE4pQLwLC07YjgAJoIJ3Mu48PYESCO+08vTwEY5GJP4zsD4g5xqD6EJe7oTybQDptgHic1+DzlNG"
    "ruZYwbnalwEIPgyeIH4bQE7SXAQIDKolaDDYo7ofQHiIrVeAvAdAZuG2eA30o9OUmJ2tbJmEKjkmO69+L94217uwyCpAWrrZPv6aZopYYAn7fcheBYBgFhYE0Z3VAFlYInYRpiFYhNPkIZrg9TR6kXxUm+YAMeDA6vGQ3ljMl2p9spWzcfBgfIFk0TsFZXztGcol/YLA"
    "cmDeCwYj66G4o41ONQyUhLQuSM6CAAbWLhp6Rk8UfJywPoO5XgAQ47OqQ6sYTLjSWnquQRYWWFHiJbOw3H4F88XiFsHJamVHuMRXmyfiuUTw2UIuUelcTwWIx8CorK8CSOdn0F56X3dCQTjaK91Mv38FIO7HsOkM3AqQXw6QRhQutSTNjoidAGksVh9IzOu70AKUZ4Ul"
    "e+tAyDz7D5IHsQ5ksxvvdQOlmI9bg+KFrfu5FW5r6lR9h8FkqLwg1BdIYI8t8Pdi5R1s86FiRAgfcXC4QXPD+CwsBIiG8Hs42oBjDMozaOfLOqhP6/XxBibOmE4FAGtpSKPyeViDRpsCAuWYINZiRxcIqjhQarRmfM4wws+n+Ep0K/r14UvgN20CvmTIBpDhBfiu8Bkt"
    "BoHgL/lyLiynokzXRUsXMlXdOtcIErQcTL5cqTrERgYhaZ34app5nP3JAIFZVyyqZw8Qs0uZusudXewn6TZVsD1QaLwtwuhfAghyevqyKkB+O0DshXDF9QC5/Jvno3H3pfHCSNux/nxMPcdWJpiFtdaMV5GzuA4gAgs2gnKGj2Rar8oBFaGyW3ofE86Yx8oobG4Hjh93"
    "NP5bYsCDCPA4AUAMpOrSWDiIYROIu2PxBVT0tdjxBIyGBgsJ0SeGZgqoM3eXt9S3IWlEbGEnJFoc+E/I/DINJj8Yn82Lhgq6sSCK2XcdJlb59/AAMaG0E+hw8PUlwWihHi3W4+b1srAgQHDE0tEYBnNibgaIIdO5PvCyAvPyOwHidDib1LPTpqe9abwcFfisnG9JGeHT"
    "vDhMoZ7bK18DiIIM97Fy0deBXD9+qALk3QDCsOfTPoDkjrG9AGlJQbAXFlY7rAndG8WJoRwLgYFG29gHy6ERlH7ozCiweAI7XHkXlvV9rrBxiIZnLBb5YQjhjPkIEFqAkzjtbjC04E/tdBSk4grfKcsHHUKXLRabZ/k+WtiXC9bSxK6VwneogkX6o7BvCqTPWXfesHZ4"
    "RzhhgzWD3tjAd8caFo8+OI8ONfUidNMSAS3wQUz7kgDRYtZzifu4wQ0AARjltW0wA2BBo+cCxDcviT0hA0DI7lNiOd/GKNqjgX0RIAbr1mcx9y8CJITRU4Acr55KUgHyfgBBd7rcdZuR1l4LEB16huQCACHo+6ddV+yl2LVXFRL6ZDJfCePbd2F+cuwviC29MAQdhmwM"
    "HgfYNdHBZoC29hoPGhiEhXyTe1y/9m2OY9wEj2baz4a3FpOl4ZnYPtHjy3dG9NxofIt5380RaOUD+PDmbIAzMf8zMN8XER5u8IThs8Bhf3yPxSGACOmj0yaNgKV5M8UXBEhBm5IbAVK6BdQ3A8S3KtThiVCJftwxfgNW8InlfBt6m3fuoohpy5aN73MXgHBIaWA6AqQL"
    "lehXTQOpAHlDgKBGHWSndg3zEVdbILbQjZd4gBjRrXfj7a6yQP6wUJo9NE2AhG9uLoMODilJ2DvdNzPRvsUJHDUkjd/xKdTx/m9/iO+23kz88RLi1MN4kkChho0vasaK8AGfhVeElrs+DO5fCf/xB4xrZ3/iZ4nrC+8lAyPHMVp+E9CMTzQjeV4OIOoKgKhwKRefd7vl"
    "lwSIngEEmt2M1zSl66vBVi1QZm6hJqtYjN+6zYyZUgtOuY/prgABewiSQZIbcs8XWQHybgBxCgkGLsluT8AM+lNf78KiZRcWueTCEvrrQ6CG0Im6MFRqnOOxzPXyGnvHWKllrSbbnKrr3vMwHjcUVg2/xp5phfnLWTKaZJpscrmQ9CUskONvAsjHHCBj49sgYqNhvfLd"
    "ULwjq1SFbthJd3zS8Pl6vwwQmQNk6jYc2ggfK0AqQNYqE1enrae5Xe4Sbm4ASD+TBCD9qtwOkCr7q/efH0TH7stxYA00IenkaSWNV4177WIMRHoPSzwXnOwoXhAgei9AuMI2z2AIlIc8iSlIgeGevBrsAQDRFSAVILsrE1ccr+nAS9VOXpG9ADErA6VCDOROLqwqPwAg"
    "vg3v55jG6/SlMauFIGMhoWZr5erGJGm8n+6fi0O/HSC+8mUUs52nq3zLYQyFzBKyOPm00Gp7/Oe8dP2LADnOAQJTQvKlV4BUgKyfcOzdHHd0IZ22Taxp8MJeWQciuoJ/Su3IwmqvnAdS5fUB4jTUP1hIyL3T38r1PuynCwA5eAUXTGflCwlPhxcDCNQVppG+CyMXcYI5"
    "jqo8lGY8nRJHATEnloXR722BLDZ/NQZSAbJVyo32ADiYACSE9D328xOpbcKThrw7K9FbsizyCJuttiN8VVR/TSV6lR8CEHY66dgL6wglNWtdSsbdtfrUKz4ud66TFLESXehTaUTVdwfRdWw5zS9kYaVWCHji2IwgAFKd3Ivwb5bN8HhwFtYeKFSAvCtAYjqoEX6EII0j"
    "uRuWVLxyKpvDSzZTrPIjALKvmWIGEE7t5mFnhVfL2bJv74UV1LOcALK7kHAe7bDeBlHZCmcAkSy9N79cB7IASK1ErwC5KnspZH9i28E2pr+m7UURIHdqpvgBFoiqAKkAKcs4MQl8/zsA0r5AM0VYhrOWZN7K5HqAQETHt2wniYLHXvFYye99vAI6sKfzq+5dSFjbuVeA"
    "XMuQgQ0wGZwQMWaEpsNHACDDdQBZnwfSm3YTIF0FyC8FiN+ibwNkitmqo9gCCB4H7QVfACDYymTWC+tT3aIOHB9OqaIHOhxg4OE0zBw63Gs58emLAOkWrUwqQCpAbhEcEZ5UDCZBEE6n4rS9Ewn7lYmDBKc23W0iYZUfAxBs+UTKI84ngNAp38isHxfP1Qr9AgCBZorT"
    "vfkVgGC3q1l/dWxAPIlmeWWNBwgtA+TzhmaKFSAVIDdplQVApovSN0O/Jo3XmnKaru973q4m8raipvH+UoD4/Cmp9QY/0qUps3UgNKyBJvbsFQBibDIr8CsAgbdIm5VAGzpMQkjEj5CfBiCuAMQAQP5eBMiynXsFyO8FCHY/XPuVOMcsvNu1Sg4Qt8cb36kX+jqAjB2a"
    "FoINp0KDwuLTunqwfilAdkiaXwTJ44cb5RsGSolbBkqVVqQdQOK3cPStFq3vwAYiEcFJGB3eTi8XhbND/k0vAAQeTrKCK0B+N0D+QBfvzT4gF/sgrcpBnGkGkEZOsYrEstkFkD+zdgiThCZTK8/C86xm8b4tQJhOdtLrDXtfDCDq9pG2RYCARo+paBJ7/ucCRZj6M0Y9"
    "oBUwk0YVfFPuoG2AQAj94HalvALkPQAyNL6r94qIG/kBvQa9U9lMTQCb5KK8GiBVKkBuEC3pbBP9+gDhGO1nYvIpfQkgRw+QwIYWU6RmccIjTEgfTRAMm5zmbED6yqmipAiQ0Et+mlJXAfLLARLblq/Jbdt33yq8G+tAPIbS8YfXurCqVIDcovbzvhk46O/VAQL3jM4a"
    "Dl8DEL6wZpzNMQbRIaVd28+52oeoRxJGxyEhs8k+flBVMj+9BBA8D0v74VWA/HaAoLPpsPL4rfoEWk93vqUJThu1oQW6jFfSLAurdxujqoYrQO4OEJiJkasl/eoAQXOAZe2p3LKbvXUgZDYExKftBkXvYyvLfqcckHGQxyTuPj/MD9kVJAt2HP4hM/8VdAFO0YOFhBUg"
    "vxwgd9cnje18pypoZQKlSr5scHCXElxb7tE2qUSHisO2ZkpVgNwdIEzPGgmionw1gNhkVBQmJftdPMm06dTKZLMjCCcdPZL4JGTB+HyuCBB5Kn7xi1QtsDamRr7Ym1GyLDASm61MSz/C7GVoAJy2RaGFViacV4BUgGxIo2NhBvbGVr0Jdecw2ZzAXXK2Yx3hH2yeZSpA"
    "KkDuDxA7T0Y93hpGfyBA9JShDj2FcZxg1nwkNFP8zCPfxXaK7nuy1h1Jok6HbljadlM/ydSKSJd9Okw2h2/Cru0ZzsOhlurTxzYonwFkCscbYbWznJjNjD6sS59n4B8vz0ivAHlrC0Tn7XGjBfInJoAYkaTWQsDE1GLxCpC7A6Qwd4oYq18PIDbkrFhQ0zjFoyO5Nj2F"
    "rvOJfJYI4p1UqLGPPr3q4NATprxhHKPsCiPgNxsDRtD+B8FjoOPJMZxHZx8LAcLimoS1vh7RzpyGfj350j/pRZ9WBchbAwT6KSYShq7iM1itled2QY1HUwFSAXJvgBSKGSAN6dUAks5bgro+Lc5ELbTpbKCU1rZEEI6uJh1nh2gs+og+JUXnoz/y8LcWYxSE/EWTIwzy"
    "kP48eawGAHLQ2dJxftXMZ4joma3c0gqQCpDt9GAvi6Gp/pmqcitAngEQcyyEi+0LAeSDgpLPZaGEwXKQhQqn0vBPTv4jpOcQ2DIMet7HtlYwek2vjXMCTS8nFxU//rXxNP48M6vIRzdmK7eLJUEMZrnyCpAKkL35XSBVw1aAfA9App6zUY7mtiDIgwBCKG0niRECvlCn"
    "6VFezuUoCMS7O4hH+K2+MUnEQYFXayX+4OhypsmyMW4elb8jQ7sYjTuu6Xw+h6UTsrT46Hzp5+x9KkAqQKpUgLwqQObyH/NaAOGzCZvl5CqnzUl5HmdZ+R7ppx8e+zdT+5C+sj7OCd5DzVBkfMsT8Vngzr6lK7V/4RUgFSBVKkBeByDLBmngz38lgHyUhmzuOGz7YMi/"
    "OtLjkZBcq/OtDNrFk3iaaFnwi2v6uGLpHxUgFSBVqrw2QPKes0nr2RcCSJUKkFcFyHfGHg5Z9CP+OT40i4vUQEkFyCMAckepAKkAeS+AwEzAgX2DjMlXK4+vPV/lZ0lItKsAqQCpAPltABmgtqJKlYeLhgqeCpAKkAqQ3wQQ6GZIqlR5tNBWNBUgFSAVIL8MINi9gFep"
    "8lgh2Ee5AqQCpALkV7mwoEFh3R9XeaD0Tmi1QCpAHgKQYwXI9wHkLAem16eEV6lyL/FtzW4CSFcBUqUC5AUB8t8/IYjeaKmrVHmkSOknzl8NEHZoxM+yQE5mP0B0BchdAHKoAPkWgPxPB5A/NVW2ynOSeW9J42WsEedrAfKdHxOAtxcgrKkAqQD5yQChoqaXVnmmXDe0"
    "VBlrxfkq5nDoy/SdIijZBxCz/9gq67+3gslYhlSAfANA/p8itEqVJ8qVN7qaN2nao1HUt6YM7F7vDZ+tSvEH39N0scoDAPKvmr9b5blSb7sqVX4PQKpUqVKlSpUKkCpVqlSpUgFSpUqVKlUqQKpUqVKlSgVIlSpVqlSpUgFSpUqVKlUqQKpUqVKlSgVIlSpVqlSpAKlS"
    "pUqVKhUgVapUqVKlSgVIlSpVqlSpAKlSpUqVKhUgVapUqVKlAqRKlSpVqlSAVKlSpUqVKhUgVapUqVKlAqRKlSpVqlSAVKlSpUqVnweQ//pXlSpVqlSpcrX81/8HMs7Nu19xnXIAAAAASUVORK5CYII="
)

# ============================================================
# GENERAR PDF (idéntico al original, sin dependencias de Tkinter)
# ============================================================

def _escapar_pdf(texto):
    """Escapa caracteres especiales de XML (&, <, >) antes de meter un
    texto en un Paragraph de ReportLab. Sin esto, cualquier campo que
    el usuario escriba con esos símbolos (p. ej. "Sala < 20m2" o "A&B")
    rompe el generador de PDF con un error de "unclosed tags"."""
    if texto is None:
        return ""
    texto = str(texto)
    return texto.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def generar_pdf(centro_id, output_path):
    try:
        import base64
        import io
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.platypus import Image as RLImage
        from reportlab.lib.utils import ImageReader
        from PIL import Image as PILImage, ImageDraw
        
        centro = get_centro(centro_id)
        if not centro:
            raise ValueError("Centro no encontrado")
        
        _, nombre, zona, fecha, img_ext, tecnico, direccion = centro
        detectores = fetch_detectores(centro_id)
        
        # ---- Logotipo (para cabecera de todas las paginas) ----
        logo_bytes = base64.b64decode(LOGO_PNG_B64)
        with PILImage.open(io.BytesIO(logo_bytes)) as _logo_im:
            logo_w_px, logo_h_px = _logo_im.size
        logo_aspect = logo_w_px / logo_h_px
        
        def _dibujar_cabecera(canvas_obj, doc_):
            canvas_obj.saveState()
            header_h = 1.3 * cm
            header_w = header_h * logo_aspect
            page_w, page_h = doc_.pagesize
            x = (page_w - header_w) / 2
            y = page_h - 0.5 * cm - header_h
            img_reader = ImageReader(io.BytesIO(logo_bytes))
            canvas_obj.drawImage(img_reader, x, y, width=header_w, height=header_h,
                                  mask='auto', preserveAspectRatio=True)
            canvas_obj.restoreState()
        
        doc = SimpleDocTemplate(output_path, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2.1*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        centrado = ParagraphStyle('Centrado', parent=styles['Normal'], alignment=TA_CENTER)
        nombre_style = ParagraphStyle('NombreCentro', parent=styles['Normal'],
                                       fontName='Helvetica-Bold', fontSize=20, leading=24,
                                       alignment=TA_CENTER, spaceAfter=4)
        zona_style = ParagraphStyle('ZonaCentro', parent=styles['Normal'],
                                     fontName='Helvetica', fontSize=20, leading=24,
                                     alignment=TA_CENTER, textColor=colors.HexColor('#444444'),
                                     spaceAfter=10)
        story = []
        
        story.append(Paragraph("Informe de colocación de detectores de Rn", styles["Title"]))
        story.append(Spacer(1, 0.6*cm))
        story.append(Paragraph(_escapar_pdf(nombre) or '-', nombre_style))
        if zona:
            story.append(Paragraph(_escapar_pdf(zona), zona_style))
        story.append(Spacer(1, 0.3*cm))
        
        if img_ext and os.path.exists(img_ext):
            try:
                with PILImage.open(img_ext) as im_ext:
                    w, h = im_ext.size
                r = min(14*cm/w, 10.5*cm/h)
                img_portada = RLImage(img_ext, width=w*r, height=h*r)
                img_portada.hAlign = 'CENTER'
                story.append(img_portada)
                story.append(Spacer(1, 0.5*cm))
            except Exception:
                pass
        
        empresa_pdf = get_empresa()
        cif_pdf = get_cif()
        if empresa_pdf:
            story.append(Paragraph(f"<b>Empresa:</b> {_escapar_pdf(empresa_pdf)}", centrado))
        if cif_pdf:
            story.append(Paragraph(f"<b>CIF:</b> {_escapar_pdf(cif_pdf)}", centrado))
        if tecnico:
            story.append(Paragraph(f"<b>Técnico:</b> {_escapar_pdf(tecnico)}", centrado))
        story.append(Paragraph(f"<b>Fecha:</b> {_escapar_pdf(fecha) or '-'}", centrado))
        story.append(Paragraph(f"<b>Detectores:</b> {len(detectores)}", centrado))
        story.append(Spacer(1, 0.5*cm))

        categorias_centro_pdf = fetch_categorias_centro(centro_id)
        if categorias_centro_pdf:
            story.append(Paragraph("<b>Categorías profesionales</b>", centrado))
            story.append(Spacer(1, 0.2*cm))
            filas_cat = [["Categoría profesional", "Nº personas expuestas"]]
            for _, _, categoria_pdf, num_personas_pdf in categorias_centro_pdf:
                filas_cat.append([_escapar_pdf(categoria_pdf), str(num_personas_pdf)])
            tabla_cat = Table(filas_cat, colWidths=[9*cm, 6*cm])
            tabla_cat.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F5A623")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]))
            story.append(tabla_cat)
            story.append(Spacer(1, 0.5*cm))
        
        for idx, d in enumerate(detectores, 1):
            (did, _, planta, sala, fecha_det, codigo, _plano_antiguo, px, py, foto_sit, foto_det, _,
             codigo_sala, profesionales_sala, hora_colocacion, turno_trabajo, nivel, plano_centro_id, fecha_retirada_real, hora_retirada_real) = d
            plano = None
            nombre_plano_actual = None
            if plano_centro_id:
                plano_info = get_plano_centro(plano_centro_id)
                if plano_info:
                    nombre_plano_actual = plano_info[2]
                    plano = plano_info[3]
            story.append(PageBreak())
            
            titulo_partes = [codigo or "-", nombre or "-"]
            if zona:
                titulo_partes.append(zona)
            titulo_detector = f"Detector {idx}: " + " - ".join(titulo_partes)
            titulo_detector = _escapar_pdf(titulo_detector)
            estilo_titulo_detector = ParagraphStyle(
                'TituloDetector', parent=styles['Heading2'], fontSize=13,
                spaceBefore=0, spaceAfter=6,
            )
            story.append(Paragraph(titulo_detector, estilo_titulo_detector))
            
            fecha_retirada_optima_pdf = calcular_fecha_retirada_optima(fecha_det)

            # Las celdas se envuelven en Paragraph para que el texto largo
            # (p.ej. varios profesionales, o "Mañana + tarde + noche") se
            # reparta en varias líneas en vez de desbordar la celda; la
            # fila crece de alto automáticamente según haga falta.
            estilo_label_celda = ParagraphStyle(
                'CeldaLabel', parent=styles['Normal'], fontSize=9, leading=11, fontName='Helvetica-Bold',
            )
            estilo_valor_celda = ParagraphStyle(
                'CeldaValor', parent=styles['Normal'], fontSize=9, leading=11,
            )

            fecha_y_hora_colocacion = _escapar_pdf(fecha_det) or "-"
            if hora_colocacion:
                fecha_y_hora_colocacion += f"<br/>{_escapar_pdf(hora_colocacion)}"

            fecha_y_hora_retirada_real = ""
            if fecha_retirada_real:
                fecha_y_hora_retirada_real = _escapar_pdf(fecha_retirada_real)
                if hora_retirada_real:
                    fecha_y_hora_retirada_real += f"<br/>{_escapar_pdf(hora_retirada_real)}"

            filas_texto = [
                ["Codigo del detector", _escapar_pdf(codigo) or "-", "Fecha y hora de colocacion", fecha_y_hora_colocacion],
                ["Planta", _escapar_pdf(planta) or "-", "Nivel", _escapar_pdf(nivel) or "-"],
                ["Sala", _escapar_pdf(sala) or "-", "Codigo de la sala", _escapar_pdf(codigo_sala) or "-"],
                ["Profesionales en la sala", _escapar_pdf(profesionales_sala) or "-", "Turno de trabajo", _escapar_pdf(turno_trabajo) or "-"],
                ["Fecha optima retirada", _escapar_pdf(fecha_retirada_optima_pdf) or "-", "Fecha y hora real de retirada", fecha_y_hora_retirada_real],
            ]
            filas_pdf = []
            for fila_txt in filas_texto:
                fila_pdf = []
                for i, val in enumerate(fila_txt):
                    if not val:
                        fila_pdf.append("")
                    else:
                        estilo = estilo_label_celda if i % 2 == 0 else estilo_valor_celda
                        fila_pdf.append(Paragraph(val, estilo))
                filas_pdf.append(fila_pdf)

            tabla = Table(filas_pdf, colWidths=[3.3*cm, 4.7*cm, 3.3*cm, 4.7*cm])
            tabla.setStyle(TableStyle([
                ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
                ("BACKGROUND", (0,0), (0,-1), colors.whitesmoke),
                ("BACKGROUND", (2,0), (2,-1), colors.whitesmoke),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
            ]))
            story.append(tabla)
            story.append(Spacer(1, 0.2*cm))

            estilo_subtitulo = ParagraphStyle(
                'SubtituloCompacto', parent=styles['Normal'], fontSize=10,
                fontName="Helvetica-Bold", spaceBefore=2, spaceAfter=2,
            )

            if plano and os.path.exists(plano):
                nombre_plano_esc = _escapar_pdf(nombre_plano_actual)
                titulo_ubicacion = f"Ubicacion en el plano ({nombre_plano_esc}):" if nombre_plano_esc else "Ubicacion en el plano:"
                story.append(Paragraph(titulo_ubicacion, estilo_subtitulo))
                try:
                    with PILImage.open(plano) as im_plano:
                        im_plano = im_plano.convert("RGB")
                        w, h = im_plano.size
                        if px is not None and py is not None and px >= 0 and py >= 0:
                            draw = ImageDraw.Draw(im_plano)
                            cx, cy = px * w, py * h
                            r = max(6, int(min(w, h) * 0.012))
                            draw.ellipse([cx - r, cy - r, cx + r, cy + r],
                                         fill=(220, 20, 20), outline=(120, 0, 0), width=2)
                        tmp_plano_path = os.path.join(get_data_dir(), f"_tmp_plano_{did}.jpg")
                        im_plano.save(tmp_plano_path, quality=90)
                    r = min(16.5*cm/w, 8.5*cm/h)
                    story.append(RLImage(tmp_plano_path, width=w*r, height=h*r))
                    story.append(Spacer(1, 0.2*cm))
                except Exception:
                    pass
            
            if foto_sit and os.path.exists(foto_sit) and foto_det and os.path.exists(foto_det):
                story.append(Paragraph("Fotos:", estilo_subtitulo))
                titulo_foto_style = ParagraphStyle(
                    'TituloFoto', parent=styles['Normal'], alignment=TA_CENTER, fontName="Helvetica-Bold",
                )
                try:
                    with PILImage.open(foto_sit) as im:
                        w, h = im.size
                    r = min(8.2*cm/w, 7.5*cm/h)
                    img1 = RLImage(foto_sit, width=w*r, height=h*r)
                except:
                    img1 = Paragraph("(no disponible)", styles["Normal"])
                try:
                    with PILImage.open(foto_det) as im:
                        w, h = im.size
                    r = min(8.2*cm/w, 7.5*cm/h)
                    img2 = RLImage(foto_det, width=w*r, height=h*r)
                except:
                    img2 = Paragraph("(no disponible)", styles["Normal"])
                cap1 = Paragraph("Situación del detector", titulo_foto_style)
                cap2 = Paragraph("Detector", titulo_foto_style)
                story.append(Table([[cap1, cap2], [img1, img2]], colWidths=[8.5*cm, 8.5*cm]))
        
        doc.build(story, onFirstPage=_dibujar_cabecera, onLaterPages=_dibujar_cabecera)
        return True
    except Exception as e:
        raise Exception(f"Error: {str(e)}")


# ============================================================
# GENERAR EXCEL (hoja de cálculo del centro, con las fotos de cada
# detector incrustadas cada una en su propia celda)
# ============================================================

# ============================================================
# CÁLCULO DE LA FECHA DE RETIRADA ÓPTIMA (90 días laborables desde la
# colocación, saltando fines de semana y festivos de Galicia)
# ============================================================

def _domingo_de_pascua(anio):
    """Algoritmo de Gauss/Meeus para calcular el Domingo de Pascua de
    un año dado (necesario para Jueves y Viernes Santo, festivos
    movibles que dependen de la Pascua cada año)."""
    a = anio % 19
    b = anio // 100
    c = anio % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes = (h + l - 7 * m + 114) // 31
    dia = ((h + l - 7 * m + 114) % 31) + 1
    return date(anio, mes, dia)


def _festivos_galicia(anio):
    """Festivos de ámbito estatal/autonómico aplicables en Galicia para
    un año dado (no incluye fiestas locales de cada ayuntamiento, que
    varían por municipio)."""
    pascua = _domingo_de_pascua(anio)
    jueves_santo = pascua - timedelta(days=3)
    viernes_santo = pascua - timedelta(days=2)
    fijos = [
        date(anio, 1, 1),    # Año Nuevo
        date(anio, 1, 6),    # Reyes
        date(anio, 5, 1),    # Fiesta del Trabajo
        date(anio, 7, 25),   # Santiago Apóstol (patrón de Galicia)
        date(anio, 8, 15),   # Asunción
        date(anio, 10, 12),  # Fiesta Nacional de España
        date(anio, 11, 1),   # Todos los Santos
        date(anio, 12, 6),   # Día de la Constitución
        date(anio, 12, 8),   # Inmaculada Concepción
        date(anio, 12, 25),  # Navidad
    ]
    return set(fijos + [jueves_santo, viernes_santo])


def _parsear_fecha(texto):
    """Intenta interpretar una fecha escrita a mano en varios formatos
    habituales. Devuelve un date, o None si no se pudo interpretar."""
    texto = (texto or "").strip()
    if not texto:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def calcular_fecha_retirada_optima(fecha_colocacion_texto):
    """90 días naturales desde la fecha de colocación; si ese día cae
    en sábado, domingo o festivo de Galicia, se pasa al siguiente día
    laborable. Devuelve el resultado como texto "DD/MM/AAAA", o ""
    si la fecha de colocación no se pudo interpretar."""
    fecha_col = _parsear_fecha(fecha_colocacion_texto)
    if fecha_col is None:
        return ""
    objetivo = fecha_col + timedelta(days=90)
    cache_festivos = {}
    while True:
        if objetivo.year not in cache_festivos:
            cache_festivos[objetivo.year] = _festivos_galicia(objetivo.year)
        if objetivo.weekday() >= 5 or objetivo in cache_festivos[objetivo.year]:
            objetivo += timedelta(days=1)
            continue
        break
    return objetivo.strftime("%d/%m/%Y")


def importar_centro_desde_excel(archivo_bytes):
    """Reconstruye un centro completo (datos, planos y detectores, con
    el punto exacto de cada uno sobre su plano) a partir de un Excel
    generado por generar_excel.

    Devuelve (centro_id, numero_de_detectores_creados). Lanza
    ValueError con un mensaje claro si el archivo no tiene la
    estructura esperada (no es un Excel exportado por esta app).
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(archivo_bytes))
    if "Detectores" not in wb.sheetnames:
        raise ValueError(
            "El archivo no parece un Excel exportado por esta app "
            "(falta la hoja 'Detectores')."
        )
    ws = wb["Detectores"]

    nombre_centro = ws.cell(row=1, column=1).value or "Centro importado"
    area_centro = ws.cell(row=1, column=5).value or ""
    tecnico_centro = ws.cell(row=1, column=10).value or ""
    fecha_centro = ws.cell(row=1, column=15).value or ""
    direccion_centro = ws.cell(row=1, column=17).value or ""

    header_row = 2
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[val] = col

    columnas_esperadas = ["Centro", "Área", "Planta", "Sala", "Código"]
    faltan = [c for c in columnas_esperadas if c not in headers]
    if faltan:
        raise ValueError(
            "El archivo no tiene el formato esperado (faltan columnas: "
            + ", ".join(faltan) + ")."
        )

    # Imágenes incrustadas de la hoja "Detectores", indexadas por la
    # posición exacta de celda (fila0, col0) en la que se anclaron al
    # generarlas, para poder recuperar la foto de cada detector.
    imagenes_por_celda = {}
    for img in ws._images:
        try:
            fr = img.anchor._from
            imagenes_por_celda[(fr.row, fr.col)] = img._data()
        except Exception:
            continue

    centro_id = crear_centro(str(nombre_centro).strip() or "Centro importado")
    imagen_exterior_path = None

    # --- Hoja "Planos": crea TODOS los planos del centro (incluidos los
    # que ningún detector tenga asignado) y construye el mapa
    # nombre -> id para poder vincular cada detector a su plano.
    # También trae la Empresa y el CIF (guardados en esta misma hoja,
    # en sus dos primeras filas). ---
    planos_por_nombre = {}
    if "Planos" in wb.sheetnames:
        ws_p = wb["Planos"]

        empresa_importada = ws_p.cell(row=1, column=2).value
        cif_importado = ws_p.cell(row=2, column=2).value
        if empresa_importada:
            set_empresa(str(empresa_importada).strip())
        if cif_importado:
            set_cif(str(cif_importado).strip())

        imagenes_planos_por_celda = {}
        for img in ws_p._images:
            try:
                fr = img.anchor._from
                imagenes_planos_por_celda[(fr.row, fr.col)] = img._data()
            except Exception:
                continue
        orden = 0
        # Las filas 1-4 son Empresa, CIF, una fila en blanco y la
        # cabecera "Nombre"/"Imagen"; los planos empiezan en la 5.
        for fila_p in range(5, ws_p.max_row + 1):
            nombre_p = ws_p.cell(row=fila_p, column=1).value
            if not nombre_p:
                continue
            datos_img = imagenes_planos_por_celda.get((fila_p - 1, 1))
            if not datos_img:
                continue
            ruta_guardada = guardar_bytes_imagen(datos_img, f"plano_centro_{centro_id}")
            if str(nombre_p) == "(Foto exterior del centro)":
                imagen_exterior_path = ruta_guardada
            else:
                nuevo_id = insert_plano_centro(centro_id, str(nombre_p), ruta_guardada, orden)
                planos_por_nombre[str(nombre_p)] = nuevo_id
                orden += 1

    update_centro(
        centro_id, str(nombre_centro).strip() or "Centro importado",
        str(area_centro).strip(), str(fecha_centro).strip(), imagen_exterior_path,
        str(direccion_centro).strip(),
    )
    if tecnico_centro:
        set_tecnico_centro(centro_id, str(tecnico_centro).strip())

    # --- Hoja "Categorías profesionales" (si existe) ---
    if "Categorías profesionales" in wb.sheetnames:
        ws_cat = wb["Categorías profesionales"]
        for fila_cat in range(2, ws_cat.max_row + 1):
            categoria_val = ws_cat.cell(row=fila_cat, column=1).value
            num_val = ws_cat.cell(row=fila_cat, column=2).value
            if categoria_val:
                try:
                    num_final = int(num_val) if num_val is not None else 0
                except (TypeError, ValueError):
                    num_final = 0
                insert_categoria_centro(centro_id, str(categoria_val).strip(), num_final)

    def _val(fila, nombre_col, default=""):
        col = headers.get(nombre_col)
        if not col:
            return default
        v = ws.cell(row=fila, column=col).value
        return v if v is not None else default

    detectores_creados = 0
    for fila_actual in range(header_row + 1, ws.max_row + 1):
        codigo_det = _val(fila_actual, "Código")
        sala_det = _val(fila_actual, "Sala")
        if not str(codigo_det).strip() and not str(sala_det).strip():
            continue  # fila vacía, se ignora

        planta = _val(fila_actual, "Planta")
        nivel = _val(fila_actual, "Nivel")
        codigo_sala = _val(fila_actual, "Código de la sala")
        profesionales_multilinea = str(_val(fila_actual, "Profesionales en la sala"))
        profesionales = ", ".join(
            linea.strip() for linea in profesionales_multilinea.split("\n") if linea.strip()
        )
        turno = _val(fila_actual, "Turno de trabajo")
        fecha_colocacion = _val(fila_actual, "Fecha de colocación")
        hora_colocacion = _val(fila_actual, "Hora de colocación")
        fecha_retirada_real = _val(fila_actual, "Fecha de retirada real")
        hora_retirada_real = _val(fila_actual, "Hora de retirada real")
        nombre_plano_fila = _val(fila_actual, "Nombre del plano")
        punto_x_fila = _val(fila_actual, "Punto X", None)
        punto_y_fila = _val(fila_actual, "Punto Y", None)

        plano_centro_id = planos_por_nombre.get(str(nombre_plano_fila)) if nombre_plano_fila else None

        try:
            punto_x_final = float(punto_x_fila) if punto_x_fila not in (None, "") else -1
            punto_y_final = float(punto_y_fila) if punto_y_fila not in (None, "") else -1
        except (TypeError, ValueError):
            punto_x_final = punto_y_final = -1

        foto_sit_path = None
        foto_det_path = None
        col_situacion = headers.get("Foto situación")
        col_detector_foto = headers.get("Foto detector")
        if col_situacion:
            datos = imagenes_por_celda.get((fila_actual - 1, col_situacion - 1))
            if datos:
                foto_sit_path = guardar_bytes_imagen(datos, "foto_situacion")
        if col_detector_foto:
            datos = imagenes_por_celda.get((fila_actual - 1, col_detector_foto - 1))
            if datos:
                foto_det_path = guardar_bytes_imagen(datos, "foto_detector")

        data = (
            centro_id, str(planta), str(sala_det), str(fecha_colocacion), str(codigo_det),
            None, punto_x_final, punto_y_final, foto_sit_path, foto_det_path,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            str(codigo_sala), profesionales, str(hora_colocacion),
            str(turno) if turno else TURNOS_TRABAJO_OPCIONES[0],
            str(nivel) if nivel else NIVEL_OPCIONES[0],
            plano_centro_id, str(fecha_retirada_real), str(hora_retirada_real),
        )
        insert_detector(data)
        detectores_creados += 1

    return centro_id, detectores_creados


def generar_excel(centro_id, output_path):
    """Genera un .xlsx en formato de tabla "plana" (sin ningún bloque de
    cabecera encima), pensado para poder ir pegando debajo los datos de
    otros centros más adelante. El nombre del centro y el área figuran
    como las dos primeras columnas de cada fila, junto con todos los
    datos de cada detector, la fecha de retirada óptima calculada
    automáticamente, una columna vacía para anotar la retirada real, y
    sus tres fotos (plano con el punto marcado, situación, detector)
    incrustadas cada una en su propia celda, centradas y manteniendo su
    proporción real."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    centro = get_centro(centro_id)
    if not centro:
        raise Exception("Centro no encontrado")
    cid, nombre, zona, fecha, img_centro, tecnico, direccion = centro
    detectores = fetch_detectores(centro_id)

    FUENTE = "Arial"
    fuente_normal = Font(name=FUENTE, size=7)
    fuente_cabecera = Font(name=FUENTE, size=8, bold=True, color="FFFFFF")
    fuente_info_grande = Font(name=FUENTE, size=12, bold=True, color="FFFFFF")
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Estilo especial para la primera columna ("Centro"): sin fondo
    # naranja, alineado a la derecha, y con el texto en naranja en vez
    # de negro/blanco.
    alineado_derecha = Alignment(horizontal="right", vertical="center", wrap_text=True)
    fuente_naranja_cabecera = Font(name=FUENTE, size=8, bold=True, color="F5A623")
    fuente_naranja_dato = Font(name=FUENTE, size=7, color="F5A623")

    wb = Workbook()
    ws = wb.active
    ws.title = "Detectores"

    borde_fino = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    # --- Línea de cabecera superior con el resumen del centro (Centro,
    # Área, Día, Técnico), y debajo la fila de cabeceras de columna
    # (al doble de alto de lo normal). ---
    fila_info = 1
    header_row = 2
    headers = [
        "Centro", "Área", "ID", "Planta", "Nivel", "Sala", "Código de la sala",
        "Profesionales en la sala", "Turno de trabajo", "Código",
        "Resultado (Bq/m³/h)", "Incertidumbre",
        "Fecha de colocación", "Hora de colocación",
        "Fecha de retirada óptima", "Fecha de retirada real", "Hora de retirada real",
        "Nombre del plano", "Punto X", "Punto Y",
        "Plano", "Foto situación", "Foto detector",
    ]

    col_letra = {h: get_column_letter(i) for i, h in enumerate(headers, start=1)}
    col_idx0 = {h: i for i, h in enumerate(headers)}  # índice base-0, para los anclajes

    # Línea superior en forma de banner: el nombre del centro ocupa las
    # 4 primeras columnas fusionadas, el área las 4 siguientes, luego
    # una columna con la palabra "Técnico" y las 4 siguientes con su
    # nombre fusionadas. El resto de columnas quedan en naranja liso
    # para que la fila entera se vea como un único banner.
    ws.merge_cells(start_row=fila_info, start_column=1, end_row=fila_info, end_column=4)
    c_centro = ws.cell(row=fila_info, column=1, value=nombre or "")
    c_centro.font = fuente_info_grande
    c_centro.fill = PatternFill("solid", fgColor="F5A623")
    c_centro.alignment = centrado

    ws.merge_cells(start_row=fila_info, start_column=5, end_row=fila_info, end_column=8)
    c_area = ws.cell(row=fila_info, column=5, value=zona or "")
    c_area.font = fuente_info_grande
    c_area.fill = PatternFill("solid", fgColor="F5A623")
    c_area.alignment = centrado

    c_tecnico_label = ws.cell(row=fila_info, column=9, value="Técnico")
    c_tecnico_label.font = fuente_cabecera
    c_tecnico_label.fill = PatternFill("solid", fgColor="F5A623")
    c_tecnico_label.alignment = centrado

    ws.merge_cells(start_row=fila_info, start_column=10, end_row=fila_info, end_column=13)
    c_tecnico_val = ws.cell(row=fila_info, column=10, value=tecnico or "")
    c_tecnico_val.font = fuente_cabecera
    c_tecnico_val.fill = PatternFill("solid", fgColor="F5A623")
    c_tecnico_val.alignment = centrado

    # Fecha del centro: no se muestra a gran tamaño como Centro/Área,
    # pero se guarda en el banner para poder reconstruirla al importar.
    c_fecha_label = ws.cell(row=fila_info, column=14, value="Fecha")
    c_fecha_label.font = fuente_cabecera
    c_fecha_label.fill = PatternFill("solid", fgColor="F5A623")
    c_fecha_label.alignment = centrado
    c_fecha_val = ws.cell(row=fila_info, column=15, value=fecha or "")
    c_fecha_val.font = fuente_cabecera
    c_fecha_val.fill = PatternFill("solid", fgColor="F5A623")
    c_fecha_val.alignment = centrado

    # Dirección del centro: igual que la fecha, se guarda en el banner
    # (aunque no se muestre a gran tamaño) para poder recuperarla si se
    # reimporta este Excel más adelante.
    c_dir_label = ws.cell(row=fila_info, column=16, value="Dirección")
    c_dir_label.font = fuente_cabecera
    c_dir_label.fill = PatternFill("solid", fgColor="F5A623")
    c_dir_label.alignment = centrado
    c_dir_val = ws.cell(row=fila_info, column=17, value=direccion or "")
    c_dir_val.font = fuente_cabecera
    c_dir_val.fill = PatternFill("solid", fgColor="F5A623")
    c_dir_val.alignment = centrado

    # El resto de la fila (hasta donde llegan las columnas de la tabla
    # de abajo) también en naranja con texto centrado, aunque no haya
    # más datos que mostrar, para que quede como un único banner.
    for col in range(18, len(headers) + 1):
        c_resto = ws.cell(row=fila_info, column=col)
        c_resto.fill = PatternFill("solid", fgColor="F5A623")
        c_resto.alignment = centrado
        c_resto.font = fuente_cabecera
    ws.row_dimensions[fila_info].height = 22

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        if i == 1:
            # "Centro": sin relleno naranja, texto naranja, a la derecha.
            c.font = fuente_naranja_cabecera
            c.alignment = alineado_derecha
        else:
            c.font = fuente_cabecera
            c.fill = PatternFill("solid", fgColor="F5A623")
            c.alignment = centrado
        c.border = borde_fino
    ws.row_dimensions[header_row].height = 30  # el doble de una fila normal (~15pt)

    ANCHO_FOTOS = 24
    for h in ("Plano", "Foto situación", "Foto detector"):
        ws.column_dimensions[col_letra[h]].width = ANCHO_FOTOS
    for h in ("Centro", "Planta", "Nivel", "Sala", "Código de la sala", "Profesionales en la sala",
              "Turno de trabajo", "Código"):
        ws.column_dimensions[col_letra[h]].width = 15
    for h in ("Fecha de colocación", "Fecha de retirada óptima", "Fecha de retirada real"):
        ws.column_dimensions[col_letra[h]].width = 16
    for h in ("Hora de colocación", "Hora de retirada real"):
        ws.column_dimensions[col_letra[h]].width = 10
    ws.column_dimensions[col_letra["Área"]].width = 13
    ws.column_dimensions[col_letra["Nombre del plano"]].width = 14
    for h in ("Punto X", "Punto Y"):
        ws.column_dimensions[col_letra[h]].width = 8

    TAM_IMG_PX = 150  # tamaño MÁXIMO (lado mayor) de cada foto dentro de su celda
    ALTO_FILA_PT = 115  # alto de fila (puntos) reservado para las fotos

    def _ancho_col_px(col_chars):
        # Aproximación estándar: píxeles ≈ caracteres * 7 + 5
        return round(col_chars * 7 + 5)

    def _alto_fila_px(alto_pt):
        # 1 punto ≈ 1.333 píxeles (96 dpi)
        return round(alto_pt * 96 / 72)

    def _anclaje_centrado(col_h, fila1, ancho_img_px, alto_img_px):
        """Ancla la imagen centrada (horizontal y verticalmente) dentro
        de su celda, calculando el hueco libre a cada lado."""
        ancho_celda_px = _ancho_col_px(ws.column_dimensions[col_letra[col_h]].width)
        alto_celda_px = _alto_fila_px(ALTO_FILA_PT)
        off_x = max(0, (ancho_celda_px - ancho_img_px) / 2)
        off_y = max(0, (alto_celda_px - alto_img_px) / 2)
        marker = AnchorMarker(
            col=col_idx0[col_h], row=fila1 - 1,
            colOff=pixels_to_EMU(off_x), rowOff=pixels_to_EMU(off_y),
        )
        size = XDRPositiveSize2D(cx=pixels_to_EMU(ancho_img_px), cy=pixels_to_EMU(alto_img_px))
        return OneCellAnchor(_from=marker, ext=size)

    fila = header_row + 1
    for d in detectores:
        (did, _, planta, sala, fecha_det, codigo, _plano_antiguo, punto_x, punto_y,
         foto_sit_p, foto_det_p, fecha_creacion, codigo_sala, profesionales_sala,
         hora_colocacion, turno_trabajo, nivel, plano_centro_id, fecha_retirada_real, hora_retirada_real) = d
        plano_p = None
        nombre_plano_d = ""
        if plano_centro_id:
            plano_info = get_plano_centro(plano_centro_id)
            if plano_info:
                plano_p = plano_info[3]
                nombre_plano_d = plano_info[2]

        fecha_retirada_optima = calcular_fecha_retirada_optima(fecha_det)

        # Los profesionales se escriben separados por comas en el
        # formulario; en la celda del Excel se muestran uno debajo de
        # otro (salto de línea + texto ajustado a la celda).
        profesionales_multilinea = "\n".join(
            p.strip() for p in (profesionales_sala or "").split(",") if p.strip()
        )

        hay_punto_valido = (
            punto_x is not None and punto_y is not None and 0 <= punto_x <= 1 and 0 <= punto_y <= 1
        )

        valores = [
            nombre or "", zona or "", did, planta or "", nivel or "", sala or "",
            codigo_sala or "", profesionales_multilinea, turno_trabajo or "", codigo or "",
            "", "",
            fecha_det or "", hora_colocacion or "",
            fecha_retirada_optima, fecha_retirada_real or "", hora_retirada_real or "",
            nombre_plano_d, round(punto_x, 4) if hay_punto_valido else "",
            round(punto_y, 4) if hay_punto_valido else "",
        ]
        for col, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=val)
            if col == 1:
                # "Centro": sin relleno naranja, texto naranja, a la derecha.
                c.font = fuente_naranja_dato
                c.alignment = alineado_derecha
            else:
                c.font = fuente_normal
                c.alignment = centrado
            c.border = borde_fino
        # Las tres celdas de fotos también con borde, aunque queden vacías
        for h in ("Plano", "Foto situación", "Foto detector"):
            ws[f"{col_letra[h]}{fila}"].border = borde_fino

        ws.row_dimensions[fila].height = ALTO_FILA_PT

        hay_punto = hay_punto_valido

        for col_name, ruta in (
            ("Plano", plano_p),
            ("Foto situación", foto_sit_p),
            ("Foto detector", foto_det_p),
        ):
            if ruta and os.path.exists(ruta):
                try:
                    with Image.open(ruta) as im_orig:
                        # IMPORTANTE: las fotos de móvil llevan a menudo
                        # metadatos EXIF de rotación que PIL no aplica
                        # solo; sin esto, el ancho/alto "en bruto" no
                        # coincide con cómo se ve realmente la foto y la
                        # proporción calculada sale mal.
                        im = ImageOps.exif_transpose(im_orig)
                        im = im.convert("RGB")
                        if col_name == "Plano" and hay_punto:
                            # Dibujar el punto rojo del detector sobre el
                            # plano, igual que en el informe PDF.
                            draw = ImageDraw.Draw(im)
                            w, h = im.size
                            cx, cy = punto_x * w, punto_y * h
                            r = max(6, int(min(w, h) * 0.02))
                            draw.ellipse(
                                [cx - r, cy - r, cx + r, cy + r],
                                fill=(220, 20, 20), outline=(120, 0, 0), width=2,
                            )
                        # La imagen se ve como mucho a ~150px en la
                        # celda, así que guardarla mucho más grande solo
                        # engorda el archivo (y puede hacer fallar el
                        # envío por WhatsApp por límite de tamaño).
                        im.thumbnail((350, 350))
                        buf = io.BytesIO()
                        im.save(buf, format="JPEG", quality=78)
                        buf.seek(0)
                        ancho_real, alto_real = im.size

                    # Escalar manteniendo la proporción real (sin
                    # deformar), ajustando al lado mayor = TAM_IMG_PX.
                    escala = min(TAM_IMG_PX / ancho_real, TAM_IMG_PX / alto_real)
                    ancho_final = max(1, round(ancho_real * escala))
                    alto_final = max(1, round(alto_real * escala))
                    xl_img = XLImage(buf)
                    xl_img.width = ancho_final
                    xl_img.height = alto_final
                    xl_img.anchor = _anclaje_centrado(col_name, fila, ancho_final, alto_final)
                    ws.add_image(xl_img)
                except Exception:
                    c_err = ws[f"{col_letra[col_name]}{fila}"]
                    c_err.value = "(no se pudo incrustar la imagen)"
                    c_err.font = fuente_normal
                    c_err.alignment = centrado
        fila += 1

    # --- Hoja "Planos": TODOS los planos del centro, uno por fila,
    # tengan o no algún detector que los use (los del "Detectores" solo
    # aparecen si algún detector los tiene asignado; esta hoja evita
    # que un plano "huérfano" se pierda al reimportar). También incluye
    # la foto exterior del centro, que no aparece en ningún otro sitio
    # del Excel. ---
    ws_planos = wb.create_sheet("Planos")

    # Empresa y CIF (datos globales de la empresa que realiza las
    # mediciones), en las dos primeras filas de esta hoja.
    empresa_xl = get_empresa()
    cif_xl = get_cif()
    ws_planos.cell(row=1, column=1, value="Empresa").font = fuente_cabecera
    ws_planos.cell(row=1, column=1).fill = PatternFill("solid", fgColor="F5A623")
    ws_planos.cell(row=1, column=2, value=empresa_xl).font = fuente_normal
    ws_planos.cell(row=2, column=1, value="CIF").font = fuente_cabecera
    ws_planos.cell(row=2, column=1).fill = PatternFill("solid", fgColor="F5A623")
    ws_planos.cell(row=2, column=2, value=cif_xl).font = fuente_normal

    header_row_planos = 4
    ws_planos.cell(row=header_row_planos, column=1, value="Nombre").font = fuente_cabecera
    ws_planos.cell(row=header_row_planos, column=1).fill = PatternFill("solid", fgColor="F5A623")
    ws_planos.cell(row=header_row_planos, column=2, value="Imagen").font = fuente_cabecera
    ws_planos.cell(row=header_row_planos, column=2).fill = PatternFill("solid", fgColor="F5A623")
    ws_planos.column_dimensions["A"].width = 20
    ws_planos.column_dimensions["B"].width = 28

    fila_planos = header_row_planos + 1

    def _incrustar_en_hoja_planos(nombre_fila, ruta_imagen, fila_actual):
        ws_planos.cell(row=fila_actual, column=1, value=nombre_fila).font = fuente_normal
        ws_planos.row_dimensions[fila_actual].height = 115
        if ruta_imagen and os.path.exists(ruta_imagen):
            try:
                with Image.open(ruta_imagen) as im_orig:
                    im = ImageOps.exif_transpose(im_orig)
                    im = im.convert("RGB")
                    im.thumbnail((500, 500))
                    buf = io.BytesIO()
                    im.save(buf, format="JPEG", quality=82)
                    buf.seek(0)
                    ancho_real, alto_real = im.size
                escala = min(180 / ancho_real, 150 / alto_real)
                xl_img = XLImage(buf)
                xl_img.width = max(1, round(ancho_real * escala))
                xl_img.height = max(1, round(alto_real * escala))
                ws_planos.add_image(xl_img, f"B{fila_actual}")
            except Exception:
                ws_planos.cell(row=fila_actual, column=2, value="(no se pudo incrustar la imagen)")

    if img_centro and os.path.exists(img_centro):
        _incrustar_en_hoja_planos("(Foto exterior del centro)", img_centro, fila_planos)
        fila_planos += 1

    for plano_c in fetch_planos_centro(centro_id):
        _, _, nombre_plano_c, ruta_plano_c, _ = plano_c
        _incrustar_en_hoja_planos(nombre_plano_c, ruta_plano_c, fila_planos)
        fila_planos += 1

    # --- Hoja "Categorías profesionales": una fila por categoría con
    # el número de personas expuestas de esa categoría en el centro. ---
    categorias_centro_xl = fetch_categorias_centro(centro_id)
    if categorias_centro_xl:
        ws_cat = wb.create_sheet("Categorías profesionales")
        ws_cat.cell(row=1, column=1, value="Categoría profesional").font = fuente_cabecera
        ws_cat.cell(row=1, column=1).fill = PatternFill("solid", fgColor="F5A623")
        ws_cat.cell(row=1, column=2, value="Nº personas expuestas").font = fuente_cabecera
        ws_cat.cell(row=1, column=2).fill = PatternFill("solid", fgColor="F5A623")
        ws_cat.column_dimensions["A"].width = 28
        ws_cat.column_dimensions["B"].width = 20
        for i, (_, _, categoria_xl, num_personas_xl) in enumerate(categorias_centro_xl, start=2):
            ws_cat.cell(row=i, column=1, value=categoria_xl).font = fuente_normal
            ws_cat.cell(row=i, column=2, value=num_personas_xl).font = fuente_normal

    wb.save(output_path)
    return True


# ============================================================
# REGISTRO PARA LABORATORIO
# Ficha calcada del formulario "FICHA DE IDENTIFICACIÓN E INFORMACIÓN
# DE LOS DETECTORES DE TRAZAS" que exige el laboratorio de análisis,
# para que la acepten sin tener que rellenarla a mano aparte.
# ============================================================

def generar_registro_laboratorio(centro_id, output_path, tipo_firma="digital"):
    """Genera la ficha de identificación de detectores en el formato
    que exige el laboratorio, con una tabla prácticamente idéntica a
    su propio formulario en papel, más un cuadro de firma del técnico
    (digital o a mano) fuera de la tabla."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors

    centro = get_centro(centro_id)
    if not centro:
        raise ValueError("Centro no encontrado")
    _, nombre, zona, fecha, img_ext, tecnico, direccion = centro
    detectores = fetch_detectores(centro_id)
    empresa = get_empresa()

    doc = SimpleDocTemplate(
        output_path, pagesize=landscape(A4),
        topMargin=1.2*cm, bottomMargin=1.2*cm, leftMargin=1.2*cm, rightMargin=1.2*cm,
    )
    styles = getSampleStyleSheet()
    story = []

    estilo_celda = ParagraphStyle(
        "CeldaLab", parent=styles["Normal"], fontSize=7, leading=8.5, alignment=TA_CENTER,
    )
    estilo_celda_izq = ParagraphStyle(
        "CeldaLabIzq", parent=styles["Normal"], fontSize=7, leading=8.5, alignment=TA_LEFT,
    )
    estilo_cab = ParagraphStyle(
        "CabLab", parent=styles["Normal"], fontSize=7, leading=8.5, alignment=TA_CENTER,
        textColor=colors.white, fontName="Helvetica-Bold",
    )

    # --- Cabecera: datos de la empresa que realiza la medición, a la
    # derecha, imitando el bloque de logo/contacto del laboratorio ---
    cab_izq = Paragraph(
        "<b>FICHA DE IDENTIFICACIÓN E INFORMACIÓN DE LOS DETECTORES DE TRAZAS</b>",
        ParagraphStyle("TituloLab", parent=styles["Normal"], fontSize=13, leading=16),
    )
    cab_der = Paragraph(
        f"<b>{_escapar_pdf(empresa or 'Empresa de medición')}</b>",
        ParagraphStyle("EmpresaLab", parent=styles["Normal"], fontSize=10, alignment=TA_CENTER),
    )
    tabla_cab = Table([[cab_izq, cab_der]], colWidths=[19*cm, 6.7*cm])
    tabla_cab.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (1, 0), (1, 0), 1, colors.black),
        ("BOX", (0, 0), (0, 0), 1.2, colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(tabla_cab)
    story.append(Spacer(1, 0.25*cm))

    # --- Datos de la obra / inmueble / cliente ---
    filas_datos = [
        [Paragraph("<b>Ref. de Obra:</b>", estilo_celda_izq), Paragraph("", estilo_celda_izq),
         Paragraph("<b>Orden de trabajo:</b>", estilo_celda_izq), Paragraph("", estilo_celda_izq)],
        [Paragraph("<b>Inmueble:</b>", estilo_celda_izq),
         Paragraph(_escapar_pdf(nombre or ""), estilo_celda_izq),
         Paragraph("<b>Cliente:</b>", estilo_celda_izq),
         Paragraph(_escapar_pdf(empresa or ""), estilo_celda_izq)],
    ]
    tabla_datos = Table(filas_datos, colWidths=[2.8*cm, 9.7*cm, 3.2*cm, 9.9*cm])
    tabla_datos.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(tabla_datos)

    # --- Barra "Condiciones del edificio durante la exposición" ---
    tabla_cond = Table(
        [[Paragraph("<b>CONDICIONES DEL EDIFICIO DURANTE LA EXPOSICIÓN:</b>", estilo_celda_izq)]],
        colWidths=[25.6*cm],
    )
    tabla_cond.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#D9D9D9")),
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(tabla_cond)
    story.append(Spacer(1, 0.3*cm))

    # --- Tabla principal de detectores (cabecera de 2 filas) ---
    headers_fila1 = [
        "Nº", "Código\nDetector", "Edificio", "Planta", "Habitación /\nEstancia",
        "Ubicado en", "Foto", "INSTALACIÓN", "", "", "DESINSTALACIÓN", "", "",
    ]
    headers_fila2 = [
        "", "", "", "", "", "", "", "Fecha", "Hora", "Técnico\nFirma", "Fecha", "Hora", "Técnico\nFirma",
    ]
    filas_tabla = [
        [Paragraph(h.replace("\n", "<br/>"), estilo_cab) for h in headers_fila1],
        [Paragraph(h.replace("\n", "<br/>"), estilo_cab) for h in headers_fila2],
    ]

    n_filas_totales = max(len(detectores) + 3, 10)
    for i in range(n_filas_totales):
        if i < len(detectores):
            d = detectores[i]
            did_d, _, planta_d, sala_d, fecha_d, codigo_d = d[0], d[1], d[2], d[3], d[4], d[5]
            codigo_sala_d, hora_colocacion_d = d[12], d[14]
            fecha_retirada_real_d, hora_retirada_real_d = d[18], d[19]
            tiene_foto = bool(d[9]) or bool(d[10])
            fila = [
                str(i + 1),
                Paragraph(_escapar_pdf(codigo_d or ""), estilo_celda),
                Paragraph(_escapar_pdf(nombre or ""), estilo_celda),
                Paragraph(_escapar_pdf(planta_d or ""), estilo_celda),
                Paragraph(_escapar_pdf(codigo_sala_d or ""), estilo_celda),
                Paragraph(_escapar_pdf(sala_d or ""), estilo_celda),
                "\u2713" if tiene_foto else "",
                Paragraph(_escapar_pdf(fecha_d or ""), estilo_celda),
                Paragraph(_escapar_pdf(hora_colocacion_d or ""), estilo_celda),
                "",
                Paragraph(_escapar_pdf(fecha_retirada_real_d or ""), estilo_celda),
                Paragraph(_escapar_pdf(hora_retirada_real_d or ""), estilo_celda),
                "",
            ]
        else:
            fila = [str(i + 1)] + [""] * 12
        filas_tabla.append(fila)

    col_widths = [0.8*cm, 2.4*cm, 2.4*cm, 1.4*cm, 2.6*cm, 3.4*cm, 1.1*cm,
                  1.9*cm, 1.4*cm, 3.0*cm, 1.9*cm, 1.4*cm, 3.0*cm]
    tabla = Table(filas_tabla, colWidths=col_widths, repeatRows=2)
    estilo_tabla = [
        ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#404040")),
        ("SPAN", (0, 0), (0, 1)), ("SPAN", (1, 0), (1, 1)), ("SPAN", (2, 0), (2, 1)),
        ("SPAN", (3, 0), (3, 1)), ("SPAN", (4, 0), (4, 1)), ("SPAN", (5, 0), (5, 1)),
        ("SPAN", (6, 0), (6, 1)),
        ("SPAN", (7, 0), (9, 0)), ("SPAN", (10, 0), (12, 0)),
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 2), (0, -1), "CENTER"),
        ("ALIGN", (6, 2), (6, -1), "CENTER"),
        ("FONTSIZE", (0, 2), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]
    tabla.setStyle(TableStyle(estilo_tabla))
    story.append(tabla)
    story.append(Spacer(1, 0.6*cm))

    # --- Cuadro de firma del técnico, FUERA de la tabla ---
    marca_digital = "[X]" if tipo_firma == "digital" else "[   ]"
    marca_manual = "[X]" if tipo_firma == "manual" else "[   ]"
    estilo_firma_nombre = ParagraphStyle(
        "FirmaNombre", parent=styles["Normal"], fontSize=11, alignment=TA_CENTER, fontName="Helvetica-Bold",
    )
    estilo_firma_asterisco = ParagraphStyle(
        "FirmaAsterisco", parent=styles["Normal"], fontSize=11, alignment=TA_CENTER,
    )
    estilo_firma_opciones = ParagraphStyle(
        "FirmaOpciones", parent=styles["Normal"], fontSize=10, alignment=TA_CENTER,
    )
    contenido_firma = [
        [Paragraph(_escapar_pdf(tecnico or "Técnico responsable"), estilo_firma_nombre)],
        [Paragraph("*", estilo_firma_asterisco)],
        [Paragraph(f"{marca_digital} Firma digital&nbsp;&nbsp;&nbsp;&nbsp;{marca_manual} Firma a mano",
                    estilo_firma_opciones)],
    ]
    caja_firma = Table(contenido_firma, colWidths=[8*cm])
    caja_firma.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    # Alineada a la derecha de la hoja, como un cuadro de firma independiente.
    tabla_firma_wrap = Table([[Paragraph("", estilo_celda), caja_firma]], colWidths=[17.6*cm, 8*cm])
    tabla_firma_wrap.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
    ]))
    story.append(tabla_firma_wrap)

    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("CF-PE-CYE/39 V.3 (30/05/2025)", ParagraphStyle(
        "CodigoForm", parent=styles["Normal"], fontSize=6.5, textColor=colors.grey, alignment=TA_LEFT,
    )))

    doc.build(story)



# ============================================================
# WIDGETS DE IMAGEN (Subir archivo / Cámara del navegador)
# Sustituyen a los botones "Seleccionar" / "Camara" de la app Tkinter.
# ============================================================

# Solo se permite un flujo de cámara abierto a la vez en TODO el
# formulario (plano, foto de situación, foto del detector, imagen del
# centro). Guarda qué campo "posee" la cámara ahora mismo; el resto no
# monta su propio st.camera_input hasta que ese campo la libere. Esto
# evita el error "no autorizado" que da el móvil cuando varios widgets
# intentan acceder a la cámara al mismo tiempo.
GLOBAL_CAM_OWNER_KEY = "_camara_activa_global"


def widget_imagen(label, state_key, key_prefix, con_camara=True, ancho_miniatura=None,
                   tab_por_defecto="subir", titulo_amarillo=False):
    """Selector de imagen con pestañas 'Subir' y 'Cámara'.

    Guarda automáticamente el archivo elegido en la carpeta de datos y
    recuerda la ruta en st.session_state[state_key]. Devuelve esa ruta
    (o None si no hay imagen). Para quitar una imagen ya subida basta
    con pulsar la "x" del propio archivo en el selector, o subir/sacar
    una nueva foto que la sustituya.

    tab_por_defecto: "subir" (normal) o "camara" (la pestaña de
    cámara aparece primero y por tanto activa por defecto al abrir
    el formulario).
    """
    if titulo_amarillo:
        st.markdown(f'<p class="subtitulo-amarillo">{label}</p>', unsafe_allow_html=True)
    else:
        st.markdown(f"**{label}**")
    file_id_key = key_prefix + "__file_id"
    cam_nonce_key = key_prefix + "__cam_nonce"
    cam_activa_key = key_prefix + "__cam_activa"
    if cam_nonce_key not in st.session_state:
        st.session_state[cam_nonce_key] = 0
    if cam_activa_key not in st.session_state:
        st.session_state[cam_activa_key] = False  # la cámara NO se activa sola

    camara_primero = con_camara and tab_por_defecto == "camara"
    if camara_primero:
        tab_labels = ["📷 Cámara", "📁 Subir"]
        idx_subir, idx_camara = 1, 0
    else:
        tab_labels = ["📁 Subir"] + (["📷 Cámara"] if con_camara else [])
        idx_subir, idx_camara = 0, 1
    tabs = st.tabs(tab_labels)
    nuevo_bytes = None
    nueva_ext = ".jpg"
    vino_de_camara = False

    with tabs[idx_subir]:
        up = st.file_uploader(
            "Selecciona una imagen", type=["png", "jpg", "jpeg"],
            key=key_prefix + "_up", label_visibility="collapsed",
        )
        if up is not None:
            fid = ("up", getattr(up, "file_id", None) or f"{up.name}_{up.size}")
            if st.session_state.get(file_id_key) != fid:
                nuevo_bytes = up.getvalue()
                nueva_ext = extension_de(up)
                st.session_state[file_id_key] = fid

    if con_camara:
        with tabs[idx_camara]:
            hay_imagen = bool(st.session_state.get(state_key))
            # Solo se permite UNA cámara activa a la vez en todo el formulario:
            # si otro campo la tiene abierta, este campo no monta su propio
            # visor (así se evita el error "no autorizado" por streams
            # de cámara simultáneos en el móvil).
            es_el_activo = st.session_state.get(GLOBAL_CAM_OWNER_KEY) == cam_activa_key
            otra_cam_en_uso = (
                st.session_state.get(GLOBAL_CAM_OWNER_KEY) is not None and not es_el_activo
            )

            if st.session_state[cam_activa_key] and es_el_activo:
                cam_key = f"{key_prefix}_cam_{st.session_state[cam_nonce_key]}"
                foto = st.camera_input(
                    "Capturar foto", key=cam_key, label_visibility="collapsed",
                )
                if foto is not None:
                    nuevo_bytes = foto.getvalue()
                    nueva_ext = ".jpg"
                    vino_de_camara = True
            else:
                if hay_imagen:
                    st.caption("Foto capturada (se muestra abajo).")
                etiqueta_btn = "📷 Tomar otra foto" if hay_imagen else "📷 Activar cámara"
                if otra_cam_en_uso:
                    st.caption("La cámara está siendo usada en otro campo. Termina esa foto primero.")
                elif st.button(etiqueta_btn, key=key_prefix + "_activar_cam"):
                    st.session_state[cam_activa_key] = True
                    st.session_state[GLOBAL_CAM_OWNER_KEY] = cam_activa_key
                    st.rerun()

    if nuevo_bytes is not None:
        path = guardar_bytes_imagen(nuevo_bytes, key_prefix, nueva_ext)
        st.session_state[state_key] = path
        # Marca para que quien llame a este widget sepa que ACABA de
        # capturarse/subirse una foto nueva en esta misma interacción
        # (se usa, por ejemplo, para fechar automáticamente la
        # colocación en el momento de sacar la foto del detector).
        st.session_state[key_prefix + "__recien_capturada"] = True
        if vino_de_camara:
            # Ocultamos el visor de la cámara y liberamos el "turno" para
            # que otro campo pueda activarla.
            st.session_state[cam_activa_key] = False
            st.session_state[cam_nonce_key] += 1
            if st.session_state.get(GLOBAL_CAM_OWNER_KEY) == cam_activa_key:
                st.session_state[GLOBAL_CAM_OWNER_KEY] = None
            st.rerun()

    path_actual = st.session_state.get(state_key)
    if path_actual and os.path.exists(path_actual):
        if ancho_miniatura:
            st.image(path_actual, width=ancho_miniatura)
        else:
            st.image(path_actual, use_container_width=True)
    else:
        st.caption("Sin imagen")

    return st.session_state.get(state_key)


def widget_seleccionar_plano_y_punto(cid, ns):
    """Elige uno de los planos ya cargados en el centro (si hay más de
    uno) y marca con un toque la posición del detector sobre ese plano.

    El PLANO se comparte entre varios detectores del mismo centro (se
    gestiona desde la pantalla del centro), pero el PUNTO es propio de
    cada detector: cada uno marca su propia ubicación sobre el plano
    que le corresponda, aunque compartan la misma imagen de fondo.
    """
    px_key = ns + "_plano_px"
    py_key = ns + "_plano_py"
    sel_key = ns + "_plano_centro_id"

    st.markdown('<p class="subtitulo-amarillo">Plano</p>', unsafe_allow_html=True)

    planos = fetch_planos_centro(cid)
    if not planos:
        st.caption(
            "Este centro todavía no tiene ningún plano cargado. "
            "Añade uno desde la pantalla del centro (sección «🗺️ Planos del centro»)."
        )
        st.session_state[sel_key] = None
        return None, None, None

    ids = [p[0] for p in planos]
    nombres = {p[0]: p[2] for p in planos}
    rutas = {p[0]: p[3] for p in planos}

    actual = st.session_state.get(sel_key)
    if actual not in ids:
        # Ni se había elegido antes, ni el que tenía guardado ya no
        # existe (se pudo borrar): se usa el primero como valor por
        # defecto. Si solo hay un plano, este es siempre el usado.
        actual = ids[0]
        st.session_state[sel_key] = actual

    if len(planos) > 1:
        idx_actual = ids.index(actual)
        idx_sel = st.selectbox(
            "Selecciona el plano", options=list(range(len(ids))),
            format_func=lambda i: nombres[ids[i]], index=idx_actual,
            key=ns + "_plano_selector",
        )
        nuevo_id = ids[idx_sel]
        if nuevo_id != actual:
            st.session_state[sel_key] = nuevo_id
            # Al cambiar de plano, el punto anterior ya no tiene sentido
            # (era una posición relativa al plano anterior).
            st.session_state[px_key] = None
            st.session_state[py_key] = None
            st.rerun()
    else:
        st.caption(f"Plano: {nombres[actual]}")

    plano_id_actual = st.session_state[sel_key]
    plano_path = rutas[plano_id_actual]

    if not (plano_path and os.path.exists(plano_path)):
        st.warning("No se pudo abrir la imagen de este plano.")
        return plano_id_actual, st.session_state.get(px_key), st.session_state.get(py_key)

    st.caption(
        "Toca sobre el plano para marcar la ubicación de ESTE detector "
        "(el punto rojo aparecerá en el informe)."
    )
    try:
        img_orig = Image.open(plano_path).convert("RGB")
    except Exception:
        st.warning("No se pudo abrir la imagen del plano.")
        return plano_id_actual, st.session_state.get(px_key), st.session_state.get(py_key)

    # Se redimensiona ANTES de mostrarla para que las coordenadas que
    # devuelve el componente coincidan exactamente con esta imagen.
    ancho_max = 680
    escala = min(1.0, ancho_max / img_orig.width)
    disp_w = max(1, int(img_orig.width * escala))
    disp_h = max(1, int(img_orig.height * escala))
    img_disp = img_orig.resize((disp_w, disp_h), Image.Resampling.LANCZOS)

    px = st.session_state.get(px_key)
    py = st.session_state.get(py_key)
    if px is not None and py is not None:
        draw = ImageDraw.Draw(img_disp)
        cx, cy = px * disp_w, py * disp_h
        r = max(6, int(min(disp_w, disp_h) * 0.02))
        draw.ellipse([cx - r, cy - r, cx + r, cy + r],
                     fill=(220, 20, 20), outline=(120, 0, 0), width=2)

    if IMG_COORD_DISPONIBLE:
        coords = streamlit_image_coordinates(img_disp, key=ns + "_plano_coords_" + str(plano_id_actual))
        if coords is not None:
            nuevo_px = max(0.0, min(1.0, coords["x"] / disp_w))
            nuevo_py = max(0.0, min(1.0, coords["y"] / disp_h))
            if (px, py) != (nuevo_px, nuevo_py):
                st.session_state[px_key] = nuevo_px
                st.session_state[py_key] = nuevo_py
                st.rerun()
    else:
        st.image(img_disp, use_container_width=False)
        st.warning(
            "Para marcar el punto sobre el plano instala el componente:\n\n"
            "`pip install streamlit-image-coordinates`"
        )

    return plano_id_actual, st.session_state.get(px_key), st.session_state.get(py_key)


# ============================================================
# COMPARTIR PDF POR WHATSAPP (Web Share API - Android)
# ============================================================

def boton_compartir_whatsapp_archivo(ruta, nombre_archivo, mime_type, texto_mensaje,
                                      etiqueta_boton, id_sufijo, titulo_compartir):
    """Botón genérico que abre el diálogo nativo de "Compartir" de
    Android con UN archivo (PDF, Excel...) ya adjunto (el usuario elige
    WhatsApp), usando la Web Share API. Requisitos del navegador/dispositivo:
      - Android + Chrome (u otro navegador compatible).
      - La app debe servirse por HTTPS (o localhost). Si usas
        `streamlit run` en tu PC y accedes desde el móvil por IP local
        (http://192.168.x.x:8501) el navegador bloqueará esta función
        por no ser un "contexto seguro": usa un túnel HTTPS (p. ej.
        Cloudflare Tunnel, ngrok) o despliega en Streamlit Community
        Cloud.
    """
    with open(ruta, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")

    texto_js = texto_mensaje.replace("\\", "\\\\").replace("`", "\\`").replace("\n", "\\n")
    nombre_js = nombre_archivo.replace("\\", "\\\\").replace("`", "\\`")
    titulo_js = titulo_compartir.replace("\\", "\\\\").replace("`", "\\`")

    html = f"""
    <div id="wrap-compartir-{id_sufijo}"></div>
    <script>
      const b64_{id_sufijo} = "{b64}";
      const nombreArchivo_{id_sufijo} = `{nombre_js}`;
      const textoMensaje_{id_sufijo} = `{texto_js}`;
      const tituloCompartir_{id_sufijo} = `{titulo_js}`;
      const mimeType_{id_sufijo} = "{mime_type}";

      function b64ToBlob_{id_sufijo}(b64Data, contentType) {{
        const byteChars = atob(b64Data);
        const byteNumbers = new Array(byteChars.length);
        for (let i = 0; i < byteChars.length; i++) {{
          byteNumbers[i] = byteChars.charCodeAt(i);
        }}
        const byteArray = new Uint8Array(byteNumbers);
        return new Blob([byteArray], {{ type: contentType }});
      }}

      // IMPORTANTE: Streamlit ejecuta este código dentro de un <iframe>,
      // y los navegadores bloquean la Web Share API dentro de iframes
      // ("Permission denied") salvo que se invoque desde el documento
      // PRINCIPAL. Por eso el botón real no se pinta aquí dentro, sino
      // que se inyecta en la página principal (fuera del iframe), justo
      // en el hueco que deja este propio iframe.
      const marco_{id_sufijo} = window.frameElement;
      const docPadre_{id_sufijo} = window.parent && window.parent.document;

      function crearUI_{id_sufijo}(doc, contenedorDestino) {{
        // Si de una ejecución anterior ya quedó un botón inyectado
        // (p.ej. al marcar/desmarcar una casilla, que vuelve a
        // ejecutar este script), se quita primero para no duplicarlo.
        const anterior = doc.getElementById("inyectado-{id_sufijo}");
        if (anterior) anterior.remove();

        const cont = doc.createElement("div");
        cont.id = "inyectado-{id_sufijo}";
        cont.style.fontFamily = "'Source Sans Pro', sans-serif";
        const boton = doc.createElement("button");
        boton.textContent = "{etiqueta_boton}";
        boton.style.cssText = "background-color:#25D366; color:white; border:none; " +
          "padding:12px 20px; border-radius:8px; font-size:16px; " +
          "font-weight:600; cursor:pointer; width:100%;";
        const msg = doc.createElement("p");
        msg.style.cssText = "margin-top:8px; font-size:13px; color:#ccc;";
        cont.appendChild(boton);
        cont.appendChild(msg);
        contenedorDestino.appendChild(cont);
        return {{ boton, msg }};
      }}

      let elementos_{id_sufijo};
      if (marco_{id_sufijo} && docPadre_{id_sufijo}) {{
        marco_{id_sufijo}.style.display = "none";
        elementos_{id_sufijo} = crearUI_{id_sufijo}(docPadre_{id_sufijo}, marco_{id_sufijo}.parentNode);
      }} else {{
        elementos_{id_sufijo} = crearUI_{id_sufijo}(document, document.getElementById("wrap-compartir-{id_sufijo}"));
      }}
      const btn_{id_sufijo} = elementos_{id_sufijo}.boton;
      const msg_{id_sufijo} = elementos_{id_sufijo}.msg;
      const navegador_{id_sufijo} = (marco_{id_sufijo} && window.parent) ? window.parent.navigator : navigator;

      btn_{id_sufijo}.addEventListener("click", async () => {{
        try {{
          const blob = b64ToBlob_{id_sufijo}(b64_{id_sufijo}, mimeType_{id_sufijo});
          const file = new File([blob], nombreArchivo_{id_sufijo}, {{ type: mimeType_{id_sufijo} }});

          if (navegador_{id_sufijo}.canShare && navegador_{id_sufijo}.canShare({{ files: [file] }})) {{
            await navegador_{id_sufijo}.share({{
              files: [file],
              title: tituloCompartir_{id_sufijo},
              text: textoMensaje_{id_sufijo},
            }});
          }} else if (navegador_{id_sufijo}.share) {{
            await navegador_{id_sufijo}.share({{ title: tituloCompartir_{id_sufijo}, text: textoMensaje_{id_sufijo} }});
            msg_{id_sufijo}.textContent = "Tu navegador no admite adjuntar el archivo directamente; descárgalo y compártelo manualmente.";
          }} else {{
            msg_{id_sufijo}.textContent = "Tu navegador no soporta compartir archivos. Descárgalo con el botón de arriba y compártelo manualmente desde WhatsApp.";
          }}
        }} catch (err) {{
          if (err.name !== "AbortError") {{
            if (err.name === "NotAllowedError") {{
              msg_{id_sufijo}.textContent = "Tu navegador/dispositivo ha bloqueado el envío directo de este archivo. Descárgalo con el botón de arriba y adjúntalo manualmente en WhatsApp.";
            }} else {{
              msg_{id_sufijo}.textContent = "No se pudo abrir el diálogo de compartir: " + err.message;
            }}
          }}
        }}
      }});
    </script>
    """
    components.html(html, height=110)


def boton_compartir_whatsapp(pdf_path, nombre_archivo, texto_mensaje):
    """Botón que comparte el PDF del informe por WhatsApp."""
    boton_compartir_whatsapp_archivo(
        pdf_path, nombre_archivo, "application/pdf", texto_mensaje,
        "📲 Enviar PDF por WhatsApp", "pdf", "Informe de detectores de Rn",
    )


def boton_compartir_whatsapp_excel(xlsx_path, nombre_archivo, texto_mensaje):
    """Botón que comparte la hoja de cálculo del informe por WhatsApp."""
    boton_compartir_whatsapp_archivo(
        xlsx_path, nombre_archivo,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        texto_mensaje, "📊 Enviar Excel por WhatsApp", "xlsx",
        "Hoja de cálculo de detectores de Rn",
    )


def boton_compartir_whatsapp_fotos(imagenes, texto_mensaje):
    """Botón que comparte por WhatsApp SOLO las fotos marcadas en la
    checklist (sin el PDF), usando la Web Share API con varios archivos
    a la vez. Se separa del PDF a propósito: compartir un PDF mezclado
    con imágenes en el mismo envío falla en bastantes navegadores/apps.

    Cada imagen se comprime/reduce de tamaño antes de codificarla (no
    hace falta la resolución original para verla en WhatsApp), para
    que el conjunto no se vuelva demasiado pesado cuantas más fotos se
    marquen — eso es lo que hacía fallar el envío conjunto cuando
    había muchas fotos de por medio.

    imagenes: lista de tuplas (ruta_en_disco, nombre_de_archivo).
    """
    archivos = []
    for ruta, nombre in imagenes:
        if not ruta or not os.path.exists(ruta):
            continue
        try:
            with Image.open(ruta) as im_orig:
                im = ImageOps.exif_transpose(im_orig)
                im = im.convert("RGB")
                im.thumbnail((1280, 1280))
                buf = io.BytesIO()
                im.save(buf, format="JPEG", quality=78)
                datos = buf.getvalue()
            nombre_final = os.path.splitext(nombre)[0] + ".jpg"
            mime_final = "image/jpeg"
        except Exception:
            # Si por lo que sea no se puede recomprimir, se manda el
            # archivo tal cual estaba en disco.
            with open(ruta, "rb") as f:
                datos = f.read()
            nombre_final = nombre
            ext = os.path.splitext(nombre)[1].lower()
            mime_final = "image/png" if ext == ".png" else "image/jpeg"

        archivos.append({
            "b64": base64.b64encode(datos).decode("utf-8"),
            "name": nombre_final,
            "type": mime_final,
        })

    if not archivos:
        st.caption("Marca al menos una foto arriba para poder enviarla.")
        return

    archivos_json = json.dumps(archivos)
    texto_json = json.dumps(texto_mensaje)
    n_fotos = len(archivos)

    html = f"""
    <div id="wrap-compartir-fotos"></div>
    <script>
      const archivosDataFotos = {archivos_json};
      const textoMensajeFotos = {texto_json};

      function b64ToBlobFotos(b64Data, contentType) {{
        const byteChars = atob(b64Data);
        const byteNumbers = new Array(byteChars.length);
        for (let i = 0; i < byteChars.length; i++) {{
          byteNumbers[i] = byteChars.charCodeAt(i);
        }}
        const byteArray = new Uint8Array(byteNumbers);
        return new Blob([byteArray], {{ type: contentType }});
      }}

      // IMPORTANTE: Streamlit ejecuta este código dentro de un <iframe>,
      // y los navegadores bloquean la Web Share API dentro de iframes
      // ("Permission denied") salvo que se invoque desde el documento
      // PRINCIPAL. Por eso el botón real no se pinta aquí dentro, sino
      // que se inyecta en la página principal (fuera del iframe), justo
      // en el hueco que deja este propio iframe.
      const marcoActual = window.frameElement;
      const docPadre = window.parent && window.parent.document;

      function crearUI(doc, contenedorDestino) {{
        // Si de una ejecución anterior ya quedó un botón inyectado
        // (p.ej. al marcar/desmarcar una casilla, que vuelve a
        // ejecutar este script), se quita primero para no duplicarlo.
        const anterior = doc.getElementById("inyectado-fotos-wa");
        if (anterior) anterior.remove();

        const cont = doc.createElement("div");
        cont.id = "inyectado-fotos-wa";
        cont.style.fontFamily = "'Source Sans Pro', sans-serif";
        const boton = doc.createElement("button");
        boton.textContent = "🖼️ Enviar {n_fotos} foto(s) por WhatsApp";
        boton.style.cssText = "background-color:#25D366; color:white; border:none; " +
          "padding:12px 20px; border-radius:8px; font-size:16px; " +
          "font-weight:600; cursor:pointer; width:100%;";
        const msg = doc.createElement("p");
        msg.style.cssText = "margin-top:8px; font-size:13px; color:#ccc;";
        cont.appendChild(boton);
        cont.appendChild(msg);
        contenedorDestino.appendChild(cont);
        return {{ boton, msg }};
      }}

      let elementos;
      if (marcoActual && docPadre) {{
        // Se oculta el iframe (se queda vacío) y se inserta el botón
        // real justo a continuación, en el documento principal.
        marcoActual.style.display = "none";
        elementos = crearUI(docPadre, marcoActual.parentNode);
      }} else {{
        // Si por algún motivo no se puede acceder al documento padre
        // (navegador muy restrictivo), se deja el botón aquí dentro
        // como respaldo, aunque pueda dar el mismo error de permiso.
        elementos = crearUI(document, document.getElementById("wrap-compartir-fotos"));
      }}
      const btnFotos = elementos.boton;
      const msgFotos = elementos.msg;
      const navegadorParaCompartir = (marcoActual && window.parent) ? window.parent.navigator : navigator;

      btnFotos.addEventListener("click", async () => {{
        try {{
          const files = archivosDataFotos.map(
            a => new File([b64ToBlobFotos(a.b64, a.type)], a.name, {{ type: a.type }})
          );
          if (navegadorParaCompartir.canShare && navegadorParaCompartir.canShare({{ files }})) {{
            await navegadorParaCompartir.share({{
              files: files,
              title: "Fotos del informe de detectores de Rn",
              text: textoMensajeFotos,
            }});
          }} else if (navegadorParaCompartir.share) {{
            await navegadorParaCompartir.share({{ title: "Fotos del informe de detectores de Rn", text: textoMensajeFotos }});
            msgFotos.textContent = "Tu navegador no admite adjuntar varias fotos a la vez.";
          }} else {{
            msgFotos.textContent = "Tu navegador no soporta compartir archivos. Descarga las fotos manualmente.";
          }}
        }} catch (err) {{
          if (err.name !== "AbortError") {{
            msgFotos.textContent = "No se pudo abrir el diálogo de compartir: " + err.message;
          }}
        }}
      }});
    </script>
    """
    components.html(html, height=110)


def _limpiar_namespace(ns):
    borrar = [k for k in st.session_state.keys()
              if k == ns or k.startswith(ns + "_") or k.startswith(ns + "__")]
    for k in borrar:
        del st.session_state[k]
    # Si la cámara global la tenía un campo de este formulario, la
    # liberamos para que no quede "atascada" al salir de la pantalla.
    owner = st.session_state.get(GLOBAL_CAM_OWNER_KEY)
    if owner and (owner == ns or owner.startswith(ns + "_") or owner.startswith(ns + "__")):
        st.session_state[GLOBAL_CAM_OWNER_KEY] = None


def abrir_detector_nuevo():
    ns_previo = st.session_state.get("detector_form_ns")
    if ns_previo:
        _limpiar_namespace(ns_previo)
    st.session_state.detector_actual = None
    st.session_state.view = "detector"


def abrir_detector_editar(detector_id):
    ns_previo = st.session_state.get("detector_form_ns")
    if ns_previo:
        _limpiar_namespace(ns_previo)
    st.session_state.detector_actual = detector_id
    st.session_state.view = "detector"


# ============================================================
# PANTALLA: INICIO (lista de centros)
# ============================================================

def pantalla_inicio():
    st.markdown(
        '<p style="color:#999999; font-size:0.85rem; font-weight:700; margin:0;">'
        'Gestión de muestreo</p>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<p class="titulo-home">☢️ Detectores de Radón</p>',
        unsafe_allow_html=True,
    )

    if "mostrar_form_nuevo_centro" not in st.session_state:
        st.session_state.mostrar_form_nuevo_centro = False
    if "mostrar_form_importar" not in st.session_state:
        st.session_state.mostrar_form_importar = False

    with st.container(border=True):
        bc1, bc2 = st.columns(2)
        with bc1:
            marcadores_nuevo = '<div class="marcador-btn-nuevo-centro"></div>'
            if st.session_state.mostrar_form_nuevo_centro:
                marcadores_nuevo += '<div class="marcador-activo-naranja"></div>'
            st.markdown(marcadores_nuevo, unsafe_allow_html=True)
            if st.button("➕ Nuevo centro", use_container_width=True, type="secondary"):
                abrir = not st.session_state.mostrar_form_nuevo_centro
                st.session_state.mostrar_form_nuevo_centro = abrir
                if abrir:
                    st.session_state.mostrar_form_importar = False
                st.rerun()
        with bc2:
            marcadores_importar = '<div class="marcador-btn-importar"></div>'
            if st.session_state.mostrar_form_importar:
                marcadores_importar += '<div class="marcador-activo-naranja"></div>'
            st.markdown(marcadores_importar, unsafe_allow_html=True)
            if st.button("Importar centro", use_container_width=True, type="secondary",
                         icon=":material/folder_open:"):
                abrir = not st.session_state.mostrar_form_importar
                st.session_state.mostrar_form_importar = abrir
                if abrir:
                    st.session_state.mostrar_form_nuevo_centro = False
                st.rerun()

    if st.session_state.mostrar_form_nuevo_centro:
        with st.form("form_nuevo_centro", clear_on_submit=True):
            nombre_nuevo = st.text_input("Nombre del centro")
            st.markdown('<div class="marcador-btn-crear-centro"></div>', unsafe_allow_html=True)
            crear = st.form_submit_button("Crear centro", type="primary")
        if crear:
            if nombre_nuevo and nombre_nuevo.strip():
                cid = crear_centro(nombre_nuevo.strip())
                st.session_state.centro_actual = cid
                st.session_state.mostrar_form_nuevo_centro = False
                st.session_state.view = "centro"
                st.rerun()
            else:
                st.warning("Escribe un nombre para el centro")

    if st.session_state.mostrar_form_importar:
        st.caption(
            "Reconstruye un centro completo (datos, planos y detectores, con "
            "el punto exacto de cada uno) a partir de un Excel generado por "
            "esta misma app."
        )
        st.markdown('<div class="marcador-uploader-importar"></div>', unsafe_allow_html=True)
        archivo_importar = st.file_uploader(
            "Selecciona el archivo .xlsx", type=["xlsx"], key="importar_centro_file",
        )
        if archivo_importar:
            st.markdown('<div class="marcador-btn-confirmar-importar"></div>', unsafe_allow_html=True)
            if st.button("Importar", key="btn_confirmar_importar", type="primary"):
                try:
                    with st.spinner("Importando centro..."):
                        nuevo_cid, n_detectores = importar_centro_desde_excel(archivo_importar.getvalue())
                    st.session_state.mostrar_form_importar = False
                    st.session_state.centro_actual = nuevo_cid
                    st.session_state.view = "centro"
                    st.success(f"Centro importado correctamente ({n_detectores} detector(es)).")
                    st.rerun()
                except ValueError as e:
                    st.error(str(e))
                except Exception as e:
                    st.error(f"No se pudo importar el archivo: {e}")

    centros = fetch_centros()
    if not centros:
        return

    st.markdown("---")
    st.markdown('<p class="subtitulo-amarillo">Centros registrados</p>', unsafe_allow_html=True)

    opciones = list(range(len(centros)))
    etiquetas = {
        i: f"{c[1] or '(sin nombre)'}"
           + (f" · {c[2]}" if c[2] else "")
        for i, c in enumerate(centros)
    }
    idx_sel = st.selectbox(
        "Centro", options=opciones, format_func=lambda i: etiquetas[i],
        key="selector_centro_home", label_visibility="collapsed",
    )
    cid_sel, nombre_sel, zona_sel, fecha_sel, img_sel = centros[idx_sel]

    b1, b2 = st.columns([3, 1])
    with b1:
        st.markdown('<div class="marcador-btn-abrir-centro"></div>', unsafe_allow_html=True)
        if st.button("Abrir centro", type="primary", use_container_width=True,
                     icon=":material/folder_open:"):
            st.session_state.centro_actual = cid_sel
            st.session_state.view = "centro"
            st.rerun()
    with b2:
        st.markdown('<div class="marcador-btn-eliminar"></div>', unsafe_allow_html=True)
        if st.button("❌ Eliminar", key="btn_eliminar_centro_home"):
            st.session_state["confirmar_borrado_centro"] = cid_sel
            st.rerun()

    if st.session_state.get("confirmar_borrado_centro") == cid_sel:
        st.warning(f"¿Eliminar el centro «{nombre_sel}» y todos sus detectores? Esta acción no se puede deshacer.")
        cc1, cc2 = st.columns(2)
        with cc1:
            if st.button("Sí, eliminar", key=f"confirmar_del_{cid_sel}", type="primary"):
                delete_centro(cid_sel)
                st.session_state["confirmar_borrado_centro"] = None
                st.rerun()
        with cc2:
            if st.button("Cancelar", key=f"cancelar_del_{cid_sel}"):
                st.session_state["confirmar_borrado_centro"] = None
                st.rerun()


# ============================================================
# PANTALLA: CENTRO (datos, detectores, generar PDF)
# ============================================================

def pantalla_centro():
    cid = st.session_state.centro_actual
    centro = get_centro(cid) if cid else None
    if not centro:
        st.error("Centro no encontrado.")
        st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
        if st.button("← Volver"):
            st.session_state.view = "inicio"
            st.rerun()
        return

    cid, nombre, zona, fecha, img_path, tecnico_centro, direccion = centro

    top1, top2 = st.columns([3, 1])
    with top1:
        st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
        if st.button("← Volver a Centros"):
            st.session_state.view = "inicio"
            st.rerun()
    with top2:
        if st.button("Datos de la empresa", use_container_width=True):
            st.session_state.view = "ajustes"
            st.rerun()

    st.markdown(
        '<p style="color:#999999; font-size:0.85rem; font-weight:700; margin:0;">'
        'Gestión de Detectores</p>',
        unsafe_allow_html=True,
    )
    st.markdown(
        f'<p class="titulo-centro">🏢 {html.escape(nombre or "")}</p>',
        unsafe_allow_html=True,
    )

    ns_centro = f"centro_{cid}"
    seccion_key = ns_centro + "_seccion_abierta"
    if seccion_key not in st.session_state:
        st.session_state[seccion_key] = "datos"  # al entrar, solo éste abierto

    def _cabecera_bloque(titulo, seccion_id):
        """Botón-cabecera de un bloque desplegable: al pulsarlo se abre
        ese bloque y se cierra cualquier otro que estuviera abierto
        (solo uno abierto a la vez, tipo acordeón)."""
        abierto = st.session_state[seccion_key] == seccion_id
        flecha = "▾" if abierto else "▸"
        if abierto:
            # El bloque activo lleva el texto en amarillo (mismo tono
            # que el logo), para distinguirlo del resto de cabeceras.
            st.markdown('<div class="marcador-bloque-activo"></div>', unsafe_allow_html=True)
        if st.button(f"{flecha} {titulo}", key=f"toggle_{seccion_id}_{cid}",
                     type="tertiary", use_container_width=True):
            st.session_state[seccion_key] = None if abierto else seccion_id
            st.rerun()
        return abierto

    # Datos que necesitan varias secciones (se calculan una sola vez).
    detectores = fetch_detectores(cid)
    planos_centro = fetch_planos_centro(cid)

    # ============================================================
    # BLOQUE 1: Datos del centro
    # ============================================================
    if _cabecera_bloque("Datos del centro", "datos"):
        with st.container(border=True):
            # OJO: Streamlit borra el valor de un text_input de su
            # session_state en cuanto ese widget deja de renderizarse en
            # una ejecución (p.ej. al cerrar este bloque). Por eso aquí
            # se usa "value=" directamente con el dato de la base de
            # datos cada vez que se abre, en vez de precargarlo solo la
            # primera vez (ese patrón perdía los datos al volver).
            img_key = ns_centro + "_img"
            if img_key not in st.session_state:
                st.session_state[img_key] = img_path

            fecha_valor_original = fecha or datetime.now().strftime("%d/%m/%Y")

            c1, c2 = st.columns(2)
            with c1:
                st.text_input("Nombre", value=nombre or "", key=ns_centro + "_nombre")
                st.text_input("Área", value=zona or "", key=ns_centro + "_zona")
                st.text_input("Dirección", value=direccion or "", key=ns_centro + "_direccion")
                st.text_input("Fecha", value=fecha_valor_original, key=ns_centro + "_fecha")
            with c2:
                # Igual de pequeña que los planos, no a tamaño completo.
                st.markdown('<div class="marcador-imagen-exterior"></div>', unsafe_allow_html=True)
                widget_imagen(
                    "Imagen exterior", img_key, key_prefix=ns_centro + "_imgw",
                    ancho_miniatura=140, titulo_amarillo=True,
                )

            # El botón de guardar solo aparece si hay algún cambio
            # respecto a lo que ya hay guardado; en cuanto guardas,
            # desaparece de nuevo hasta que vuelvas a tocar algo.
            hay_cambios = (
                st.session_state[ns_centro + "_nombre"] != (nombre or "")
                or st.session_state[ns_centro + "_zona"] != (zona or "")
                or st.session_state[ns_centro + "_direccion"] != (direccion or "")
                or st.session_state[ns_centro + "_fecha"] != fecha_valor_original
                or st.session_state.get(img_key) != img_path
            )

            if hay_cambios:
                st.markdown('<div class="marcador-btn-guardar-centro"></div>', unsafe_allow_html=True)
                if st.button("💾 Guardar centro", type="primary"):
                    nombre_in = st.session_state[ns_centro + "_nombre"].strip()
                    if not nombre_in:
                        st.warning("El nombre es obligatorio")
                    else:
                        update_centro(
                            cid, nombre_in,
                            st.session_state[ns_centro + "_zona"].strip(),
                            st.session_state[ns_centro + "_fecha"].strip(),
                            st.session_state.get(img_key),
                            st.session_state[ns_centro + "_direccion"].strip(),
                        )
                        st.success("Centro guardado")
                        st.rerun()

    # ============================================================
    # BLOQUE 2: Categorías profesionales expuestas
    # ============================================================
    if _cabecera_bloque("Categorías profesionales", "categorias"):
        with st.container(border=True):
            # Formulario para añadir, ARRIBA del listado.
            cnew1, cnew2, cnew3 = st.columns([2.2, 1.8, 1.3])
            with cnew1:
                nueva_categoria = st.text_input(
                    "Categoría profesional", key=f"nueva_categoria_nombre_{cid}",
                )
            with cnew2:
                st.markdown('<div class="marcador-num-personas"></div>', unsafe_allow_html=True)
                nuevo_num_personas = st.number_input(
                    "Nº de profesionales", key=f"nueva_categoria_num_{cid}",
                    min_value=0, max_value=999, step=1,
                )
            with cnew3:
                st.markdown('<div class="marcador-btn-anadir-categoria"></div>', unsafe_allow_html=True)
                if st.button("➕ Añadir", key=f"add_categoria_{cid}", type="primary"):
                    if not nueva_categoria.strip():
                        st.warning("Escribe el nombre de la categoría")
                    else:
                        insert_categoria_centro(cid, nueva_categoria.strip(), int(nuevo_num_personas))
                        st.rerun()

            st.markdown("---")

            categorias_centro = fetch_categorias_centro(cid)
            if categorias_centro:
                seleccionadas_cat = []
                for cat_id, _, categoria, num_personas in categorias_centro:
                    marcado = st.checkbox(
                        f"{categoria}: {num_personas}", key=f"chk_cat_{cat_id}",
                    )
                    if marcado:
                        seleccionadas_cat.append(cat_id)

                if seleccionadas_cat:
                    st.markdown('<div class="marcador-btn-eliminar"></div>', unsafe_allow_html=True)
                    if st.button(f"❌ Eliminar seleccionadas ({len(seleccionadas_cat)})",
                                 key=f"del_cats_sel_{cid}"):
                        for cat_id_del in seleccionadas_cat:
                            delete_categoria_centro(cat_id_del)
                        st.rerun()
            else:
                st.caption("Todavía no se ha añadido ninguna categoría profesional.")

    # ============================================================
    # BLOQUE 3: Planos del centro
    # ============================================================
    if _cabecera_bloque("Planos del centro", "planos"):
        with st.container(border=True):
            st.caption(
                "Carga aquí los planos de las plantas del centro. Cada "
                "detector elegirá luego, en su propia pantalla, sobre cuál de "
                "estos planos marcar su ubicación."
            )

            for plano_c in planos_centro:
                plano_id_c, _, nombre_plano_c, ruta_plano_c, _ = plano_c
                with st.container(border=True):
                    pcol1, pcol2, pcol3 = st.columns([1, 2, 1])
                    with pcol1:
                        if ruta_plano_c and os.path.exists(ruta_plano_c):
                            st.image(ruta_plano_c, width=140)
                        else:
                            st.caption("(imagen no encontrada)")
                    with pcol2:
                        st.markdown(f"**{nombre_plano_c}**")
                    with pcol3:
                        st.markdown('<div class="marcador-btn-eliminar"></div>', unsafe_allow_html=True)
                        if st.button("❌ Eliminar", key=f"del_plano_{plano_id_c}"):
                            st.session_state["confirmar_borrado_plano"] = plano_id_c
                            st.rerun()

                if st.session_state.get("confirmar_borrado_plano") == plano_id_c:
                    st.warning(
                        f"¿Eliminar el plano «{nombre_plano_c}»? Los detectores que lo "
                        "tuvieran asignado se quedarán sin plano ni punto marcado."
                    )
                    cc1, cc2 = st.columns(2)
                    with cc1:
                        if st.button("Sí, eliminar", key=f"conf_del_plano_{plano_id_c}", type="primary"):
                            delete_plano_centro(plano_id_c)
                            st.session_state["confirmar_borrado_plano"] = None
                            st.rerun()
                    with cc2:
                        if st.button("Cancelar", key=f"cancel_del_plano_{plano_id_c}"):
                            st.session_state["confirmar_borrado_plano"] = None
                            st.rerun()

            mostrar_add_plano_key = f"mostrar_add_plano_{cid}"
            if mostrar_add_plano_key not in st.session_state:
                st.session_state[mostrar_add_plano_key] = False

            st.markdown('<div class="marcador-btn-plano-amarillo"></div>', unsafe_allow_html=True)
            if st.button("➕ Añadir plano", type="tertiary"):
                st.session_state[mostrar_add_plano_key] = not st.session_state[mostrar_add_plano_key]
                st.rerun()

            if st.session_state[mostrar_add_plano_key]:
                nombre_plano_nuevo = st.text_input(
                    "Nombre del plano (p.ej. «Planta baja», «Planta 1»...)",
                    key=f"nuevo_plano_nombre_{cid}",
                )
                st.markdown('<div class="marcador-btn-plano-amarillo"></div>', unsafe_allow_html=True)
                archivo_plano_nuevo = st.file_uploader(
                    "Selecciona la imagen del plano", type=["png", "jpg", "jpeg"],
                    key=f"nuevo_plano_file_{cid}",
                )
                if archivo_plano_nuevo:
                    st.markdown('<div class="marcador-btn-plano-amarillo"></div>', unsafe_allow_html=True)
                    if st.button("Guardar plano", key=f"guardar_plano_{cid}", type="primary"):
                        if not nombre_plano_nuevo.strip():
                            st.warning("Ponle un nombre al plano")
                        else:
                            ruta_nueva = guardar_bytes_imagen(
                                archivo_plano_nuevo.getvalue(), f"plano_centro_{cid}",
                                extension_de(archivo_plano_nuevo),
                            )
                            insert_plano_centro(cid, nombre_plano_nuevo.strip(), ruta_nueva, len(planos_centro))
                            st.session_state[mostrar_add_plano_key] = False
                            st.success("Plano añadido")
                        st.rerun()

    # ============================================================
    # BLOQUE 4: Detectores colocados
    # ============================================================
    if _cabecera_bloque("Detectores colocados", "detectores"):
        with st.container(border=True):
            st.markdown('<div class="marcador-btn-nuevo-detector"></div>', unsafe_allow_html=True)
            if st.button("➕ Nuevo detector", type="tertiary"):
                abrir_detector_nuevo()
                st.rerun()

            if not detectores:
                st.info("Todavía no se han añadido detectores para este centro.")
            else:
                opciones = list(range(len(detectores)))
                etiquetas = {}
                for i, d in enumerate(detectores):
                    did, _, planta, sala, fecha_det, codigo = d[0], d[1], d[2], d[3], d[4], d[5]
                    partes = [codigo or f"Detector {did}"]
                    if sala:
                        partes.append(sala)
                    if planta:
                        partes.append(f"Planta {planta}")
                    etiquetas[i] = " · ".join(partes)

                idx_sel = st.selectbox(
                    "Detector", options=opciones, format_func=lambda i: etiquetas[i],
                    key=f"selector_detector_{cid}", label_visibility="collapsed",
                )
                d_sel = detectores[idx_sel]
                did_sel = d_sel[0]

                b1, b2 = st.columns([3, 1])
                with b1:
                    st.markdown('<div class="marcador-btn-abrir-detector"></div>', unsafe_allow_html=True)
                    if st.button("📂 Abrir detector", type="primary", use_container_width=True):
                        abrir_detector_editar(did_sel)
                        st.rerun()
                with b2:
                    st.markdown('<div class="marcador-btn-eliminar"></div>', unsafe_allow_html=True)
                    if st.button("❌ Eliminar", key="btn_eliminar_detector_home"):
                        st.session_state["confirmar_borrado_det"] = did_sel
                        st.rerun()

                if st.session_state.get("confirmar_borrado_det") == did_sel:
                    st.warning("¿Eliminar este detector? Esta acción no se puede deshacer.")
                    cc1, cc2 = st.columns(2)
                    with cc1:
                        if st.button("Sí, eliminar", key=f"conf_del_det_{did_sel}", type="primary"):
                            delete_detector(did_sel)
                            st.session_state["confirmar_borrado_det"] = None
                            st.rerun()
                    with cc2:
                        if st.button("Cancelar", key=f"cancel_del_det_{did_sel}"):
                            st.session_state["confirmar_borrado_det"] = None
                            st.rerun()

    # ============================================================
    # BLOQUE (nuevo): Retirada de detectores
    # ============================================================
    if _cabecera_bloque("Retirada de detectores", "retirada"):
        with st.container(border=True):
            if not detectores:
                st.info("Todavía no se han añadido detectores para este centro.")
            else:
                opciones_r = list(range(len(detectores)))
                etiquetas_r = {}
                for i, d in enumerate(detectores):
                    did_i, _, planta_i, sala_i, _, codigo_i = d[0], d[1], d[2], d[3], d[4], d[5]
                    partes_i = [codigo_i or f"Detector {did_i}"]
                    if sala_i:
                        partes_i.append(sala_i)
                    if planta_i:
                        partes_i.append(f"Planta {planta_i}")
                    etiquetas_r[i] = " · ".join(partes_i)

                idx_sel_r = st.selectbox(
                    "Detector a retirar", options=opciones_r, format_func=lambda i: etiquetas_r[i],
                    key=f"selector_retirada_{cid}", label_visibility="collapsed",
                )
                d_r = detectores[idx_sel_r]
                did_r = d_r[0]
                codigo_r = d_r[5]
                sala_r = d_r[3]
                foto_sit_r = d_r[9]
                punto_x_r = d_r[7]
                punto_y_r = d_r[8]
                plano_centro_id_r = d_r[17]  # columna plano_centro_id (ver orden de init_db)
                fecha_retirada_real_r = d_r[18]
                hora_retirada_real_r = d_r[19]

                st.markdown(f"**Código:** {codigo_r or '-'}")
                st.markdown(f"**Sala:** {sala_r or '-'}")

                # Plano de este detector, con su propio punto rojo
                # dibujado encima (no el plano "en blanco").
                if (plano_centro_id_r and punto_x_r is not None and punto_y_r is not None
                        and punto_x_r >= 0 and punto_y_r >= 0):
                    plano_info_r = get_plano_centro(plano_centro_id_r)
                    if plano_info_r and os.path.exists(plano_info_r[3]):
                        ruta_tmp_punto_r = os.path.join(
                            get_data_dir(), f"_tmp_plano_punto_retirada_{did_r}.jpg"
                        )
                        if generar_plano_con_punto(plano_info_r[3], punto_x_r, punto_y_r, ruta_tmp_punto_r):
                            st.image(ruta_tmp_punto_r, use_container_width=True)
                        else:
                            st.caption("No se pudo generar el plano con el punto")
                    else:
                        st.caption("Sin plano asignado")
                else:
                    st.caption("Este detector no tiene plano ni punto marcado")

                if foto_sit_r and os.path.exists(foto_sit_r):
                    st.image(foto_sit_r, use_container_width=True)
                else:
                    st.caption("Sin foto de situación")

                fr_key = f"retirada_fecha_{did_r}"
                hr_key = f"retirada_hora_{did_r}"
                pend_fr_key = f"retirada_fecha_pend_{did_r}"
                pend_hr_key = f"retirada_hora_pend_{did_r}"

                # Igual que con la fecha/hora de colocación: al ser
                # widgets ya creados, su valor solo se puede actualizar
                # mediante un valor "pendiente" aplicado justo ANTES de
                # crearlos (+ rerun), nunca después.
                if pend_fr_key in st.session_state:
                    st.session_state[fr_key] = st.session_state.pop(pend_fr_key)
                if pend_hr_key in st.session_state:
                    st.session_state[hr_key] = st.session_state.pop(pend_hr_key)

                st.text_input("Fecha de retirada real", value=fecha_retirada_real_r or "", key=fr_key)
                st.text_input("Hora de retirada real", value=hora_retirada_real_r or "", key=hr_key)

                rc1, rc2 = st.columns(2)
                with rc1:
                    st.markdown('<div class="marcador-btn-retirada-amarillo"></div>', unsafe_allow_html=True)
                    if st.button("🕒 Capturar fecha y hora", key=f"capturar_retirada_{did_r}"):
                        ahora = datetime.now()
                        st.session_state[pend_fr_key] = ahora.strftime("%d/%m/%Y")
                        st.session_state[pend_hr_key] = ahora.strftime("%H:%M")
                        st.rerun()
                with rc2:
                    st.markdown('<div class="marcador-btn-retirada-amarillo"></div>', unsafe_allow_html=True)
                    if st.button("💾 Guardar", key=f"guardar_retirada_{did_r}", type="primary"):
                        actualizar_retirada_detector(
                            did_r,
                            st.session_state[fr_key].strip(),
                            st.session_state[hr_key].strip(),
                        )
                        st.success("Fecha y hora de retirada guardadas")
                        st.rerun()

    # ============================================================
    # BLOQUE 5: Informes y descargas
    # ============================================================
    if _cabecera_bloque("📄 Informes y descargas", "informes"):
      with st.container(border=True):
        if not detectores:
            st.caption("Añade al menos un detector para poder generar los archivos.")
        else:
            # Recopilar todas las fotos del informe (imagen exterior del
            # centro y plano-con-punto/situación/detector de cada uno).
            # Se usa tanto para el ZIP de descarga como para la checklist
            # de WhatsApp.
            fotos_disponibles = []
            if img_path and os.path.exists(img_path):
                ext = os.path.splitext(img_path)[1] or ".jpg"
                fotos_disponibles.append({
                    "ruta": img_path,
                    "nombre_archivo": f"exterior_{_slug(nombre)}{ext}",
                    "etiqueta": "Foto exterior del centro",
                })
            # Los planos se listan una sola vez cada uno (se comparten
            # entre varios detectores, así que no se repiten).
            for plano_c in planos_centro:
                _, _, nombre_plano_c, ruta_plano_c, _ = plano_c
                if ruta_plano_c and os.path.exists(ruta_plano_c):
                    ext = os.path.splitext(ruta_plano_c)[1] or ".jpg"
                    fotos_disponibles.append({
                        "ruta": ruta_plano_c,
                        "nombre_archivo": f"plano_{_slug(nombre_plano_c)}{ext}",
                        "etiqueta": f"Plano — {nombre_plano_c}",
                    })
            for d in detectores:
                did_d, codigo_d = d[0], d[5]
                foto_sit_d, foto_det_d = d[9], d[10]
                punto_x_d, punto_y_d = d[7], d[8]
                plano_centro_id_d = d[17]  # columna plano_centro_id (ver orden de init_db)
                base_id = codigo_d or f"Detector {did_d}"

                # Plano de ESTE detector, con su propio punto rojo dibujado
                # encima (no la imagen del plano "en blanco").
                if (plano_centro_id_d and punto_x_d is not None and punto_y_d is not None
                        and punto_x_d >= 0 and punto_y_d >= 0):
                    plano_info_d = get_plano_centro(plano_centro_id_d)
                    if plano_info_d and os.path.exists(plano_info_d[3]):
                        ruta_tmp_punto = os.path.join(
                            get_data_dir(), f"_tmp_plano_punto_det{did_d}.jpg"
                        )
                        if generar_plano_con_punto(plano_info_d[3], punto_x_d, punto_y_d, ruta_tmp_punto):
                            fotos_disponibles.append({
                                "ruta": ruta_tmp_punto,
                                "nombre_archivo": f"plano_punto_det{did_d}.jpg",
                                "etiqueta": f"Plano con punto — {base_id}",
                            })

                if foto_sit_d and os.path.exists(foto_sit_d):
                    ext = os.path.splitext(foto_sit_d)[1] or ".jpg"
                    fotos_disponibles.append({
                        "ruta": foto_sit_d, "nombre_archivo": f"situacion_det{did_d}{ext}",
                        "etiqueta": f"Foto situación — {base_id}",
                    })
                if foto_det_d and os.path.exists(foto_det_d):
                    ext = os.path.splitext(foto_det_d)[1] or ".jpg"
                    fotos_disponibles.append({
                        "ruta": foto_det_d, "nombre_archivo": f"detector_det{did_d}{ext}",
                        "etiqueta": f"Foto detector — {base_id}",
                    })

            # --- Checklist de qué documentos generar ---
            st.markdown(
                '<p class="subtitulo-amarillo">¿Qué quieres generar?</p>',
                unsafe_allow_html=True,
            )
            gen_pdf = st.checkbox("Informe PDF", value=True, key=f"gen_chk_pdf_{cid}")
            gen_excel = st.checkbox("Hoja Excel", value=True, key=f"gen_chk_excel_{cid}")
            gen_fotos = st.checkbox("Fotos", value=True, key=f"gen_chk_fotos_{cid}")
            gen_lab = st.checkbox("Registro para laboratorio", value=False, key=f"gen_chk_lab_{cid}")
            tipo_firma_lab = "digital"
            if gen_lab:
                tipo_firma_lab = st.radio(
                    "Tipo de firma del técnico",
                    options=["digital", "manual"],
                    format_func=lambda v: "Firma digital" if v == "digital" else "Firma a mano",
                    key=f"gen_firma_tipo_{cid}", horizontal=True,
                )

            # --- Un único botón que genera lo marcado, pequeño y
            # alineado a la izquierda ---
            col_gen, _col_resto = st.columns([1, 3])
            with col_gen:
                st.markdown('<div class="marcador-btn-generar"></div>', unsafe_allow_html=True)
                generar_clic = st.button(
                    "📦 Generar", type="tertiary", use_container_width=True,
                )

            if generar_clic:
                if not (gen_pdf or gen_excel or gen_fotos or gen_lab):
                    st.warning("Marca al menos un documento para generar.")
                else:
                    try:
                        marca_tiempo = datetime.now().strftime('%Y%m%d_%H%M%S')
                        with st.spinner("Generando documentos..."):
                            if gen_pdf:
                                nombre_pdf = f"informe_{_slug(nombre)}_{marca_tiempo}.pdf"
                                ruta_pdf = os.path.join(get_data_dir(), nombre_pdf)
                                generar_pdf(cid, ruta_pdf)
                                st.session_state["ultimo_pdf"] = ruta_pdf
                                st.session_state["ultimo_pdf_nombre"] = nombre_pdf
                                st.session_state["ultimo_pdf_centro"] = cid

                            if gen_excel:
                                nombre_xlsx = f"detectores_{_slug(nombre)}_{marca_tiempo}.xlsx"
                                ruta_xlsx = os.path.join(get_data_dir(), nombre_xlsx)
                                generar_excel(cid, ruta_xlsx)
                                st.session_state["ultimo_excel"] = ruta_xlsx
                                st.session_state["ultimo_excel_nombre"] = nombre_xlsx
                                st.session_state["ultimo_excel_centro"] = cid

                            if gen_fotos:
                                nombre_zip = f"fotos_{_slug(nombre)}_{marca_tiempo}.zip"
                                ruta_zip = os.path.join(get_data_dir(), nombre_zip)
                                modo_fotos_gen = st.session_state.get(f"modo_fotos_{cid}", "todas")
                                if modo_fotos_gen == "individual":
                                    fotos_a_incluir = [
                                        foto for i, foto in enumerate(fotos_disponibles)
                                        if st.session_state.get(
                                            f"wa_chk_{cid}_{i}_{_slug(foto['nombre_archivo'])}")
                                    ]
                                else:
                                    fotos_a_incluir = fotos_disponibles
                                with zipfile.ZipFile(ruta_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                                    for foto in fotos_a_incluir:
                                        if os.path.exists(foto["ruta"]):
                                            zf.write(foto["ruta"], arcname=foto["nombre_archivo"])
                                st.session_state["ultimo_zip_fotos"] = ruta_zip
                                st.session_state["ultimo_zip_fotos_nombre"] = nombre_zip
                                st.session_state["ultimo_zip_fotos_centro"] = cid

                            if gen_lab:
                                nombre_lab = f"registro_laboratorio_{_slug(nombre)}_{marca_tiempo}.pdf"
                                ruta_lab = os.path.join(get_data_dir(), nombre_lab)
                                generar_registro_laboratorio(cid, ruta_lab, tipo_firma=tipo_firma_lab)
                                st.session_state["ultimo_lab"] = ruta_lab
                                st.session_state["ultimo_lab_nombre"] = nombre_lab
                                st.session_state["ultimo_lab_centro"] = cid

                        st.success("Documentos generados correctamente")
                    except Exception as e:
                        st.error(f"Error al generar los documentos: {e}")

            hay_pdf = st.session_state.get("ultimo_pdf_centro") == cid
            hay_excel = st.session_state.get("ultimo_excel_centro") == cid
            hay_zip = st.session_state.get("ultimo_zip_fotos_centro") == cid
            hay_lab = st.session_state.get("ultimo_lab_centro") == cid

            if hay_pdf or hay_excel or hay_zip or hay_lab:
                ultimo_pdf = st.session_state.get("ultimo_pdf")
                ultimo_excel = st.session_state.get("ultimo_excel")
                ultimo_zip = st.session_state.get("ultimo_zip_fotos")
                ultimo_lab = st.session_state.get("ultimo_lab")

                # 1) Descargar PDF
                if hay_pdf and ultimo_pdf and os.path.exists(ultimo_pdf):
                    with open(ultimo_pdf, "rb") as f:
                        pdf_bytes = f.read()
                    st.markdown('<div class="marcador-btn-descarga-amarillo"></div>', unsafe_allow_html=True)
                    st.download_button(
                        "Descargar PDF", data=pdf_bytes,
                        file_name=st.session_state.get("ultimo_pdf_nombre", "informe.pdf"),
                        mime="application/pdf", use_container_width=True,
                        icon=":material/picture_as_pdf:",
                    )

                # 2) Enviar PDF por WhatsApp
                if hay_pdf and ultimo_pdf and os.path.exists(ultimo_pdf):
                    texto_wa = f"Informe de colocación de detectores de Rn – {nombre or ''}"
                    boton_compartir_whatsapp(
                        ultimo_pdf,
                        st.session_state.get("ultimo_pdf_nombre", "informe.pdf"),
                        texto_wa,
                    )

                # 3) Descargar Excel
                if hay_excel and ultimo_excel and os.path.exists(ultimo_excel):
                    with open(ultimo_excel, "rb") as f:
                        excel_bytes = f.read()
                    st.markdown('<div class="marcador-btn-descarga-amarillo"></div>', unsafe_allow_html=True)
                    st.download_button(
                        "Descargar Excel", data=excel_bytes,
                        file_name=st.session_state.get("ultimo_excel_nombre", "detectores.xlsx"),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        icon=":material/grid_on:",
                    )

                # 4) Enviar Excel por WhatsApp
                if hay_excel and ultimo_excel and os.path.exists(ultimo_excel):
                    texto_wa_excel = f"Hoja de cálculo de detectores de Rn – {nombre or ''}"
                    boton_compartir_whatsapp_excel(
                        ultimo_excel,
                        st.session_state.get("ultimo_excel_nombre", "detectores.xlsx"),
                        texto_wa_excel,
                    )

                # 5) Descargar Registro para laboratorio
                if hay_lab and ultimo_lab and os.path.exists(ultimo_lab):
                    with open(ultimo_lab, "rb") as f:
                        lab_bytes = f.read()
                    st.markdown('<div class="marcador-btn-descarga-amarillo"></div>', unsafe_allow_html=True)
                    st.download_button(
                        "Descargar registro para laboratorio", data=lab_bytes,
                        file_name=st.session_state.get("ultimo_lab_nombre", "registro_laboratorio.pdf"),
                        mime="application/pdf", use_container_width=True,
                        icon=":material/picture_as_pdf:",
                    )

                # 6) Enviar Registro para laboratorio por WhatsApp
                if hay_lab and ultimo_lab and os.path.exists(ultimo_lab):
                    texto_wa_lab = f"Registro para laboratorio – {nombre or ''}"
                    boton_compartir_whatsapp(
                        ultimo_lab,
                        st.session_state.get("ultimo_lab_nombre", "registro_laboratorio.pdf"),
                        texto_wa_lab,
                    )

                # 7) Fotos: por defecto se incluyen todas; el botón
                # "Selección individual" muestra la checklist para
                # elegir solo algunas (y así no alargar la pantalla
                # con el listado completo salvo que se pida).
                seleccionadas = []
                if fotos_disponibles:
                    st.markdown(
                        '<p class="subtitulo-amarillo">📷 Fotos del informe</p>',
                        unsafe_allow_html=True,
                    )

                    modo_key = f"modo_fotos_{cid}"
                    if modo_key not in st.session_state:
                        st.session_state[modo_key] = "todas"
                    modo_actual = st.session_state[modo_key]

                    st.markdown('<div class="marcador-btn-descarga-amarillo"></div>', unsafe_allow_html=True)
                    etiqueta_toggle = (
                        "🔎 Selección individual" if modo_actual == "todas" else "☑️ Seleccionar todos"
                    )
                    if st.button(etiqueta_toggle, key=f"toggle_modo_fotos_{cid}", use_container_width=True):
                        if modo_actual == "todas":
                            st.session_state[modo_key] = "individual"
                        else:
                            st.session_state[modo_key] = "todas"
                            for i, foto in enumerate(fotos_disponibles):
                                st.session_state[
                                    f"wa_chk_{cid}_{i}_{_slug(foto['nombre_archivo'])}"
                                ] = True
                        st.rerun()

                    if st.session_state[modo_key] == "individual":
                        for i, foto in enumerate(fotos_disponibles):
                            marcado = st.checkbox(
                                foto["etiqueta"],
                                key=f"wa_chk_{cid}_{i}_{_slug(foto['nombre_archivo'])}",
                            )
                            if marcado:
                                seleccionadas.append((foto["ruta"], foto["nombre_archivo"]))
                    else:
                        st.caption(
                            f"Se incluirán las {len(fotos_disponibles)} fotos disponibles. "
                            "Pulsa «Selección individual» para elegir solo algunas."
                        )
                        seleccionadas = [(f["ruta"], f["nombre_archivo"]) for f in fotos_disponibles]

                # 8) Descargar fotos
                if hay_zip and ultimo_zip and os.path.exists(ultimo_zip):
                    with open(ultimo_zip, "rb") as f:
                        zip_bytes = f.read()
                    st.markdown('<div class="marcador-btn-descarga-amarillo"></div>', unsafe_allow_html=True)
                    st.download_button(
                        "⬇️ Descargar fotos", data=zip_bytes,
                        file_name=st.session_state.get("ultimo_zip_fotos_nombre", "fotos.zip"),
                        mime="application/zip", use_container_width=True,
                    )

                # 9) Enviar fotos seleccionadas por WhatsApp
                if fotos_disponibles:
                    texto_wa_fotos = f"Fotos del informe – {nombre or ''}"
                    boton_compartir_whatsapp_fotos(seleccionadas, texto_wa_fotos)


# ============================================================
# PANTALLA: DETECTOR (nuevo / editar)
# ============================================================

TURNOS_TRABAJO_OPCIONES = ["Mañana", "Mañana + tarde", "Mañana + tarde + noche", "Rotatorio"]
NIVEL_OPCIONES = ["Bajo rasante", "Planta Baja"]


def pantalla_detector():
    cid = st.session_state.centro_actual
    detector_id = st.session_state.detector_actual
    ns = f"det_{detector_id or 'nuevo'}"
    st.session_state["detector_form_ns"] = ns

    init_key = ns + "__cargado"
    if not st.session_state.get(init_key):
        if detector_id:
            d = get_detector(detector_id)
            (did, _, planta, sala, fecha, codigo, _plano_antiguo, px, py, foto_sit, foto_det, _,
             codigo_sala, profesionales_sala, hora_colocacion, turno_trabajo, nivel, plano_centro_id, fecha_retirada_real, hora_retirada_real) = d
        else:
            centro = get_centro(cid)
            planta = sala = codigo = codigo_sala = profesionales_sala = hora_colocacion = turno_trabajo = nivel = ""
            fecha = centro[3] if centro else datetime.now().strftime("%d/%m/%Y")
            foto_sit = foto_det = None
            plano_centro_id = None
            fecha_retirada_real = hora_retirada_real = ""
            px = py = None

        st.session_state[ns + "_planta"] = planta or ""
        st.session_state[ns + "_sala"] = sala or ""
        st.session_state[ns + "_fecha"] = fecha or datetime.now().strftime("%d/%m/%Y")
        st.session_state[ns + "_codigo"] = codigo or ""
        st.session_state[ns + "_codigo_sala"] = codigo_sala or ""
        st.session_state[ns + "_profesionales_sala"] = profesionales_sala or ""
        st.session_state[ns + "_hora_colocacion"] = hora_colocacion or ""
        st.session_state[ns + "_fecha_retirada_real"] = fecha_retirada_real or ""
        st.session_state[ns + "_hora_retirada_real"] = hora_retirada_real or ""
        st.session_state[ns + "_turno_trabajo"] = (
            turno_trabajo if turno_trabajo in TURNOS_TRABAJO_OPCIONES else TURNOS_TRABAJO_OPCIONES[0]
        )
        st.session_state[ns + "_nivel"] = (
            nivel if nivel in NIVEL_OPCIONES else NIVEL_OPCIONES[0]
        )
        st.session_state[ns + "_plano_centro_id"] = plano_centro_id
        st.session_state[ns + "_plano_px"] = px if (px is not None and px >= 0) else None
        st.session_state[ns + "_plano_py"] = py if (py is not None and py >= 0) else None
        st.session_state[ns + "_foto_sit"] = foto_sit
        st.session_state[ns + "_foto_det"] = foto_det
        st.session_state[init_key] = True

    # Si en la última interacción se acaba de capturar/subir la foto del
    # detector, se fecha y hora de colocación automáticamente con el
    # momento actual. La "Fecha" es un widget ya creado en pantallas
    # anteriores, así que su valor se actualiza mediante un valor
    # "pendiente" + rerun (no se puede reasignar un session_state de un
    # widget ya instanciado en la misma ejecución).
    if st.session_state.get(ns + "_fecha_pendiente") is not None:
        st.session_state[ns + "_fecha"] = st.session_state.pop(ns + "_fecha_pendiente")

    st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
    if st.button("← Volver al centro"):
        _limpiar_namespace(ns)
        st.session_state.view = "centro"
        st.rerun()

    texto_titulo_detector = "Datos del detector" if detector_id else "Nuevo detector"
    st.markdown(
        f'<p class="titulo-home">{texto_titulo_detector}</p>',
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns(2)
    with c1:
        st.text_input("Código del detector", key=ns + "_codigo")
        st.text_input("Planta", key=ns + "_planta")
        st.text_input("Sala", key=ns + "_sala")
        st.text_input("Profesionales en la sala", key=ns + "_profesionales_sala")
    with c2:
        st.text_input("Fecha de colocación", key=ns + "_fecha")
        st.selectbox("Nivel", options=NIVEL_OPCIONES, key=ns + "_nivel")
        st.text_input("Código de la sala", key=ns + "_codigo_sala")
        st.selectbox("Turnos de trabajo", options=TURNOS_TRABAJO_OPCIONES, key=ns + "_turno_trabajo")

    st.markdown("---")
    widget_seleccionar_plano_y_punto(cid, ns)

    st.markdown("---")
    st.markdown('<p class="subtitulo-amarillo">Fotos detector</p>', unsafe_allow_html=True)
    fc1, fc2 = st.columns(2)
    with fc1:
        st.markdown('<div class="marcador-imagen-exterior"></div>', unsafe_allow_html=True)
        widget_imagen("Situación del detector", ns + "_foto_sit", key_prefix=ns + "_fsit",
                      tab_por_defecto="camara", titulo_amarillo=True)
    with fc2:
        st.markdown('<div class="marcador-imagen-exterior"></div>', unsafe_allow_html=True)
        widget_imagen("Detector", ns + "_foto_det", key_prefix=ns + "_fdet",
                      tab_por_defecto="camara", titulo_amarillo=True)

    # Al capturar/subir la foto del detector, se fecha y hora la
    # colocación automáticamente con el momento exacto (en vez de
    # tener que escribirla a mano).
    if st.session_state.get(ns + "_fdet__recien_capturada"):
        st.session_state[ns + "_fdet__recien_capturada"] = False
        ahora = datetime.now()
        st.session_state[ns + "_fecha_pendiente"] = ahora.strftime("%d/%m/%Y")
        st.session_state[ns + "_hora_colocacion"] = ahora.strftime("%H:%M")
        st.rerun()

    hora_coloc_actual = st.session_state.get(ns + "_hora_colocacion")
    if hora_coloc_actual:
        st.caption(f"🕒 Hora de colocación registrada: {hora_coloc_actual} (se toma al hacer la foto del detector)")

    st.markdown("---")
    st.markdown('<div class="marcador-btn-guardar-detector"></div>', unsafe_allow_html=True)
    guardar = st.button("💾 Guardar detector", type="primary", use_container_width=True)

    if guardar:
        sala_val = st.session_state[ns + "_sala"].strip()
        codigo_val = st.session_state[ns + "_codigo"].strip()
        if not sala_val:
            st.warning("La sala es obligatoria")
        elif not codigo_val:
            st.warning("El código es obligatorio")
        else:
            px = st.session_state.get(ns + "_plano_px")
            py = st.session_state.get(ns + "_plano_py")
            data = (
                cid,
                st.session_state[ns + "_planta"].strip(),
                sala_val,
                st.session_state[ns + "_fecha"].strip(),
                codigo_val,
                None,
                px if px is not None else -1,
                py if py is not None else -1,
                st.session_state.get(ns + "_foto_sit"),
                st.session_state.get(ns + "_foto_det"),
                datetime.now().strftime("%Y-%m-%d %H:%M"),
                st.session_state[ns + "_codigo_sala"].strip(),
                st.session_state[ns + "_profesionales_sala"].strip(),
                st.session_state.get(ns + "_hora_colocacion", "").strip(),
                st.session_state.get(ns + "_turno_trabajo", TURNOS_TRABAJO_OPCIONES[0]),
                st.session_state.get(ns + "_nivel", NIVEL_OPCIONES[0]),
                st.session_state.get(ns + "_plano_centro_id"),
                st.session_state.get(ns + "_fecha_retirada_real", ""),
                st.session_state.get(ns + "_hora_retirada_real", ""),
            )
            if detector_id:
                update_detector(detector_id, data)
            else:
                insert_detector(data)

            _limpiar_namespace(ns)
            st.session_state.view = "centro"
            st.success("Detector guardado")
            st.rerun()


# ============================================================
# PANTALLA: AJUSTES
# ============================================================

def pantalla_ajustes():
    cid = st.session_state.get("centro_actual")
    centro = get_centro(cid) if cid else None

    st.markdown('<p class="titulo-home">Datos de la empresa</p>', unsafe_allow_html=True)
    st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
    if st.button("← Volver"):
        st.session_state.view = "centro" if cid else "inicio"
        st.rerun()

    if not centro:
        st.error("No hay ningún centro seleccionado.")
        return

    _, nombre_centro, _, _, _, tecnico_actual, _ = centro
    st.caption(f"Centro: {nombre_centro}")

    empresa_actual = get_empresa()
    cif_actual = get_cif()

    st.text_input(
        "Empresa",
        value=empresa_actual,
        key="ajustes_empresa",
    )
    st.text_input(
        "CIF",
        value=cif_actual,
        key="ajustes_cif",
    )
    st.text_input(
        "Técnico (aparece en el PDF de este centro)",
        value=tecnico_actual or "",
        key="ajustes_tecnico",
    )

    hay_cambios_ajustes = (
        st.session_state["ajustes_empresa"] != (empresa_actual or "")
        or st.session_state["ajustes_cif"] != (cif_actual or "")
        or st.session_state["ajustes_tecnico"] != (tecnico_actual or "")
    )

    if hay_cambios_ajustes:
        st.markdown('<div class="marcador-btn-guardar-ajustes"></div>', unsafe_allow_html=True)
        if st.button("Guardar", type="primary"):
            set_empresa(st.session_state["ajustes_empresa"].strip())
            set_cif(st.session_state["ajustes_cif"].strip())
            set_tecnico_centro(cid, st.session_state["ajustes_tecnico"].strip())
            st.session_state.view = "centro"
            st.rerun()


# ============================================================
# TRADUCCIÓN DE TEXTOS FIJOS EN INGLÉS (cámara, subida de archivos...)
# ============================================================
# Streamlit no permite traducir directamente los textos internos de
# algunos widgets nativos (p.ej. "Take Photo" del selector de cámara,
# o "Drag and drop file here" del subidor de archivos). Se inyecta un
# pequeño script que busca esos textos en la página y los sustituye por
# su equivalente en castellano, y los vuelve a aplicar cada vez que
# Streamlit repinta la interfaz. Si el navegador bloquea el acceso al
# documento padre (restricción de origen cruzado), simplemente no se
# traduce y el texto se queda en inglés, sin romper nada más.

def inyectar_traduccion_widgets():
    html = """
    <script>
    (function() {
      const traducciones = [
        ["Take Photo", "Sacar foto"],
        ["Clear photo", "Quitar foto"],
        ["Switch camera", "Cambiar cámara"],
        ["Drag and drop file here", "Arrastra el archivo aquí"],
        ["Drag and drop files here", "Arrastra los archivos aquí"],
        ["Browse files", "Buscar archivo"],
        ["Upload", "Subir archivo"],
        ["200MB per file", "200MB max."],
      ];

      function traducirNodo(nodo) {
        if (nodo.nodeType === Node.TEXT_NODE) {
          const original = nodo.nodeValue;
          for (const [en, es] of traducciones) {
            if (original.includes(en)) {
              nodo.nodeValue = original.split(en).join(es);
            }
          }
        } else if (nodo.nodeType === Node.ELEMENT_NODE) {
          nodo.childNodes.forEach(traducirNodo);
        }
      }

      try {
        const doc = window.parent.document;
        traducirNodo(doc.body);
        const observer = new MutationObserver(function() {
          traducirNodo(doc.body);
        });
        observer.observe(doc.body, { childList: true, subtree: true, characterData: true });
      } catch (e) {
        // Si el navegador no permite acceder al documento padre, se
        // deja tal cual (seguirá en inglés en ese caso concreto).
      }
    })();
    </script>
    """
    components.html(html, height=0)


# ============================================================
# MAIN
# ============================================================

def main():
    init_db()
    inyectar_traduccion_widgets()

    if "view" not in st.session_state:
        st.session_state.view = "inicio"
    if "centro_actual" not in st.session_state:
        st.session_state.centro_actual = None
    if "detector_actual" not in st.session_state:
        st.session_state.detector_actual = None

    view = st.session_state.view
    if view == "inicio":
        pantalla_inicio()
    elif view == "centro":
        pantalla_centro()
    elif view == "detector":
        pantalla_detector()
    elif view == "ajustes":
        pantalla_ajustes()
    else:
        st.session_state.view = "inicio"
        st.rerun()


if __name__ == "__main__":
    main()
